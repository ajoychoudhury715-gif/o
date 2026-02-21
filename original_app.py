# pyright: reportMissingImports=false, reportMissingModuleSource=false, reportUnknownVariableType=false, reportUnknownArgumentType=false, reportUnknownParameterType=false, reportUnknownMemberType=false, reportGeneralTypeIssues=false
import streamlit as st  # pyright: ignore[reportUndefinedVariable]
import pandas as pd # pyright: ignore[reportMissingModuleSource]
from datetime import datetime, time as time_type, timezone, timedelta
from typing import Any, Optional
from pandas import DataFrame
import os
import time as time_module  # for retry delays
import zipfile  # for BadZipFile exception handling
from pathlib import Path
# Add missing import
import hashlib
import re  # for creating safe keys for buttons
import uuid  # for generating stable row IDs
import json
import io
import html
import textwrap
import openpyxl
from openpyxl.utils import get_column_letter
try:
    # Altair was previously used for a status dashboard chart.
    # Kept as a try-block placeholder to avoid breaking older deployments that
    # may still have altair installed, but the app no longer requires it.
    pass
except Exception:
    pass
# Supabase integration (Postgres) for persistent cloud storage (no Google)
_supabase_available = False
try:
    from supabase import create_client  # type: ignore
    _supabase_available = True
except Exception:
    pass
SUPABASE_AVAILABLE = _supabase_available
# To install required packages, run in your terminal:
# pip install --upgrade pip
# pip install pandas openpyxl streamlit supabase
# Page config
st.set_page_config(page_title="THE DENTAL BOND", layout="wide", initial_sidebar_state="expanded")
# Session defaults for role/user (replace with real auth later)
if "user_role" not in st.session_state:
    st.session_state.user_role = "admin"
if "current_user" not in st.session_state:
    st.session_state.current_user = "admin"
if "nav_category" not in st.session_state:
    st.session_state.nav_category = "Scheduling"
if "nav_sched" not in st.session_state:
    st.session_state.nav_sched = "Full Schedule"
# ================ GLOBAL CONSTANTS (MOVED TO TOP TO FIX BUGS) ================
# Indian Standard Time (IST = UTC+5:30)
IST = timezone(timedelta(hours=5, minutes=30))
# File paths
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Putt Allotment.xlsx")
# Color schemes
# Medical Blue & White Theme with Glassmorphism
LIGHT_COLORS = {
    "bg_primary": "#f8fafc",        # Soft blue-tinted white
    "bg_secondary": "#f1f5f9",      # Light blue-gray
    "bg_card": "rgba(255, 255, 255, 0.7)",  # Translucent white for glass effect
    "text_primary": "#1e293b",      # Dark slate for text
    "text_secondary": "#64748b",    # Medium slate gray
    "button_bg": "#2563eb",         # Medical blue (primary actions)
    "button_text": "#ffffff",       # White text on buttons
    "accent": "#3b82f6",            # Lighter blue
    "accent_primary": "#2563eb",    # Medical blue (primary actions)
    "accent_secondary": "#3b82f6",  # Lighter blue
    "accent_gradient_start": "#3b82f6",
    "accent_gradient_end": "#2563eb",
    "success": "#10b981",           # Green
    "warning": "#f59e0b",           # Amber
    "danger": "#ef4444",            # Red
    "info": "#0ea5e9",              # Sky blue
    "glass_bg": "rgba(255, 255, 255, 0.25)",  # Glass background
    "glass_border": "rgba(255, 255, 255, 0.18)", # Glass border
    "glass_shadow": "0 8px 32px 0 rgba(37, 99, 235, 0.15)", # Blue-tinted shadow
    "backdrop_blur": "blur(16px)",  # Blur strength
}
DARK_COLORS = {
    "bg_primary": "#0f172a",        # Dark slate
    "bg_secondary": "#1e293b",      # Lighter dark slate
    "bg_card": "rgba(30, 41, 59, 0.7)",
    "text_primary": "#f1f5f9",
    "text_secondary": "#94a3b8",
    "button_bg": "#3b82f6",
    "button_text": "#ffffff",
    "accent": "#60a5fa",
    "accent_primary": "#3b82f6",
    "accent_secondary": "#60a5fa",
    "accent_gradient_start": "#60a5fa",
    "accent_gradient_end": "#3b82f6",
    "success": "#34d399",
    "warning": "#fbbf24",
    "danger": "#f87171",
    "info": "#38bdf8",
    "glass_bg": "rgba(30, 41, 59, 0.4)",
    "glass_border": "rgba(148, 163, 184, 0.18)",
    "glass_shadow": "0 8px 32px 0 rgba(59, 130, 246, 0.2)",
    "backdrop_blur": "blur(16px)",
}
COLORS = LIGHT_COLORS
# Weekly off configuration
WEEKLY_OFF: dict[int, list[str]] = {
    0: ["RAJA"],                          # Monday
    1: ["PRAMOTH", "ANYA"],              # Tuesday
    2: ["ANSHIKA", "MUKHILA"],           # Wednesday
    3: ["RESHMA", "LAVANYA"],            # Thursday
    4: ["ROHINI"],                        # Friday
    5: [],                                 # Saturday (no offs)
    6: ["NITIN", "BABU"],                # Sunday
}
# Profile and storage configuration
PROFILE_SUPABASE_TABLE = "profiles"
PROFILE_ASSISTANT_SHEET = "Assistants"
PROFILE_DOCTOR_SHEET = "Doctors"
PROFILE_COLUMNS = [
    "id",
    "name",
    "kind",
    "department",
    "contact_email",
    "contact_phone",
    "status",
    "weekly_off",
    "pref_first",
    "pref_second",
    "pref_third",
    "created_at",
    "updated_at",
    "created_by",
    "updated_by",
]
# Storage configuration
# USE_SUPABASE: If True and credentials are available, uses Supabase Postgres for cloud sync
# Otherwise falls back to local Excel file (Putt Allotment.xlsx)
USE_SUPABASE = True
# Attendance configuration
ATTENDANCE_SHEET = "Assistants_Attendance"
ATTENDANCE_COLUMNS = ["DATE", "ASSISTANT", "PUNCH IN", "PUNCH OUT"]
# Duties and Patients configuration (Excel sheets)
DUTIES_MASTER_SHEET = "Duties_Master"
DUTY_ASSIGNMENTS_SHEET = "Duty_Assignments"
DUTY_RUNS_SHEET = "Duty_Runs"
PATIENTS_SHEET = "Patients"
# Supabase configuration
supabase_client = None
supabase_table_name = "tdb_allotment_state"
supabase_row_id = "main"
SUPABASE_CHECK_TTL_SECONDS = 60
# Session state initialization for profiles
if "profiles_cache_bust" not in st.session_state:
    st.session_state.profiles_cache_bust = 0
# ================ UTILITY FUNCTIONS ================
def now_ist():
    """Get current datetime in IST timezone."""
    return datetime.now(IST)
def time_to_minutes(time_value: Any) -> Optional[int]:
    """Convert time values to minutes since midnight for comparison."""
    from datetime import time as time_type
    # Handle None
    if time_value is None:
        return None
    # Handle time objects
    if isinstance(time_value, time_type):
        return time_value.hour * 60 + time_value.minute
    # Handle strings
    if isinstance(time_value, str):
        time_value = time_value.strip()
        if not time_value:
            return None
        try:
            # Try parsing HH:MM format
            parts = time_value.split(":")
            if len(parts) == 2:
                h, m = int(parts[0]), int(parts[1])
                if 0 <= h < 24 and 0 <= m < 60:
                    return h * 60 + m
        except Exception:
            pass
    # Handle numbers (assume minutes or decimal hours)
    try:
        val = float(time_value)
        if 0 <= val < 24:  # Assume decimal hours
            hours = int(val)
            minutes = int((val - hours) * 60)
            return hours * 60 + minutes
        elif 0 <= val < 1440:  # Assume minutes
            return int(val)
    except Exception:
        pass
    return None
def _get_profiles_cache_snapshot() -> dict[str, Any]:
    cached = st.session_state.get("profiles_cache")
    return cached if isinstance(cached, dict) else {}
# -----------------------------
# Premium sidebar CSS (white pastel)
# -----------------------------
def inject_white_pastel_sidebar():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&display=swap');
    /* Glassmorphism Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, rgba(248, 250, 252, 0.95) 0%, rgba(241, 245, 249, 0.95) 100%);
        backdrop-filter: blur(20px);
        -webkit-backdrop-filter: blur(20px);
        border-right: 1px solid rgba(37, 99, 235, 0.1);
        box-shadow: 4px 0 24px rgba(37, 99, 235, 0.08);
    }
    [data-testid="stSidebarContent"] { padding: 20px 16px; }
    [data-testid="stSidebar"] .stSelectbox, [data-testid="stSidebar"] .stRadio {
        background: rgba(255, 255, 255, 0.6);
        backdrop-filter: blur(10px);
        border: 1px solid rgba(37, 99, 235, 0.2);
        border-radius: 12px;
        padding: 10px 12px 8px 12px;
        box-shadow: 0 4px 16px rgba(37, 99, 235, 0.1);
        transition: all 0.3s ease;
    }
    [data-testid="stSidebar"] button {
        border-radius: 12px !important;
        padding: 0.7rem 1rem !important;
        border: 1px solid rgba(255, 255, 255, 0.3) !important;
        background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%) !important;
        box-shadow: 0 4px 20px rgba(37, 99, 235, 0.3) !important;
        backdrop-filter: blur(10px) !important;
        font-weight: 600;
        color: white !important;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
    }
    [data-testid="stSidebar"] button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 30px rgba(37, 99, 235, 0.4) !important;
        background: linear-gradient(135deg, #60a5fa 0%, #3b82f6 100%) !important;
    }
    .sidebar-title {
        font-size: 24px;
        font-weight: 800;
        background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        margin-bottom: 20px;
        font-family: 'Inter', sans-serif;
    }
    .live-pill {
        display: inline-flex;
        align-items: center;
        gap: 8px;
        padding: 8px 16px;
        border-radius: 20px;
        font-size: 13px;
        font-weight: 600;
        background: linear-gradient(135deg, rgba(16, 185, 129, 0.2), rgba(5, 150, 105, 0.2));
        border: 1px solid rgba(16, 185, 129, 0.3);
        backdrop-filter: blur(10px);
        color: #059669;
        margin-bottom: 12px;
    }
    .live-dot {
        width: 10px;
        height: 10px;
        border-radius: 50%;
        background: linear-gradient(135deg, #10b981, #059669);
        box-shadow: 0 0 10px rgba(16, 185, 129, 0.6);
        animation: pulse 2s ease-in-out infinite;
    }
    @keyframes pulse {
        0%, 100% { opacity: 1; }
        50% { opacity: 0.6; }
    }
    </style>
    """, unsafe_allow_html=True)
inject_white_pastel_sidebar()
# ================= ASSISTANTS ATTENDANCE TAB =================
ATTENDANCE_SHEET = "Assistants_Attendance"
ATTENDANCE_COLUMNS = ["DATE", "ASSISTANT", "PUNCH IN", "PUNCH OUT"]
def _attendance_excel_path(path_override: Optional[str] = None) -> str:
    """Return a safe attendance Excel path (defaults to local workbook)."""
    if path_override:
        return path_override
    try:
        return file_path
    except NameError:
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), "Putt Allotment.xlsx")
# Use only one definition for safe_str_to_time_obj, and ensure it is robust
def safe_str_to_time_obj(s):
    """Convert HH:MM string to time object. Returns None if invalid."""
    if not s or not isinstance(s, str):
        return None
    try:
        parts = s.strip().split(":")
        if len(parts) != 2:
            return None
        h, m = int(parts[0]), int(parts[1])
        if 0 <= h < 24 and 0 <= m < 60:
            return time_type(hour=h, minute=m)
        return None
    except Exception:
        return None
def safe_time_to_minutes(t):
    if t is None:
        return None
    return t.hour * 60 + t.minute
def ist_today_and_time():
    now = datetime.now(IST)
    return now.date().isoformat(), now.strftime("%H:%M:%S")
# Fix None handling in calc_worked_minutes
def calc_worked_minutes(in_t, out_t, now_t):
    in_min = safe_time_to_minutes(in_t)
    now_min = safe_time_to_minutes(now_t)
    if in_min is None:
        return None, "ABSENT"
    if out_t is None:
        if now_min is None:
            return None, "PARTIAL"
        worked = now_min - in_min
        if worked < 0:
            worked += 1440
        return worked, "PARTIAL"
    out_min = safe_time_to_minutes(out_t)
    if out_min is None:
        return None, "PARTIAL"
    worked = out_min - in_min
    if worked < 0:
        worked += 1440
    return worked, "PRESENT"
def mins_to_hhmm(m):
    if m is None:
        return ""
    h = m // 60
    mm = m % 60
    return f"{h:02d}:{mm:02d}"
def _normalize_html(block: str) -> str:
    return "\n".join(
        line.strip()
        for line in textwrap.dedent(block).splitlines()
        if line.strip()
    )
@st.cache_data(ttl=30)
def _get_active_assistant_profile_names() -> list[str]:
    try:
        df = load_profiles(PROFILE_ASSISTANT_SHEET)
    except Exception:
        return []
    if df is None or df.empty or "name" not in df.columns:
        return []
    names = df["name"].astype(str).str.strip().str.upper()
    if "status" in df.columns:
        status = df["status"].astype(str).str.upper()
        names = names[status == "ACTIVE"]
    out = [n for n in names.tolist() if n]
    seen = set()
    deduped: list[str] = []
    for name in out:
        if name in seen:
            continue
        seen.add(name)
        deduped.append(name)
    return deduped
def get_assistants_list(schedule_df):
    try:
        profiles = _get_active_assistant_profile_names()
        if profiles:
            return profiles
    except Exception:
        pass
    if schedule_df is None or schedule_df.empty:
        return []
    cols = [c for c in ["FIRST", "SECOND", "Third"] if c in schedule_df.columns]
    names = set()
    for c in cols:
        names.update([x.strip() for x in schedule_df[c].dropna().astype(str).tolist() if x.strip()])
    return sorted(names)
def extract_assistants(schedule_df):
    return get_assistants_list(schedule_df)
def mark_busy_assistants(schedule_df: pd.DataFrame) -> pd.DataFrame:
    """Mark assistants as busy (🔴 BUSY) if they have an active duty.
    Modifies FIRST, SECOND, Third columns to show busy status for assistants
    with IN_PROGRESS duties.
    """
    if schedule_df is None or schedule_df.empty:
        return schedule_df
    df_copy = schedule_df.copy()
    # Get all unique assistants from schedule
    assistants = extract_assistants(df_copy)
    # Check which assistants have active duties
    busy_assistants = set()
    for assistant in assistants:
        active_duty = fetch_active_duty_run(None, assistant)
        if active_duty:
            busy_assistants.add(assistant)
    # Mark busy assistants in the schedule
    for col in ["FIRST", "SECOND", "Third"]:
        if col in df_copy.columns:
            for idx in df_copy.index:
                cell_value = str(df_copy.loc[idx, col]).strip()
                if cell_value and cell_value in busy_assistants:
                    df_copy.loc[idx, col] = f"🔴 {cell_value}"
    return df_copy
def ensure_attendance_sheet_exists(excel_path: Optional[str] = None):
    """Create/align the attendance sheet with expected columns."""
    path = Path(_attendance_excel_path(excel_path))
    try:
        if not path.exists():
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                pd.DataFrame(columns=ATTENDANCE_COLUMNS).to_excel(writer, sheet_name=ATTENDANCE_SHEET, index=False)
            return
        xls = pd.ExcelFile(path, engine="openpyxl")
        if ATTENDANCE_SHEET not in xls.sheet_names:
            with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="overlay") as writer:
                pd.DataFrame(columns=ATTENDANCE_COLUMNS).to_excel(writer, sheet_name=ATTENDANCE_SHEET, index=False)
            return
        current = pd.read_excel(xls, sheet_name=ATTENDANCE_SHEET)
        if list(current.columns) != ATTENDANCE_COLUMNS:
            aligned = pd.DataFrame(columns=ATTENDANCE_COLUMNS)
            if not current.empty:
                for col in ATTENDANCE_COLUMNS:
                    if col in current.columns:
                        aligned[col] = current[col]
            # Use safe sheet saving to preserve all other sheets
            save_excel_sheet(aligned, ATTENDANCE_SHEET, str(path))
    except Exception:
        # Non-fatal alignment failure; callers will handle empty frame
        pass
def load_attendance_sheet(excel_path: Optional[str] = None):
    """Load attendance from Supabase or Excel sheet."""
    if USE_SUPABASE and supabase_client:
        df = _sb_load("assistant_attendance")
        if df.empty:
            df = pd.DataFrame(columns=ATTENDANCE_COLUMNS)
        return df if not df.empty else pd.DataFrame(columns=ATTENDANCE_COLUMNS)
    ensure_attendance_sheet_exists(excel_path)
    path = _attendance_excel_path(excel_path)
    try:
        df = pd.read_excel(path, sheet_name=ATTENDANCE_SHEET, engine="openpyxl")
        if df.empty:
            df = pd.DataFrame(columns=ATTENDANCE_COLUMNS)
    except Exception:
        return pd.DataFrame(columns=ATTENDANCE_COLUMNS)
    for col in ATTENDANCE_COLUMNS:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].astype(str).replace("nan", "").fillna("")
    return df[ATTENDANCE_COLUMNS]
def save_attendance_sheet(excel_path: Optional[str], att_df: pd.DataFrame):
    """Save attendance to Supabase or Excel sheet."""
    if USE_SUPABASE and supabase_client:
        _sb_upsert(att_df, "assistant_attendance")
        return
    ensure_attendance_sheet_exists(excel_path)
    path = _attendance_excel_path(excel_path)
    try:
        clean_df = att_df.copy()
        for col in ATTENDANCE_COLUMNS:
            if col not in clean_df.columns:
                clean_df[col] = ""
        clean_df = clean_df[ATTENDANCE_COLUMNS]
        # Use safe sheet saving to preserve all other sheets
        save_excel_sheet(clean_df, ATTENDANCE_SHEET, path)
        try:
            _load_attendance_today_excel.clear()
        except Exception:
            pass
    except Exception as e:
        st.error(f"Attendance save failed: {e}")
# ==================== GENERIC EXCEL SHEET HELPERS ====================
def load_excel_sheet(sheet_name: str, expected_columns: Optional[list] = None, excel_path: Optional[str] = None) -> pd.DataFrame:
    """Load any Excel sheet, return empty DataFrame with expected columns if sheet missing."""
    path = excel_path or file_path
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
        if df.empty and expected_columns:
            return pd.DataFrame(columns=expected_columns)
        if expected_columns:
            for col in expected_columns:
                if col not in df.columns:
                    df[col] = ""
        return df
    except Exception:
        return pd.DataFrame(columns=expected_columns) if expected_columns else pd.DataFrame()
def save_excel_sheet(df: pd.DataFrame, sheet_name: str, excel_path: Optional[str] = None):
    """Save DataFrame to any Excel sheet (preserves all other sheets)."""
    import os
    path = excel_path or file_path
    try:
        # CRITICAL: Only create new workbook if file doesn't exist
        # If file exists but can't be loaded, this is a serious error
        if not os.path.exists(path):
            # File doesn't exist, create new workbook
            wb = openpyxl.Workbook()
            if wb.active:
                wb.remove(wb.active)
        else:
            # File exists, MUST load it successfully
            try:
                wb = openpyxl.load_workbook(path)
            except Exception as e:
                # File exists but can't be loaded - DON'T create new one!
                # This protects against data loss
                print(f"CRITICAL: Cannot load existing Excel file {path}: {e}")
                print(f"Skipping save to prevent data loss of sheet: {sheet_name}")
                return
        # Remove sheet if it exists (to avoid duplicates)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        # Create new sheet and write data using openpyxl directly
        ws = wb.create_sheet(sheet_name)
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # Write data rows
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        # Ensure at least one sheet is visible
        if not any(ws_check.sheet_state == 'visible' for ws_check in wb.worksheets):
            if wb.sheetnames:
                wb[wb.sheetnames[0]].sheet_state = 'visible'
        wb.save(path)
        wb.close()
    except Exception as e:
        print(f"Error saving {sheet_name}: {e}")
# ==================== SUPABASE HELPERS ====================
def _sb_load(table: str, columns: str = "*") -> pd.DataFrame:
    """Load all rows from a Supabase table into a DataFrame."""
    try:
        resp = supabase_client.table(table).select(columns).execute()
        return pd.DataFrame(resp.data or [])
    except Exception as e:
        print(f"Supabase load error [{table}]: {e}")
        return pd.DataFrame()
def _sb_upsert(df: pd.DataFrame, table: str):
    """Upsert a DataFrame to a Supabase table."""
    try:
        records = df.where(pd.notna(df), None).to_dict("records")
        if records:
            supabase_client.table(table).upsert(records).execute()
    except Exception as e:
        print(f"Supabase upsert error [{table}]: {e}")
# ==================== DUTIES MASTER SHEET FUNCTIONS ====================
def load_duties_master_sheet(excel_path: Optional[str] = None) -> pd.DataFrame:
    """Load duties master from Supabase or Excel sheet."""
    columns = ["id", "title", "frequency", "default_minutes", "op", "active", "created_at"]
    if USE_SUPABASE and supabase_client:
        df = _sb_load("duties_master")
        return df if not df.empty else pd.DataFrame(columns=columns)
    return load_excel_sheet(DUTIES_MASTER_SHEET, columns, excel_path)
def save_duties_master_sheet(df: pd.DataFrame, excel_path: Optional[str] = None):
    """Save duties master to Supabase or Excel sheet."""
    if USE_SUPABASE and supabase_client:
        _sb_upsert(df, "duties_master")
        return
    save_excel_sheet(df, DUTIES_MASTER_SHEET, excel_path)
# ==================== DUTY ASSIGNMENTS SHEET FUNCTIONS ====================
def load_duty_assignments_sheet(excel_path: Optional[str] = None) -> pd.DataFrame:
    """Load duty assignments from Supabase or Excel sheet."""
    columns = ["id", "duty_id", "assistant", "op", "est_minutes", "active"]
    if USE_SUPABASE and supabase_client:
        df = _sb_load("duty_assignments")
        return df if not df.empty else pd.DataFrame(columns=columns)
    return load_excel_sheet(DUTY_ASSIGNMENTS_SHEET, columns, excel_path)
def save_duty_assignments_sheet(df: pd.DataFrame, excel_path: Optional[str] = None):
    """Save duty assignments to Supabase or Excel sheet."""
    if USE_SUPABASE and supabase_client:
        _sb_upsert(df, "duty_assignments")
        return
    save_excel_sheet(df, DUTY_ASSIGNMENTS_SHEET, excel_path)
# ==================== DUTY RUNS SHEET FUNCTIONS ====================
def load_duty_runs_sheet(excel_path: Optional[str] = None) -> pd.DataFrame:
    """Load duty runs from Supabase or Excel sheet."""
    columns = ["id", "date", "assistant", "duty_id", "status", "started_at", "due_at", "ended_at", "est_minutes", "op"]
    if USE_SUPABASE and supabase_client:
        df = _sb_load("duty_runs")
        return df if not df.empty else pd.DataFrame(columns=columns)
    return load_excel_sheet(DUTY_RUNS_SHEET, columns, excel_path)
def save_duty_runs_sheet(df: pd.DataFrame, excel_path: Optional[str] = None):
    """Save duty runs to Supabase or Excel sheet."""
    if USE_SUPABASE and supabase_client:
        _sb_upsert(df, "duty_runs")
        return
    save_excel_sheet(df, DUTY_RUNS_SHEET, excel_path)
# ==================== PATIENTS SHEET FUNCTIONS ====================
def load_patients_sheet(excel_path: Optional[str] = None) -> pd.DataFrame:
    """Load patients from Supabase or Excel sheet."""
    columns = ["id", "name"]
    if USE_SUPABASE and supabase_client:
        df = _sb_load("patients")
        return df if not df.empty else pd.DataFrame(columns=columns)
    return load_excel_sheet(PATIENTS_SHEET, columns, excel_path)
def save_patients_sheet(df: pd.DataFrame, excel_path: Optional[str] = None):
    """Save patients to Supabase or Excel sheet."""
    if USE_SUPABASE and supabase_client:
        _sb_upsert(df, "patients")
        return
    save_excel_sheet(df, PATIENTS_SHEET, excel_path)
# ==================== EXCEL-BASED DUTY FUNCTIONS ====================
def fetch_active_duty_assignments_excel(assistant: str) -> list[dict[str, Any]]:
    """Fetch active duty assignments from Excel (replaces Supabase version)."""
    if not assistant:
        return []
    try:
        assignments_df = load_duty_assignments_sheet()
        duties_df = load_duties_master_sheet()
        if assignments_df.empty or duties_df.empty:
            return []
        # Filter assignments by assistant and active=True
        matching = assignments_df[
            (assignments_df["assistant"].astype(str).str.strip() == assistant) &
            (assignments_df["active"].astype(str).str.lower() == "true")
        ]
        if matching.empty:
            return []
        # Create lookup of duties
        duty_map = {}
        for _, duty_row in duties_df.iterrows():
            duty_id = str(duty_row.get("id", ""))
            if duty_row.get("active") and duty_id:
                duty_map[duty_id] = {
                    "id": duty_id,
                    "title": str(duty_row.get("title", "")),
                    "frequency": str(duty_row.get("frequency", "")).upper(),
                    "default_minutes": _safe_int(duty_row.get("default_minutes"), 15),
                    "op": str(duty_row.get("op", "")),
                }
        # Combine assignments with duty details
        result = []
        for _, assign_row in matching.iterrows():
            duty_id = str(assign_row.get("duty_id", ""))
            duty = duty_map.get(duty_id)
            if not duty:
                continue
            est = assign_row.get("est_minutes")
            if est is None or (isinstance(est, float) and pd.isna(est)):
                est = duty.get("default_minutes", 15)
            else:
                est = _safe_int(est, 15)
            result.append({
                "assignment_id": str(assign_row.get("id", "")),
                "duty_id": duty_id,
                "title": duty["title"],
                "frequency": duty["frequency"],
                "est_minutes": est,
                "op": assign_row.get("op") or duty["op"],
            })
        return result
    except Exception as e:
        print(f"Error fetching duty assignments: {e}")
        return []
def fetch_duty_runs_since_excel(assistant: str, start_date_iso: str) -> list[dict[str, Any]]:
    """Fetch duty runs since date from Excel (replaces Supabase version)."""
    if not assistant or not start_date_iso:
        return []
    try:
        runs_df = load_duty_runs_sheet()
        if runs_df.empty:
            return []
        # Filter by assistant and date >= start_date
        matching = runs_df[
            (runs_df["assistant"].astype(str).str.strip() == assistant) &
            (runs_df["date"].astype(str) >= start_date_iso)
        ]
        return matching.to_dict("records") if not matching.empty else []
    except Exception as e:
        print(f"Error fetching duty runs: {e}")
        return []
def fetch_active_duty_run_excel(assistant: str) -> Optional[dict[str, Any]]:
    """Fetch active (IN_PROGRESS) duty run from Excel (replaces Supabase version)."""
    if not assistant:
        return None
    try:
        runs_df = load_duty_runs_sheet()
        if runs_df.empty:
            return None
        # Filter by assistant and status=IN_PROGRESS, sorted by started_at descending
        matching = runs_df[
            (runs_df["assistant"].astype(str).str.strip() == assistant) &
            (runs_df["status"].astype(str).str.upper() == "IN_PROGRESS")
        ]
        if matching.empty:
            return None
        # Sort by started_at descending and get first
        matching_sorted = matching.sort_values("started_at", ascending=False)
        return matching_sorted.iloc[0].to_dict()
    except Exception:
        return None
def start_duty_run_excel(assistant: str, duty: dict[str, Any]) -> tuple[Optional[str], dict[str, Any]]:
    """Start a duty run in Excel (replaces Supabase version)."""
    now_dt = now_ist()
    est = _safe_int(duty.get("est_minutes"), 15)
    due_dt = now_dt + timedelta(minutes=est)
    payload = {
        "id": str(uuid.uuid4()),
        "date": now_dt.date().isoformat(),
        "assistant": assistant,
        "duty_id": duty.get("duty_id"),
        "status": "IN_PROGRESS",
        "started_at": now_dt.isoformat(),
        "due_at": due_dt.isoformat(),
        "op": duty.get("op"),
        "est_minutes": est,
    }
    try:
        runs_df = load_duty_runs_sheet()
        new_row = pd.DataFrame([payload])
        combined = pd.concat([runs_df, new_row], ignore_index=True)
        save_duty_runs_sheet(combined)
        return payload["id"], payload
    except Exception as e:
        print(f"Failed to start duty: {e}")
        return None, payload
def mark_duty_done_excel(run_id: str) -> bool:
    """Mark duty run as DONE in Excel (replaces Supabase version)."""
    if not run_id:
        return False
    try:
        runs_df = load_duty_runs_sheet()
        if runs_df.empty:
            return False
        # Update the row with matching id
        runs_df.loc[runs_df["id"] == run_id, "status"] = "DONE"
        runs_df.loc[runs_df["id"] == run_id, "ended_at"] = now_ist().isoformat()
        save_duty_runs_sheet(runs_df)
        return True
    except Exception as e:
        print(f"Failed to mark duty done: {e}")
        return False
def search_patients_excel(query: str, limit: int = 50) -> list[dict[str, str]]:
    """Search patients from Excel sheet (replaces Supabase version)."""
    q = (query or "").strip().lower()
    if not q:
        return []
    try:
        patients_df = load_patients_sheet()
        if patients_df.empty:
            return []
        # Filter by id or name containing query
        matching = patients_df[
            (patients_df["id"].astype(str).str.lower().str.contains(q, na=False)) |
            (patients_df["name"].astype(str).str.lower().str.contains(q, na=False))
        ].head(limit)
        return matching.to_dict("records") if not matching.empty else []
    except Exception as e:
        print(f"Error searching patients: {e}")
        return []
@st.cache_data(ttl=5)
def db_get_one_attendance(_supabase, date_str: str, assistant: str):
    try:
        res = (
            _supabase.table("assistant_attendance")
            .select("date,assistant,punch_in,punch_out")
            .eq("date", date_str)
            .eq("assistant", assistant)
            .limit(1)
            .execute()
        )
        return res.data[0] if res.data else None
    except Exception as e:
        st.warning(f"Attendance fetch failed: {e}")
        return None
@st.cache_data(ttl=5)
def _load_attendance_today_supabase(_supabase, date_str: str) -> list[dict[str, Any]]:
    try:
        res = (
            _supabase.table("assistant_attendance")
            .select("assistant,punch_in,punch_out")
            .eq("date", date_str)
            .execute()
        )
        return res.data or []
    except Exception:
        return []
@st.cache_data(ttl=30)
def _load_attendance_range_supabase(
    _supabase,
    start_date: str,
    end_date: str,
) -> list[dict[str, Any]]:
    if not _supabase or not start_date or not end_date:
        return []
    try:
        res = (
            _supabase.table("assistant_attendance")
            .select("date,assistant,punch_in,punch_out")
            .gte("date", start_date)
            .lte("date", end_date)
            .execute()
        )
        return res.data or []
    except Exception as e:
        st.warning(f"Attendance fetch failed: {e}")
        return []
@st.cache_data(ttl=5)
def _load_attendance_today_excel(_excel_path: Optional[str], date_str: str) -> list[dict[str, Any]]:
    try:
        att_df = load_attendance_sheet(_excel_path)
        if att_df is None or att_df.empty:
            return []
        day_df = att_df[att_df["DATE"] == date_str]
        if day_df.empty:
            return []
        return day_df.to_dict(orient="records")
    except Exception:
        return []
def _build_punch_map_from_records(records: list[dict[str, Any]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for rec in records or []:
        if not isinstance(rec, dict):
            continue
        name = str(rec.get("assistant") or rec.get("ASSISTANT") or "").strip().upper()
        if not name:
            continue
        punch_in = str(rec.get("punch_in") or rec.get("PUNCH IN") or "").strip()
        punch_out = str(rec.get("punch_out") or rec.get("PUNCH OUT") or "").strip()
        out[name] = {"punch_in": punch_in, "punch_out": punch_out}
    return out
def _get_today_punch_map() -> dict[str, dict[str, str]]:
    date_str, _ = ist_today_and_time()
    if USE_SUPABASE and supabase_client is not None:
        records = _load_attendance_today_supabase(supabase_client, date_str)
        return _build_punch_map_from_records(records)
    records = _load_attendance_today_excel(None, date_str)
    return _build_punch_map_from_records(records)
def _format_punch_time(val: str) -> str:
    s = str(val or "").strip()
    if len(s) >= 5:
        return s[:5]
    return s
def _assistant_punch_state(
    assistant_upper: str,
    punch_map: Optional[dict[str, dict[str, str]]],
) -> tuple[str, str, str]:
    if not punch_map:
        return "NONE", "", ""
    rec = punch_map.get(assistant_upper)
    if not rec:
        return "NONE", "", ""
    punch_in = str(rec.get("punch_in", "") or "").strip()
    punch_out = str(rec.get("punch_out", "") or "").strip()
    if punch_in and not punch_out:
        return "IN", punch_in, ""
    if punch_in and punch_out:
        return "OUT", punch_in, punch_out
    return "NONE", punch_in, punch_out
def _calc_worked_minutes(punch_in: str, punch_out: str) -> Optional[int]:
    in_min = time_to_minutes(punch_in)
    out_min = time_to_minutes(punch_out)
    if in_min is None or out_min is None:
        return None
    worked = out_min - in_min
    if worked < 0:
        worked += 1440
    return worked
def _attendance_status(punch_in: str, punch_out: str) -> str:
    if punch_in and punch_out:
        return "COMPLETE"
    if punch_in and not punch_out:
        return "IN PROGRESS"
    return "MISSING"
def db_punch_in(supabase, date_str: str, assistant: str, now_time: str):
    try:
        payload = {"date": date_str, "assistant": assistant, "punch_in": now_time}
        supabase.table("assistant_attendance").upsert(payload).execute()
        try:
            db_get_one_attendance.clear()
        except Exception:
            pass
        try:
            _load_attendance_today_supabase.clear()
        except Exception:
            pass
    except Exception as e:
        st.error(f"Punch in failed: {e}")
def db_punch_out(supabase, date_str: str, assistant: str, now_time: str):
    try:
        supabase.table("assistant_attendance").update({"punch_out": now_time}).eq("date", date_str).eq("assistant", assistant).execute()
        try:
            db_get_one_attendance.clear()
        except Exception:
            pass
        try:
            _load_attendance_today_supabase.clear()
        except Exception:
            pass
    except Exception as e:
        st.error(f"Punch out failed: {e}")
def sidebar_punch_widget(schedule_df: pd.DataFrame, excel_path: Optional[str] = None):
    today = datetime.now(IST).date().isoformat()
    now_hhmm = datetime.now(IST).strftime("%I:%M %p")
    att = load_attendance_sheet(excel_path)
    # Load assistants from Assistants sheet (not from schedule)
    try:
        assistants_df = load_profiles("Assistants")
        assistants = sorted(assistants_df["name"].dropna().astype(str).str.strip().unique().tolist()) if not assistants_df.empty else []
    except Exception:
        assistants = []
    st.markdown("### 👇 Punch System")
    if not assistants:
        st.caption("No assistants found in Assistants sheet. Add assistants first.")
        return
    assistant = st.selectbox("Select Assistant", assistants, key="sb_assistant")
    mask = (att["DATE"] == today) & (att["ASSISTANT"] == assistant)
    row = att[mask].head(1)
    punch_in = row["PUNCH IN"].iloc[0] if not row.empty else ""
    punch_out = row["PUNCH OUT"].iloc[0] if not row.empty else ""
    if punch_in and not punch_out:
        st.success(f"Status: PUNCHED IN at {punch_in}")
    elif punch_in and punch_out:
        st.info(f"Status: COMPLETED • In {punch_in} • Out {punch_out}")
    else:
        st.warning("Status: Not punched in")
    c1, c2 = st.columns(2)
    with c1:
        disabled_in = bool(punch_in)
        if st.button("✅ Punch In", use_container_width=True, disabled=disabled_in):
            if row.empty:
                new_row = pd.DataFrame([{
                    "DATE": today,
                    "ASSISTANT": assistant,
                    "PUNCH IN": now_hhmm,
                    "PUNCH OUT": "",
                }])
                att = pd.concat([att, new_row], ignore_index=True)
            else:
                att.loc[mask, "PUNCH IN"] = now_hhmm
            save_attendance_sheet(excel_path, att)
            st.toast(f"{assistant} punched in at {now_hhmm}", icon="✅")
            st.rerun()
    with c2:
        disabled_out = (not punch_in) or bool(punch_out)
        if st.button("⏹ Punch Out", use_container_width=True, disabled=disabled_out):
            att.loc[mask, "PUNCH OUT"] = now_hhmm
            save_attendance_sheet(excel_path, att)
            st.toast(f"{assistant} punched out at {now_hhmm}", icon="⏹")
            updated_df = _remove_assistant_assignments(schedule_df, assistant)
            if updated_df is not None:
                _maybe_save(updated_df, message=f"{assistant} removed from allotment after punch out")
            st.rerun()
    with st.expander("Admin actions"):
        if st.button("♻️ Reset today for this assistant", use_container_width=True):
            att = att[~mask].copy()
            save_attendance_sheet(excel_path, att)
            st.toast("Reset done", icon="♻️")
            st.rerun()
def sidebar_punch_widget_supabase(schedule_df: pd.DataFrame, supabase):
    date_str, now_time = ist_today_and_time()
    now_hhmm = now_time[:5]
    st.markdown("### 👇 Punch System")
    # Load assistants from Assistants sheet (not from schedule)
    try:
        assistants_df = load_profiles("Assistants")
        assistants = sorted(assistants_df["name"].dropna().astype(str).str.strip().unique().tolist()) if not assistants_df.empty else []
    except Exception:
        assistants = []
    if not assistants:
        st.caption("No assistants found in Assistants sheet. Add assistants first.")
        return
    assistant = st.selectbox("Select Assistant", assistants, key="sb_assistant")
    rec = db_get_one_attendance(supabase, date_str, assistant)
    punch_in = (rec.get("punch_in") if rec else None) or ""
    punch_out = (rec.get("punch_out") if rec else None) or ""
    if punch_in and not punch_out:
        st.success(f"Status: PUNCHED IN at {str(punch_in)[:5]}")
    elif punch_in and punch_out:
        st.info(f"Status: COMPLETED • In {str(punch_in)[:5]} • Out {str(punch_out)[:5]}")
    else:
        st.warning("Status: Not punched in")
    c1, c2 = st.columns(2)
    with c1:
        if st.button("✅ Punch In", use_container_width=True, disabled=bool(punch_in)):
            db_punch_in(supabase, date_str, assistant, now_time)
            st.toast(f"{assistant} punched in at {now_hhmm}", icon="✅")
            st.rerun()
    with c2:
        if st.button("⏹ Punch Out", use_container_width=True, disabled=(not punch_in) or bool(punch_out)):
            db_punch_out(supabase, date_str, assistant, now_time)
            st.toast(f"{assistant} punched out at {now_hhmm}", icon="⏹")
            updated_df = _remove_assistant_assignments(schedule_df, assistant)
            if updated_df is not None:
                _maybe_save(updated_df, message=f"{assistant} removed from allotment after punch out")
            st.rerun()
# ================= DUTY REMINDER (SUPABASE) =================
@st.cache_data(ttl=5)
def fetch_active_duty_assignments(_supabase, assistant: str) -> list[dict[str, Any]]:
    # Using Excel backend (Supabase disabled)
    return fetch_active_duty_assignments_excel(assistant)
@st.cache_data(ttl=5)
def fetch_duty_runs_since(_supabase, assistant: str, start_date_iso: str):
    # Using Excel backend (Supabase disabled)
    return fetch_duty_runs_since_excel(assistant, start_date_iso)
@st.cache_data(ttl=5)
def fetch_active_duty_run(_supabase, assistant: str):
    # Using Excel backend (Supabase disabled)
    return fetch_active_duty_run_excel(assistant)
def compute_pending_duties(assignments: list[dict[str, Any]], runs: list[dict[str, Any]], today_date) -> dict[str, list[dict[str, Any]]]:
    week_start = today_date - timedelta(days=today_date.weekday())
    month_start = today_date.replace(day=1)
    done_week: set = set()
    done_month: set = set()
    for r in runs:
        if str(r.get("status", "")).upper() != "DONE":
            continue
        r_date = _date_from_any(r.get("date"))
        if r_date is None:
            continue
        if r_date >= week_start:
            done_week.add(r.get("duty_id"))
        if r_date >= month_start:
            done_month.add(r.get("duty_id"))
    pending = {"WEEKLY": [], "MONTHLY": []}
    for a in assignments:
        freq = str(a.get("frequency", "")).upper()
        if freq == "WEEKLY" and a.get("duty_id") not in done_week:
            pending["WEEKLY"].append(a)
        elif freq == "MONTHLY" and a.get("duty_id") not in done_month:
            pending["MONTHLY"].append(a)
    return pending
def start_duty_run_supabase(supabase, assistant: str, duty: dict[str, Any]):
    # Using Excel backend (Supabase disabled)
    return start_duty_run_excel(assistant, duty)
def mark_duty_done_supabase(supabase, run_id: str):
    # Using Excel backend (Supabase disabled)
    return mark_duty_done_excel(run_id)
def compute_free_minutes_for_assistant(schedule_df: pd.DataFrame, assistant: str) -> Optional[int]:
    if schedule_df is None or schedule_df.empty or not assistant:
        return None
    assistant_upper = str(assistant).strip().upper()
    now_dt = now_ist()
    now_min = now_dt.hour * 60 + now_dt.minute
    def _assigned(row) -> bool:
        for col in ["FIRST", "SECOND", "Third"]:
            if col in row and str(row.get(col, "")).strip().upper() == assistant_upper:
                return True
        return False
    def _minutes(val):
        try:
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return None
            return int(val)
        except Exception:
            try:
                return time_to_minutes(val)
            except Exception:
                return None
    next_in = None
    for _, row in schedule_df.iterrows():
        if not _assigned(row):
            continue
        status = str(row.get("STATUS", "")).strip().upper()
        if any(s in status for s in ["CANCELLED", "DONE", "COMPLETED", "SHIFTED"]):
            continue
        in_min = _minutes(row.get("In_min"))
        out_min = _minutes(row.get("Out_min"))
        if in_min is None:
            in_min = _minutes(row.get("In Time")) or _minutes(row.get("In Time Str"))
        if out_min is None:
            out_min = _minutes(row.get("Out Time")) or _minutes(row.get("Out Time Str"))
        if in_min is None:
            continue
        if out_min is None:
            out_min = in_min
        if out_min < in_min:
            out_min += 1440
        if in_min <= now_min <= out_min:
            return 0
        if in_min > now_min:
            if next_in is None or in_min < next_in:
                next_in = in_min
    if next_in is None:
        return 999
    return max(0, next_in - now_min)
def render_duty_reminder_widget(schedule_df: pd.DataFrame, supabase):
    st.markdown("### 🧭 Duties")
    # Duties are now stored in Excel (supabase parameter kept for compatibility but not used)
    # Load assistants from Assistants sheet (not from schedule)
    try:
        assistants_df = load_profiles("Assistants")
        assistants = sorted(assistants_df["name"].dropna().astype(str).str.strip().unique().tolist()) if not assistants_df.empty else []
    except Exception:
        assistants = []
    if not assistants:
        st.caption("No assistants found in Assistants sheet. Add assistants first.")
        return
    default_idx = 0
    try:
        if st.session_state.get("duty_current_assistant") in assistants:
            default_idx = assistants.index(st.session_state.get("duty_current_assistant"))
    except Exception:
        default_idx = 0
    selected_assistant = st.selectbox(
        "Assistant (for this device)",
        options=assistants,
        index=default_idx if default_idx < len(assistants) else 0,
        key="duty_assistant_select",
    )
    st.session_state.duty_current_assistant = selected_assistant
    if not selected_assistant:
        st.caption("Select an assistant to view duties.")
        return
    # Sync active run from Supabase (server truth)
    active_run = fetch_active_duty_run(supabase, selected_assistant)
    if active_run and st.session_state.get("active_duty_run_id") != active_run.get("id"):
        st.session_state.active_duty_run_id = active_run.get("id")
        st.session_state.active_duty_due_at = active_run.get("due_at")
        st.session_state.active_duty_started_at = active_run.get("started_at")
        st.session_state.active_duty_est_minutes = active_run.get("est_minutes")
    if not active_run and st.session_state.get("duty_current_assistant") == selected_assistant and st.session_state.get("active_duty_run_id"):
        # Clear stale local state if server shows none
        for k in ["active_duty_run_id", "active_duty_due_at", "active_duty_started_at", "active_duty_est_minutes"]:
            st.session_state[k] = None
    active_run_id = st.session_state.get("active_duty_run_id")
    if active_run_id:
        due_dt = _parse_iso_ts(st.session_state.get("active_duty_due_at"))
        started_dt = _parse_iso_ts(st.session_state.get("active_duty_started_at"))
        remaining_msg = ""
        if due_dt:
            delta = due_dt - now_ist()
            if delta.total_seconds() > 0:
                mins = int(delta.total_seconds() // 60)
                secs = int(delta.total_seconds() % 60)
                remaining_msg = f"{mins:02d}:{secs:02d} remaining"
                st.info(f"⏱ Duty timer running • {remaining_msg}")
            else:
                st.error("⚠️ Time over! Please finish and mark Done.")
        if started_dt:
            st.caption(f"Started at {started_dt.strftime('%I:%M %p')} IST")
        if st.button("✅ Mark Done", use_container_width=True, key="duty_mark_done_btn"):
            ok = mark_duty_done_supabase(supabase, active_run_id)
            if ok:
                for k in ["active_duty_run_id", "active_duty_due_at", "active_duty_started_at", "active_duty_est_minutes"]:
                    st.session_state[k] = None
                st.toast("Duty marked DONE ✅", icon="✅")
                st.rerun()
        return
    today = now_ist().date()
    assignments = fetch_active_duty_assignments(supabase, selected_assistant)
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    runs = fetch_duty_runs_since(
        supabase,
        selected_assistant,
        (week_start if week_start < month_start else month_start).isoformat(),
    )
    pending = compute_pending_duties(assignments, runs, today)
    total_pending = len(pending["WEEKLY"]) + len(pending["MONTHLY"])
    free_minutes = compute_free_minutes_for_assistant(schedule_df, selected_assistant)
    if free_minutes is None:
        st.caption("Free window unknown from schedule; showing pending duties.")
    elif free_minutes <= 0:
        if total_pending > 0:
            st.warning("Currently busy with a case. Duties will appear when free.")
        else:
            st.caption("No pending duties right now.")
        return
    if total_pending == 0:
        st.success("No pending duties 🎉")
        return
    st.warning(f"✅ You are free for ~{free_minutes} min. Pending duties: {total_pending}")
    def _fits(duty):
        if free_minutes is None:
            return True
        try:
            return int(duty.get("est_minutes", 0)) <= int(free_minutes or 0)
        except Exception:
            return False
    weekly_fit = [d for d in pending["WEEKLY"] if _fits(d)]
    monthly_fit = [d for d in pending["MONTHLY"] if _fits(d)]
    with st.expander("Pick a duty"):
        tab1, tab2 = st.tabs([f"Weekly ({len(weekly_fit)})", f"Monthly ({len(monthly_fit)})"])
        def _pick_ui(duties: list[dict[str, Any]], tab_key: str):
            if not duties:
                st.info("No duties fit in the current free window.")
                return
            labels = [f"{d['title']} • {d['est_minutes']} min" for d in duties]
            idx = st.selectbox(
                "Select duty",
                options=list(range(len(labels))),
                format_func=lambda i: labels[i],
                key=f"duty_select_{tab_key}",
            )
            if st.button("▶ Start", use_container_width=True, key=f"duty_start_{tab_key}"):
                run_id, payload = start_duty_run_supabase(supabase, selected_assistant, duties[idx])
                if run_id:
                    st.session_state.active_duty_run_id = run_id
                    st.session_state.active_duty_due_at = payload.get("due_at")
                    st.session_state.active_duty_started_at = payload.get("started_at")
                    st.session_state.active_duty_est_minutes = payload.get("est_minutes")
                    st.toast("Duty timer started ✅", icon="✅")
                    st.rerun()
                else:
                    st.error("Failed to start duty. Please try again.")
        with tab1:
            _pick_ui(weekly_fit, "weekly")
        with tab2:
            _pick_ui(monthly_fit, "monthly")
def render_assistant_overview_widget():
    """Display all assistants and their current status (BUSY with timer or AVAILABLE)."""
    st.markdown("### 👥 Assistant Overview")
    # Load all assistants
    try:
        assistants_df = load_profiles("Assistants")
        assistants = sorted(assistants_df["name"].dropna().astype(str).str.strip().unique().tolist()) if not assistants_df.empty else []
    except Exception:
        assistants = []
    if not assistants:
        st.caption("No assistants found in Assistants sheet.")
        return
    # Get all active duty runs
    try:
        duty_runs_df = load_duty_runs_sheet()
        active_runs = duty_runs_df[duty_runs_df["status"].astype(str).str.upper() == "IN_PROGRESS"] if not duty_runs_df.empty else pd.DataFrame()
    except Exception:
        active_runs = pd.DataFrame()
    # Build overview
    overview_data = []
    for assistant in assistants:
        # Check if assistant has active duty
        assistant_run = active_runs[active_runs["assistant"].astype(str).str.strip() == assistant] if not active_runs.empty else pd.DataFrame()
        if not assistant_run.empty:
            run = assistant_run.iloc[0]
            due_dt = _parse_iso_ts(run.get("due_at"))
            if due_dt:
                delta = due_dt - now_ist()
                if delta.total_seconds() > 0:
                    mins = int(delta.total_seconds() // 60)
                    secs = int(delta.total_seconds() % 60)
                    status = f"🔴 BUSY • {mins:02d}:{secs:02d} remaining"
                else:
                    status = "🔴 BUSY • ⚠️ Time over!"
            else:
                status = "🔴 BUSY"
        else:
            status = "🟢 AVAILABLE"
        overview_data.append({
            "Assistant": assistant,
            "Status": status,
        })
    # Display as table
    overview_df = pd.DataFrame(overview_data)
    st.dataframe(overview_df, use_container_width=True, hide_index=True)
# ================ DUTY ADMIN (SUPABASE) ================
def render_duties_master_admin(supabase):
    st.subheader("🗂 Duties Master (Weekly / Monthly)")
    # Load assistants for selection
    try:
        assistants_df = load_profiles("Assistants")
        available_assistants = assistants_df["name"].dropna().unique().tolist() if not assistants_df.empty else []
        available_assistants = sorted([str(a).strip() for a in available_assistants if str(a).strip()])
    except Exception:
        available_assistants = []
    with st.form("add_duty_form"):
        col1, col2 = st.columns(2)
        with col1:
            title = st.text_input("Duty Name *", key="duty_title_input")
            frequency = st.selectbox("Frequency *", ["WEEKLY", "MONTHLY"], key="duty_freq_select")
            op = st.selectbox("OP", ["ANY", "OP1", "OP2", "OP3"], key="duty_op_select")
        with col2:
            default_minutes = st.number_input("Estimated Time (minutes) *", min_value=5, step=5, value=15, key="duty_minutes_input")
            active = st.checkbox("Active", value=True, key="duty_active_check")
        # Add assistants multi-select
        selected_assistants = st.multiselect(
            "Assign to Assistants (optional)",
            options=available_assistants,
            key="duty_assistants_multiselect",
            help="Select assistants to auto-assign this duty to them"
        )
        submitted = st.form_submit_button("➕ Add Duty")
        if submitted:
            if not title:
                st.error("Duty name required")
            else:
                try:
                    # Create duty in Duties_Master
                    duties_df = load_duties_master_sheet()
                    duty_id = str(uuid.uuid4())
                    new_duty = pd.DataFrame([{
                        "id": duty_id,
                        "title": title,
                        "frequency": frequency,
                        "default_minutes": int(default_minutes),
                        "op": op,
                        "active": active,
                        "created_at": now_ist().isoformat(),
                    }])
                    combined = pd.concat([duties_df, new_duty], ignore_index=True)
                    save_duties_master_sheet(combined)
                    # Auto-assign to selected assistants
                    if selected_assistants:
                        assignments_df = load_duty_assignments_sheet()
                        new_assignments = []
                        for assistant in selected_assistants:
                            new_assignments.append({
                                "id": str(uuid.uuid4()),
                                "duty_id": duty_id,
                                "assistant": assistant,
                                "op": op,
                                "est_minutes": int(default_minutes),
                                "active": "true",
                            })
                        new_assignments_df = pd.DataFrame(new_assignments)
                        combined_assignments = pd.concat([assignments_df, new_assignments_df], ignore_index=True)
                        save_duty_assignments_sheet(combined_assignments)
                        st.success(f"Duty added and assigned to {len(selected_assistants)} assistant(s) ✅")
                    else:
                        st.success("Duty added successfully ✅")
                    st.rerun()
                except Exception as e:
                    st.error(f"Failed to add duty: {e}")
    try:
        duties = load_duties_master_sheet().to_dict("records")
    except Exception as e:
        st.error(f"Failed to load duties: {e}")
        duties = []
    if duties:
        st.data_editor(
            duties,
            use_container_width=True,
            disabled=["id", "created_at"],
            num_rows="dynamic",
            key="duties_master_editor",
        )
def render_duty_assignment_admin(supabase, assistants: list[str]):
    st.subheader("👥 Assign Duties to Assistants")
    # Load duties
    try:
        duties_df = load_duties_master_sheet()
        active_duties = duties_df[duties_df["active"].astype(str).str.lower() == "true"]
        duties = active_duties.to_dict("records") if not active_duties.empty else []
    except Exception as e:
        st.error(f"Failed to load duties: {e}")
        duties = []
    if not duties:
        st.warning("No active duties found.")
        return
    # Load assistants from Assistants sheet (not from schedule)
    try:
        assistants_df = load_profiles("Assistants")
        sheet_assistants = assistants_df["name"].dropna().unique().tolist() if not assistants_df.empty else []
        sheet_assistants = sorted([str(a).strip() for a in sheet_assistants if str(a).strip()])
    except Exception:
        sheet_assistants = []
    if not sheet_assistants:
        st.warning("No assistants found in Assistants sheet. Add assistants first.")
        return
    duty_map = {d["title"]: d["id"] for d in duties if d.get("title")}
    with st.form("assign_duty_form"):
        col1, col2 = st.columns(2)
        with col1:
            duty_title = st.selectbox("Duty *", list(duty_map.keys()), key="assign_duty_select")
            assistant = st.selectbox("Assistant *", sheet_assistants, key="assign_assistant_select")
        with col2:
            est_minutes = st.number_input("Time for this Assistant (minutes)", min_value=5, step=5, value=15, key="assign_minutes_input")
            op = st.selectbox("OP (optional)", ["", "OP1", "OP2", "OP3"], key="assign_op_select")
            active = st.checkbox("Active", value=True, key="assign_active_check")
        submitted = st.form_submit_button("📌 Assign Duty")
        if submitted:
            try:
                assigns_df = load_duty_assignments_sheet()
                # Check if assignment exists and update or append
                existing = assigns_df[
                    (assigns_df["duty_id"] == duty_map.get(duty_title)) &
                    (assigns_df["assistant"].astype(str).str.strip() == assistant)
                ]
                new_assign = {
                    "id": str(uuid.uuid4()),
                    "duty_id": duty_map.get(duty_title),
                    "assistant": assistant,
                    "est_minutes": int(est_minutes),
                    "op": op or "",
                    "active": active,
                }
                if not existing.empty:
                    # Update existing
                    assigns_df.loc[existing.index[0]] = new_assign
                else:
                    # Add new
                    new_row = pd.DataFrame([new_assign])
                    assigns_df = pd.concat([assigns_df, new_row], ignore_index=True)
                save_duty_assignments_sheet(assigns_df)
                st.success("Duty assigned successfully ✅")
                st.rerun()
            except Exception as e:
                st.error(f"Failed to assign duty: {e}")
    try:
        assigns = load_duty_assignments_sheet().to_dict("records")
    except Exception as e:
        st.error(f"Failed to load assignments: {e}")
        assigns = []
    if assigns:
        st.data_editor(
            assigns,
            use_container_width=True,
            disabled=["id"],
            num_rows="dynamic",
            key="duty_assign_editor",
        )
def render_assistant_attendance_tab(schedule_df, excel_path):
    st.header("Assistants Attendance")
    today_str = datetime.now(IST).date().isoformat()
    assistants = get_assistants_list(schedule_df)
    att_df = load_attendance_sheet(excel_path)
    today_att = att_df[att_df["DATE"] == today_str].copy() if not att_df.empty else pd.DataFrame(columns=ATTENDANCE_COLUMNS)
    for name in assistants:
        if today_att.empty or name not in today_att["ASSISTANT"].values:
            new_row = {"DATE": today_str, "ASSISTANT": name, "PUNCH IN": "", "PUNCH OUT": ""}
            today_att = pd.concat([today_att, pd.DataFrame([new_row])], ignore_index=True)
    now_time = datetime.now(IST).time()
    def _decorate(row):
        in_str = str(row.get("PUNCH IN", "")).strip()
        out_str = str(row.get("PUNCH OUT", "")).strip()
        in_t = safe_str_to_time_obj(in_str) if in_str else None
        out_t = safe_str_to_time_obj(out_str) if out_str else None
        worked_mins, status = calc_worked_minutes(in_t, out_t, now_time)
        row["WORKED MINS"] = mins_to_hhmm(worked_mins)
        row["STATUS"] = status
        return row
    display_df = pd.DataFrame([_decorate(row.copy()) for _, row in today_att.iterrows()]) if not today_att.empty else pd.DataFrame(columns=ATTENDANCE_COLUMNS + ["WORKED MINS", "STATUS"])
    edited = st.data_editor(
        display_df,
        use_container_width=True,
        num_rows="fixed",
        column_config={
            "ASSISTANT": st.column_config.TextColumn(disabled=True),
            "PUNCH IN": st.column_config.TextColumn(help="HH:MM"),
            "PUNCH OUT": st.column_config.TextColumn(help="HH:MM"),
            "WORKED MINS": st.column_config.TextColumn(disabled=True),
            "STATUS": st.column_config.TextColumn(disabled=True),
        },
        key="assistants_attendance_editor"
    )
    out_rows = []
    for _, row in edited.iterrows():
        in_str = str(row.get("PUNCH IN", "")).strip()
        out_str = str(row.get("PUNCH OUT", "")).strip()
        in_t = safe_str_to_time_obj(in_str) if in_str else None
        out_t = safe_str_to_time_obj(out_str) if out_str else None
        worked_mins, status = calc_worked_minutes(in_t, out_t, now_time)
        row["WORKED MINS"] = mins_to_hhmm(worked_mins)
        row["STATUS"] = status
        out_rows.append(row)
    edited_final = pd.DataFrame(out_rows)
    if st.button("💾 Save Attendance"):
        att_df = att_df[att_df["DATE"] != today_str]
        att_df = pd.concat([att_df, edited_final[ATTENDANCE_COLUMNS]], ignore_index=True)
        save_attendance_sheet(excel_path, att_df)
        st.success("Attendance saved!")
        st.rerun()
def render_schedule_summary_chips(df: pd.DataFrame):
    """Render top summary chips for schedule STATUS counts."""
    if df is None or df.empty or "STATUS" not in df.columns:
        return
    status_series = df["STATUS"].astype(str).str.upper().str.strip()
    total = len(status_series)
    ongoing = status_series.str.contains("ON GOING|ONGOING").sum()
    waiting = status_series.str.contains("WAITING").sum()
    arrived = status_series.str.contains("ARRIVED").sum()
    completed = status_series.str.contains("DONE|COMPLETED").sum()
    cancelled = status_series.str.contains("CANCEL").sum()
    chips = [
        ("Total Cases", total, "info"),
        ("Ongoing", ongoing, "success"),
        ("Waiting", waiting, "warning"),
        ("Arrived", arrived, "secondary"),
        ("Completed", completed, "info"),
        ("Cancelled", cancelled, "danger"),
    ]
    chips_html = "".join(
        f'<div class="summary-chip {cls}"><div class="label">{label}</div><div class="value">{val}</div></div>'
        for label, val, cls in chips
    )
    st.markdown(f'<div class="summary-row">{chips_html}</div>', unsafe_allow_html=True)
def render_compact_dashboard(df_schedule: pd.DataFrame):
    """Compact single-screen dashboard with weekly off + schedule summary (pixel-matched layout)."""
    st.markdown(
        """
        <style>
        body, .stApp {
            background: #f8fafc !important;
        }
        .block-container {padding-top:0.3rem !important;}
        h1,h2,h3{margin:0.3rem 0 !important;}
        .dash-title {text-align:center; color:#1e293b; font-size:28px; font-weight:800; letter-spacing:0.5px;}
        .dash-subtitle {text-align:center; margin-top:-10px; color:#64748b; font-weight:700;}
        div[data-testid="stVerticalBlockBorderWrapper"] {
            background: #f8fafc;
            border: 1px solid #3b82f6;
            border-radius: 20px;
            box-shadow: 0 18px 36px rgba(20, 17, 15, 0.18);
        }
        div[data-testid="stVerticalBlockBorderWrapper"] > div {
            padding: 16px 18px 18px 18px;
            border-radius: 20px;
        }
        .v-divider {width: 1px; background: #3b82f6; min-height: 280px; margin: 8px auto;}
        .panel-title {font-size: 20px; font-weight: 800; margin-bottom: 8px; display:flex; align-items:center; gap:8px;}
        .panel-title .link {font-size: 14px; opacity: 0.6; margin-left: 4px;}
        .alert-card {background: #3b82f6; border: 1px solid #64748b; border-radius: 12px; padding: 12px; color: #1e293b; margin-bottom: 8px; display:flex; gap:10px; align-items:center;}
        .alert-icon {width: 28px; height: 28px; border-radius: 50%; border: 2px solid #2563eb; display:flex; align-items:center; justify-content:center; color:#2563eb; font-weight:700;}
        .alert-title {font-weight:700; margin-bottom:2px;}
        .alert-sub {opacity:0.85;}
        .manage-pill {background: #f8fafc; border: 1px solid #3b82f6; border-radius: 12px; padding: 10px 12px; display:inline-flex; align-items:center; gap:8px; margin-top:6px; color:#1e293b;}
        .metric-card {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.7), rgba(255, 255, 255, 0.4));
            backdrop-filter: blur(16px);
            -webkit-backdrop-filter: blur(16px);
            border: 1px solid rgba(255, 255, 255, 0.3);
            border-radius: 20px;
            padding: 20px;
            text-align: center;
            min-height: 100px;
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            box-shadow: 0 4px 16px rgba(37, 99, 235, 0.08);
        }
        .metric-card:hover {
            transform: translateY(-4px) scale(1.02);
            box-shadow: 0 12px 32px rgba(37, 99, 235, 0.15);
            border-color: rgba(59, 130, 246, 0.3);
        }
        .metric-value {
            font-size: 32px;
            font-weight: 800;
            background: linear-gradient(135deg, #3b82f6, #2563eb);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 8px;
            line-height: 1;
        }
        .metric-title {
            font-size: 11px;
            font-weight: 700;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 1px;
        }
        .metric-icon {
            width: 40px;
            height: 40px;
            margin: 0 auto 12px;
            display: flex;
            align-items: center;
            justify-content: center;
            border-radius: 12px;
            background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(37, 99, 235, 0.05));
            color: #2563eb;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) {
            animation: cardEntrance 0.6s cubic-bezier(0.34, 1.56, 0.64, 1) backwards;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker):nth-child(1) { animation-delay: 0.1s; }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker):nth-child(2) { animation-delay: 0.2s; }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker):nth-child(3) { animation-delay: 0.3s; }
        @keyframes cardEntrance {
            from {
                opacity: 0;
                transform: translateY(20px) scale(0.95);
            }
            to {
                opacity: 1;
                transform: translateY(0) scale(1);
            }
        }
        .metrics-grid {display:grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap:10px; margin-bottom:6px;}
        #metrics-anchor + div[data-testid="stHorizontalBlock"] {margin-top: 0 !important;}
        .controls-row .stButton>button {height: 34px !important; border-radius: 10px !important;}
        .controls-row [data-baseweb="select"] > div {min-height: 42px !important; border-radius: 12px !important;}
        button[kind="primary"] {background:linear-gradient(135deg, #3b82f6 0%, #2563eb 100%) !important; border:1px solid rgba(255, 255, 255, 0.3) !important; color:#ffffff !important; box-shadow:0 4px 20px rgba(37, 99, 235, 0.3) !important; backdrop-filter:blur(10px) !important; transition:all 0.3s ease !important;}
        button[kind="primary"]:hover {transform:translateY(-2px) !important; box-shadow:0 8px 30px rgba(37, 99, 235, 0.4) !important; background:linear-gradient(135deg, #60a5fa 0%, #3b82f6 100%) !important;}
        button[kind="secondary"] {background:linear-gradient(135deg, rgba(255, 255, 255, 0.7), rgba(255, 255, 255, 0.3)) !important; border:1px solid rgba(37, 99, 235, 0.3) !important; color:#2563eb !important; backdrop-filter:blur(10px) !important; box-shadow:0 4px 16px rgba(37, 99, 235, 0.1) !important;}
        .section-divider {height:1px; background: #3b82f6; margin: 14px 0;}
        .search-row input {background:#f8fafc !important; border-radius:10px !important; border:1px solid #3b82f6 !important;}
        [data-testid="stDataFrameContainer"] {border-radius: 14px !important; border: 1px solid #3b82f6 !important; box-shadow: 0 8px 20px rgba(20,17,15,0.08) !important;}
        [data-testid="stDataFrameContainer"] thead th {background:#f8fafc !important; color:#2563eb !important; font-weight:700 !important;}
        .summary-bar {background: #f8fafc; border: 1px solid #3b82f6; border-radius: 14px; padding: 6px 10px; margin-top: 12px;}
        .compact-dashboard [data-testid="stVerticalBlock"] {gap: 0.5rem;}
        .compact-dashboard [data-testid="stHorizontalBlock"] {gap: 0.6rem;}
        .schedule-cards {display:grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap:12px; margin-top: 6px;}
        .schedule-card {background:#f8fafc; border:1px solid #3b82f6; border-radius:14px; padding:10px; box-shadow:0 6px 14px rgba(20,17,15,0.06); display:flex; flex-direction:column; gap:8px; min-height:180px;}
        .card-shell-marker {display:none;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) {
            position: relative;
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.9), rgba(255, 255, 255, 0.6));
            backdrop-filter: blur(24px);
            -webkit-backdrop-filter: blur(24px);
            border: 1px solid rgba(255, 255, 255, 0.4);
            border-radius: 24px;
            box-shadow: 0 8px 32px rgba(37, 99, 235, 0.08), 0 1px 2px rgba(0, 0, 0, 0.02), inset 0 1px 0 rgba(255, 255, 255, 0.8);
            transition: all 0.5s cubic-bezier(0.4, 0, 0.2, 1);
            overflow: hidden;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker)::before {
            content: '';
            position: absolute;
            top: 0;
            left: 0;
            right: 0;
            height: 1px;
            background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.8), transparent);
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker):hover {
            transform: translateY(-8px) scale(1.01);
            box-shadow: 0 20px 60px rgba(37, 99, 235, 0.15), 0 4px 8px rgba(0, 0, 0, 0.05), inset 0 1px 0 rgba(255, 255, 255, 1);
            border-color: rgba(59, 130, 246, 0.3);
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) > div {padding:0; display:flex; flex-direction:column; gap:10px; min-height:200px;}
        .card-inner {background:#fff; border-radius:18px; border:1px solid #eceff5; box-shadow:0 12px 24px rgba(22, 24, 31, 0.14); padding:0 16px 12px; display:flex; flex-direction:column; gap:8px;}
        .card-inner .card-status-banner {margin-top:0;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) [data-testid="stHorizontalBlock"] {gap: 0.6rem; align-items:center; justify-content:flex-start;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stButton>button,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) [data-testid="stButton"] > button,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind] {height: 16px !important; border-radius: 8px !important; font-weight: 600; text-transform: none; letter-spacing: 0; white-space: nowrap; word-break: keep-all; overflow-wrap: normal; min-width: 55px; padding: 0 5px !important; font-size: 9px; line-height: 1; flex-shrink: 0; display: inline-flex; align-items: center; justify-content: center; gap: 3px; width: 100%; box-shadow: 0 4px 10px rgba(22, 24, 31, 0.12); color:#3f434a;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stButton>button *,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind] * {white-space: nowrap;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="stHorizontalBlock"]:has(.card-action-marker) {flex-wrap: wrap; row-gap: 0.5rem; align-items:center; justify-content:flex-end;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="stHorizontalBlock"]:has(.card-action-marker) > div {min-width: 70px; flex: 1 1 70px;}
        @media (min-width: 1100px) {
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="stHorizontalBlock"]:has(.card-action-marker) {flex-wrap: nowrap;}
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="stHorizontalBlock"]:has(.card-action-marker) > div {flex: 0 0 auto;}
        }
        .card-actions-row {
            margin-top: 8px;
            display: flex;
            flex-wrap: wrap;
            gap: 0.5rem;
            align-items: center;
            justify-content: flex-start;
        }
        .card-actions-row .stCheckbox {
            margin: 0;
        }
        .card-actions-row .stCheckbox label {
            font-size: 13px;
            font-weight: 600;
            letter-spacing: 0.1px;
        }
        .card-actions-row .stButton>button,
        .card-actions-row button[kind] {
            min-width: 70px;
            border-radius: 10px !important;
            height: 30px !important;
            font-size: 12px !important;
        }
        .card-actions-row div[data-testid="stHorizontalBlock"] {align-items:center;}
        .card-actions-row div[data-testid="stHorizontalBlock"]:has(.stCheckbox) {justify-content:flex-start !important;}
        .card-actions-row div[data-testid="stHorizontalBlock"]:has(.stCheckbox) div[data-testid="column"]:has(.card-action-done) {margin-left:auto;}
        .card-actions-row div[data-testid="column"]:has(.card-action-done),
        .card-actions-row div[data-testid="column"]:has(.card-action-edit),
        .card-actions-row div[data-testid="column"]:has(.card-action-cancel) {min-width: 72px;}
        .card-details-row {
            margin-top: 8px;
            border: 1px solid #d9dde3;
            border-radius: 12px;
            padding: 5px 8px;
            background: #fff;
            box-shadow: 0 6px 14px rgba(24, 28, 36, 0.10);
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox {margin-top: 8px;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox label {font-size: 13px; font-weight: 600; color:#2f333a; letter-spacing:0.1px; text-transform:none; white-space: normal; line-height: 1.3;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox div[data-baseweb="checkbox"] > div,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox div[data-baseweb="checkbox"] > label > div {width: 18px; height: 18px; border-radius: 5px; border: 1.5px solid #c3c8d0; background: #ffffff;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox div[data-baseweb="checkbox"] input:checked + div {background:#2f63e8; border-color:#2f63e8;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind="primary"] {
            background: linear-gradient(135deg, #10b981, #059669) !important;
            border: 1px solid rgba(255, 255, 255, 0.3) !important;
            color: white !important;
            box-shadow: 0 4px 12px rgba(16, 185, 129, 0.25) !important;
            backdrop-filter: blur(10px) !important;
            -webkit-backdrop-filter: blur(10px) !important;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind="primary"]:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 8px 20px rgba(16, 185, 129, 0.35) !important;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind="secondary"] {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.8), rgba(255, 255, 255, 0.5)) !important;
            border: 1px solid rgba(59, 130, 246, 0.3) !important;
            color: #2563eb !important;
            box-shadow: 0 4px 12px rgba(59, 130, 246, 0.15) !important;
            backdrop-filter: blur(10px) !important;
            -webkit-backdrop-filter: blur(10px) !important;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind="secondary"]:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 8px 20px rgba(59, 130, 246, 0.25) !important;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="column"]:has(.card-action-cancel) button {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.8), rgba(255, 255, 255, 0.5)) !important;
            border: 1px solid rgba(239, 68, 68, 0.3) !important;
            color: #dc2626 !important;
            box-shadow: 0 4px 12px rgba(239, 68, 68, 0.15) !important;
            backdrop-filter: blur(10px) !important;
            -webkit-backdrop-filter: blur(10px) !important;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="column"]:has(.card-action-cancel) button:hover {
            transform: translateY(-2px) !important;
            box-shadow: 0 8px 20px rgba(239, 68, 68, 0.25) !important;
        }
        .card-action-marker {display:none;}
        .card-status-banner {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            padding: 8px 16px;
            border-radius: 12px;
            font-weight: 700;
            font-size: 12px;
            letter-spacing: 0.5px;
            text-transform: uppercase;
            backdrop-filter: blur(10px);
            -webkit-backdrop-filter: blur(10px);
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.08);
            margin: 0 0 16px 0;
            transition: all 0.3s ease;
        }
        .card-status-banner.waiting {
            background: linear-gradient(135deg, rgba(251, 191, 36, 0.2), rgba(245, 158, 11, 0.15));
            border: 1px solid rgba(251, 191, 36, 0.3);
            color: #d97706;
        }
        .card-status-banner.ongoing {
            background: linear-gradient(135deg, rgba(96, 165, 250, 0.2), rgba(59, 130, 246, 0.15));
            border: 1px solid rgba(59, 130, 246, 0.3);
            color: #2563eb;
        }
        .card-status-banner.arrived {
            background: linear-gradient(135deg, rgba(148, 163, 184, 0.2), rgba(100, 116, 139, 0.15));
            border: 1px solid rgba(100, 116, 139, 0.3);
            color: #475569;
        }
        .card-status-banner.completed {
            background: linear-gradient(135deg, rgba(52, 211, 153, 0.2), rgba(16, 185, 129, 0.15));
            border: 1px solid rgba(16, 185, 129, 0.3);
            color: #059669;
        }
        .card-status-banner.cancelled {
            background: linear-gradient(135deg, rgba(248, 113, 113, 0.2), rgba(239, 68, 68, 0.15));
            border: 1px solid rgba(239, 68, 68, 0.3);
            color: #dc2626;
        }
        .status-icon {
            width: 16px;
            height: 16px;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .status-dot {
            width: 8px;
            height: 8px;
            border-radius: 50%;
            background: currentColor;
            box-shadow: 0 0 12px currentColor, 0 0 24px currentColor;
            animation: pulse-glow 2s ease-in-out infinite;
        }
        @keyframes pulse-glow {
            0%, 100% { opacity: 1; transform: scale(1); }
            50% { opacity: 0.8; transform: scale(0.95); }
        }
        .card-head {display:flex; align-items:center; gap:12px;}
        .card-title {display:flex; flex-direction:column; gap:2px;}
        .card-avatar {
            width: 60px;
            height: 60px;
            border-radius: 50%;
            background: linear-gradient(135deg, #3b82f6, #2563eb);
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: 800;
            font-size: 22px;
            position: relative;
            box-shadow: 0 8px 24px rgba(37, 99, 235, 0.25), 0 2px 8px rgba(0, 0, 0, 0.1), inset 0 -2px 8px rgba(0, 0, 0, 0.15), inset 0 2px 8px rgba(255, 255, 255, 0.3);
            transition: all 0.4s cubic-bezier(0.34, 1.56, 0.64, 1);
            border: 3px solid rgba(255, 255, 255, 0.5);
        }
        .card-avatar::before {
            content: '';
            position: absolute;
            top: 10%;
            left: 15%;
            width: 40%;
            height: 40%;
            background: radial-gradient(circle at center, rgba(255, 255, 255, 0.5), transparent);
            border-radius: 50%;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker):hover .card-avatar {
            transform: scale(1.1) translateY(-2px);
            box-shadow: 0 12px 32px rgba(37, 99, 235, 0.35), 0 4px 12px rgba(0, 0, 0, 0.15), inset 0 -2px 8px rgba(0, 0, 0, 0.2), inset 0 2px 12px rgba(255, 255, 255, 0.4);
        }
        .card-name {
            font-size: 20px;
            font-weight: 800;
            letter-spacing: 0.3px;
            text-transform: uppercase;
            background: linear-gradient(135deg, #1e293b 0%, #475569 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
            margin-bottom: 4px;
        }
        .card-time {font-size:13px; color:#6f757d;}
        .card-info {
            display: flex;
            flex-direction: column;
            gap: 12px;
            margin: 16px 0;
        }
        .info-row {
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 10px 12px;
            border-radius: 12px;
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.6), rgba(255, 255, 255, 0.3));
            backdrop-filter: blur(8px);
            -webkit-backdrop-filter: blur(8px);
            border: 1px solid rgba(255, 255, 255, 0.3);
            transition: all 0.3s ease;
        }
        .info-row:hover {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.8), rgba(255, 255, 255, 0.5));
            border-color: rgba(59, 130, 246, 0.2);
            transform: translateX(4px);
        }
        .info-icon {
            width: 36px;
            height: 36px;
            border-radius: 12px;
            background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(37, 99, 235, 0.05));
            border: 1px solid rgba(59, 130, 246, 0.15);
            display: flex;
            align-items: center;
            justify-content: center;
            color: #2563eb;
            backdrop-filter: blur(8px);
            -webkit-backdrop-filter: blur(8px);
            transition: all 0.3s ease;
            flex-shrink: 0;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker):hover .info-icon {
            background: linear-gradient(135deg, rgba(59, 130, 246, 0.15), rgba(37, 99, 235, 0.08));
            border-color: rgba(59, 130, 246, 0.25);
            transform: scale(1.05);
        }
        .info-icon.doctor-icon {font-size:16px;}
        .info-icon.staff-icon {font-size:16px;}
        .info-icon-svg {
            display: block;
            width: 20px;
            height: 20px;
        }
        .info-text {
            font-size: 14px;
            font-weight: 500;
            color: #1e293b;
            flex: 1;
        }
        .info-label {
            font-size: 11px;
            font-weight: 600;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 2px;
        }
        .card-subdivider {height:1px; background:#e4e6eb; margin: 10px 0 6px;}
        .card-divider {height:1px; background:#e4e6eb; margin: 12px 0;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) [data-testid="stExpander"] {border:1px solid #d9dde3; border-radius:12px; background:#f7f8fa; margin-top:6px;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) summary {padding:10px 12px; font-weight:600; color:#60656c; display:flex; align-items:center; gap:10px; font-size:13px; background:#fff; border-radius:12px; box-shadow:0 4px 12px rgba(20,17,15,0.08);}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) summary::before {content:"›"; color:#7a8087; font-size:18px;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) summary::after {content:"⋯"; margin-left:auto; color:#9aa0a7; font-size:18px;}
        .card-expand {font-size:12px; color:#6f757d; border-top:1px solid #d9dde3; padding-top:8px; display:flex; align-items:center; justify-content:space-between; margin-top:4px;}
        </style>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("<div class='compact-dashboard'>", unsafe_allow_html=True)
    st.write("")
    with st.container(border=True):
        left, divider, right = st.columns([1.05, 0.04, 1.6], gap="small")
        with left:
            st.markdown("<div class='panel-title'>🗓️ Assistants Weekly Off <span class='link'>🔗</span></div>", unsafe_allow_html=True)
            today_idx = now_ist().weekday()
            tomorrow_idx = (today_idx + 1) % 7
            weekday_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
            weekly_off_map = _get_profiles_cache().get("weekly_off_map", WEEKLY_OFF)
            today_off = weekly_off_map.get(today_idx, [])
            tomorrow_off = weekly_off_map.get(tomorrow_idx, [])
            if today_off:
                st.markdown(
                    "<div class='alert-card'>"
                    "<div class='alert-icon'>⛔</div>"
                    f"<div><div class='alert-title'>Today ({weekday_names[today_idx]})</div>"
                    f"<div class='alert-sub'>{', '.join(today_off)} – Cannot be allocated</div></div></div>",
                    unsafe_allow_html=True,
                )
            else:
                st.success(f"Today ({weekday_names[today_idx]}): All assistants available")
            if tomorrow_off:
                st.markdown(
                    "<div class='alert-card'>"
                    "<div class='alert-icon'>⛔</div>"
                    f"<div><div class='alert-title'>Tomorrow ({weekday_names[tomorrow_idx]})</div>"
                    f"<div class='alert-sub'>{', '.join(tomorrow_off)} – Cannot be allocated</div></div></div>",
                    unsafe_allow_html=True,
                )
            else:
                st.info(f"Tomorrow ({weekday_names[tomorrow_idx]}): All assistants available")
            manage_clicked = st.button("⚠️ Manage Reminders", key="compact_manage_reminders")
            if manage_clicked:
                st.session_state["show_compact_reminders"] = True
            if st.session_state.get("show_compact_reminders"):
                with st.expander("🔔 Manage Reminders", expanded=True):
                    st.checkbox("Enable 15-minute reminders", value=True, key="compact_enable_reminders")
                    st.selectbox(
                        "Default snooze (seconds)",
                        options=[30, 60, 90, 120, 150, 180, 300],
                        index=0,
                        key="compact_default_snooze_seconds",
                    )
                    st.caption("Reminders alert 15 minutes before a patient's In Time.")
        with divider:
            st.markdown("<div class='v-divider'></div>", unsafe_allow_html=True)
        with right:
            st.markdown("<div class='panel-title'>📋 Full Schedule</div>", unsafe_allow_html=True)
            status_series = df_schedule["STATUS"].astype(str).str.upper().str.strip() if ("STATUS" in df_schedule.columns and not df_schedule.empty) else pd.Series(dtype=str)
            total = len(status_series)
            ongoing = status_series.str.contains("ON GOING|ONGOING").sum()
            waiting = status_series.str.contains("WAITING").sum()
            arrived = status_series.str.contains("ARRIVED").sum()
            completed = status_series.str.contains("DONE|COMPLETED").sum()
            cancelled = status_series.str.contains("CANCEL").sum()
            metrics_html = (
                "<div class='metrics-grid'>"
                f"<div class='metric-card'><div class='metric-title'>TOTAL</div><div class='metric-value'>{total}</div></div>"
                f"<div class='metric-card'><div class='metric-title'>ONGOING</div><div class='metric-value'>{ongoing}</div></div>"
                f"<div class='metric-card'><div class='metric-title'>WAITING</div><div class='metric-value'>{waiting}</div></div>"
                f"<div class='metric-card'><div class='metric-title'>ARRIVED</div><div class='metric-value'>{arrived}</div></div>"
                f"<div class='metric-card'><div class='metric-title'>COMPLETED</div><div class='metric-value'>{completed}</div></div>"
                f"<div class='metric-card'><div class='metric-title'>CANCELLED</div><div class='metric-value'>{cancelled}</div></div>"
                "</div>"
            )
            st.markdown(metrics_html, unsafe_allow_html=True)
            st.markdown("<div id='metrics-anchor'></div>", unsafe_allow_html=True)
            st.markdown("<div class='controls-row'>", unsafe_allow_html=True)
            b1, b2, b3 = st.columns([1.2, 1.2, 1.6], gap="small")
            with b1:
                st.button("➕ Add Patient", use_container_width=True, key="compact_add_patient", type="primary")
            with b2:
                st.button(
                    "?? Save Changes",
                    use_container_width=True,
                    key="compact_save_changes",
                    type="secondary",
                    disabled=bool(st.session_state.get("is_saving")),
                )
            with b3:
                st.selectbox("Delete row", ["Delete row..."], label_visibility="collapsed", key="compact_delete_row")
            st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("<div class='section-divider'></div>", unsafe_allow_html=True)
        header_left, header_right = st.columns([3, 1], gap="small")
        with header_left:
            st.markdown("<div class='panel-title'>📋 Full Schedule</div>", unsafe_allow_html=True)
        with header_right:
            st.markdown("<div class='search-row'>", unsafe_allow_html=True)
            st.text_input("Search patient...", label_visibility="collapsed", placeholder="Search patient...", key="compact_search")
            st.markdown("</div>", unsafe_allow_html=True)
        if df_schedule is None or df_schedule.empty:
            df_display = pd.DataFrame({
                "Patient Name": ["AJOY CHOUDHURY", "SHRUTI LAD"],
                "In Time": ["01:09 AM", "01:09 AM"],
                "Out Time": ["01:14 AM", "01:14 AM"],
                "Procedure": ["PLT/INE", "PSE/IENN"],
                "Doctor": ["DR. FARHATH", "DR. SHRUTI"],
                "FIRST": ["ANISHA", "LAWANA"],
                "SECOND": ["ANISHA", "LAWANA"],
                "THIRD": ["NITIN", "MUKHILA"],
                "CASE PAPER": ["None", "None"],
                "SUCTION": ["None", "None"],
                "Status": ["WAITING", "WAITING"],
            })
        else:
            df_display = df_schedule.copy()
            rename_map = {}
            if "Patient Name" not in df_display.columns and "Patient" in df_display.columns:
                rename_map["Patient"] = "Patient Name"
            if "DR." in df_display.columns and "Doctor" not in df_display.columns:
                rename_map["DR."] = "Doctor"
            df_display = df_display.rename(columns=rename_map)
            desired_cols = [
                c
                for c in [
                    "Patient Name",
                    "In Time",
                    "Out Time",
                    "Procedure",
                    "Doctor",
                    "DR.",
                    "FIRST",
                    "SECOND",
                    "Third",
                    "THIRD",
                    "CASE PAPER",
                    "SUCTION",
                    "REMINDER_ROW_ID",
                    "STATUS",
                    "Status",
                ]
                if c in df_display.columns
            ]
            if desired_cols:
                df_display = df_display[desired_cols]
            if "STATUS" in df_display.columns and "Status" not in df_display.columns:
                df_display = df_display.rename(columns={"STATUS": "Status"})
        view_options = ["Cards"]
        if st.session_state.get("user_role") == "admin":
            view_options.append("Table")
        view_mode = st.radio(
            "View",
            view_options,
            horizontal=True,
            key="compact_view_mode",
            label_visibility="collapsed",
        )
        def _clean_text(val) -> str:
            if val is None or (isinstance(val, float) and pd.isna(val)):
                return ""
            text = str(val).strip()
            if text.lower() in {"nan", "none"}:
                return ""
            return text
        def _truthy(val) -> bool:
            if isinstance(val, bool):
                return val
            text = _clean_text(val).lower()
            return text in {"yes", "y", "true", "1", "done", "checked"}
        def _initials(name: str) -> str:
            parts = [p for p in name.strip().split() if p]
            if not parts:
                return "--"
            if len(parts) == 1:
                return parts[0][:2].upper()
            return (parts[0][0] + parts[-1][0]).upper()
        def _status_class(status: str) -> str:
            status_up = status.upper()
            if "WAIT" in status_up:
                return "waiting"
            if "ONGOING" in status_up or "ON GOING" in status_up:
                return "ongoing"
            if "ARRIVED" in status_up:
                return "arrived"
            if "DONE" in status_up or "COMPLETED" in status_up:
                return "completed"
            if "CANCEL" in status_up or "SHIFT" in status_up:
                return "cancelled"
            return "waiting"
        def _open_compact_edit_dialog(context: dict[str, Any]) -> None:
            st.session_state["compact_edit_context"] = context
            st.session_state["compact_edit_open"] = True
            row_key = str(context.get("row_key", "")).strip()
            if not row_key:
                return
            in_time_value = str(context.get("in_time", "") or "").strip()
            out_time_value = str(context.get("out_time", "") or "").strip()
            if in_time_value.upper() in {"N/A", "NONE", "NAT"}:
                in_time_value = ""
            if out_time_value.upper() in {"N/A", "NONE", "NAT"}:
                out_time_value = ""
            st.session_state[f"compact_popup_patient_{row_key}"] = str(context.get("patient", "") or "")
            in_hour, in_minute, in_ampm = _time_to_picker_parts(in_time_value)
            out_hour, out_minute, out_ampm = _time_to_picker_parts(out_time_value)
            st.session_state[f"compact_popup_in_hour_{row_key}"] = in_hour
            st.session_state[f"compact_popup_in_min_{row_key}"] = in_minute
            st.session_state[f"compact_popup_in_ampm_{row_key}"] = in_ampm
            st.session_state[f"compact_popup_out_hour_{row_key}"] = out_hour
            st.session_state[f"compact_popup_out_min_{row_key}"] = out_minute
            st.session_state[f"compact_popup_out_ampm_{row_key}"] = out_ampm
            st.session_state[f"compact_popup_status_{row_key}"] = str(context.get("status", "") or "")
            st.session_state[f"compact_popup_doctor_{row_key}"] = str(context.get("doctor", "") or "")
            st.session_state[f"compact_popup_procedure_{row_key}"] = str(context.get("procedure", "") or "")
            st.session_state[f"compact_popup_first_{row_key}"] = str(context.get("staff_first", "") or "")
            st.session_state[f"compact_popup_second_{row_key}"] = str(context.get("staff_second", "") or "")
            st.session_state[f"compact_popup_third_{row_key}"] = str(context.get("staff_third", "") or "")
            st.session_state[f"compact_popup_case_{row_key}"] = bool(context.get("case_paper", False))
            st.session_state[f"compact_popup_suction_{row_key}"] = bool(context.get("suction", False))
        def _close_compact_edit_dialog() -> None:
            st.session_state["compact_edit_open"] = False
            st.session_state["compact_edit_context"] = {}
        def _compact_normalize_time_input(raw_value: str) -> tuple[str, Optional[str]]:
            text = str(raw_value or "").strip()
            if not text:
                return "", None
            t = _coerce_to_time_obj(text)
            if t is None:
                return "", "Invalid time format. Use HH:MM or 09:30 AM."
            return f"{t.hour:02d}:{t.minute:02d}", None
        def _compact_build_select_options(options: list[str], current_value: str) -> tuple[list[str], int]:
            current = str(current_value or "").strip()
            opts = [opt for opt in options if str(opt).strip()]
            if current and current not in opts:
                opts = [current] + opts
            opts = [""] + opts
            index = opts.index(current) if current in opts else 0
            return opts, index
        def _apply_compact_card_edit(row_id, patient_name, in_time_str, updates: dict[str, Any]) -> bool:
            df_source = df_raw if "df_raw" in globals() else df_schedule
            if df_source is None or df_source.empty:
                st.warning("No schedule data to update.")
                return False
            df_updated = df_source.copy()
            idx = None
            if row_id and "REMINDER_ROW_ID" in df_updated.columns:
                matches = df_updated["REMINDER_ROW_ID"].astype(str) == str(row_id)
                if matches.any():
                    idx = matches.idxmax()
            if idx is None and "Patient Name" in df_updated.columns and patient_name:
                name_mask = df_updated["Patient Name"].astype(str).str.upper() == str(patient_name).upper()
                if in_time_str and "In Time" in df_updated.columns:
                    time_mask = df_updated["In Time"].astype(str) == str(in_time_str)
                    match = df_updated[name_mask & time_mask]
                else:
                    match = df_updated[name_mask]
                if not match.empty:
                    idx = match.index[0]
            if idx is None:
                st.warning("Unable to locate row for update.")
                return False
            status_col = "STATUS" if "STATUS" in df_updated.columns else "Status" if "Status" in df_updated.columns else ""
            old_status_norm = str(df_updated.at[idx, status_col]).strip().upper() if status_col else ""
            for col, val in updates.items():
                if col in df_updated.columns:
                    df_updated.at[idx, col] = val
            if status_col:
                new_status_norm = str(df_updated.at[idx, status_col]).strip().upper()
                if new_status_norm and new_status_norm != old_status_norm:
                    ts = _now_iso()
                    if "STATUS_CHANGED_AT" in df_updated.columns:
                        df_updated.at[idx, "STATUS_CHANGED_AT"] = ts
                    if ("ONGOING" in new_status_norm or "ON GOING" in new_status_norm) and "ACTUAL_START_AT" in df_updated.columns:
                        if not str(df_updated.at[idx, "ACTUAL_START_AT"]).strip():
                            df_updated.at[idx, "ACTUAL_START_AT"] = ts
                    if ("DONE" in new_status_norm or "COMPLETED" in new_status_norm) and "ACTUAL_END_AT" in df_updated.columns:
                        if not str(df_updated.at[idx, "ACTUAL_END_AT"]).strip():
                            df_updated.at[idx, "ACTUAL_END_AT"] = ts
                    if "STATUS_LOG" in df_updated.columns:
                        existing_log = str(df_updated.at[idx, "STATUS_LOG"])
                        try:
                            df_updated.at[idx, "STATUS_LOG"] = _append_status_log(
                                existing_log,
                                {"at": ts, "from": old_status_norm, "to": new_status_norm},
                            )
                        except Exception:
                            df_updated.at[idx, "STATUS_LOG"] = existing_log
            if bool(st.session_state.get("auto_assign_assistants", True)):
                only_empty = bool(st.session_state.get("auto_assign_only_empty", True))
                _auto_fill_assistants_for_row(df_updated, int(idx), only_fill_empty=only_empty)
            _maybe_save(df_updated, show_toast=False, message=f"Updated {patient_name or 'patient'}")
            if st.session_state.get("auto_save_enabled", False):
                st.toast("Changes saved.", icon="✅")
            else:
                st.toast("Changes queued. Click 'Save Changes'.", icon="📝")
            return True
        def _render_compact_edit_dialog_body() -> None:
            context = st.session_state.get("compact_edit_context") or {}
            if not context:
                _close_compact_edit_dialog()
                return
            row_key = str(context.get("row_key", "")).strip()
            if not row_key:
                _close_compact_edit_dialog()
                return
            lookup_patient = str(context.get("lookup_patient", "") or "")
            lookup_in_time = str(context.get("lookup_in_time", "") or "")
            row_id = str(context.get("row_id", "") or "")
            with st.form(key=f"compact_popup_form_{row_key}"):
                form_left, form_right = st.columns(2, gap="small")
                with form_left:
                    patient_input = st.text_input(
                        "Patient Name",
                        key=f"compact_popup_patient_{row_key}",
                    )
                    with st.container():
                        st.markdown("<div class='time-select-marker'></div>", unsafe_allow_html=True)
                        st.markdown("In Time")
                        in_time_cols = st.columns(3, gap="small")
                        with in_time_cols[0]:
                            in_hour = st.selectbox(
                                "Hour",
                                options=TIME_PICKER_HOURS,
                                key=f"compact_popup_in_hour_{row_key}",
                            )
                        with in_time_cols[1]:
                            in_minute = st.selectbox(
                                "Minute",
                                options=TIME_PICKER_MINUTES,
                                key=f"compact_popup_in_min_{row_key}",
                            )
                        with in_time_cols[2]:
                            in_ampm = st.selectbox(
                                "AM/PM",
                                options=TIME_PICKER_AMPM,
                                key=f"compact_popup_in_ampm_{row_key}",
                            )
                    with st.container():
                        st.markdown("<div class='time-select-marker'></div>", unsafe_allow_html=True)
                        st.markdown("Out Time")
                        out_time_cols = st.columns(3, gap="small")
                        with out_time_cols[0]:
                            out_hour = st.selectbox(
                                "Hour",
                                options=TIME_PICKER_HOURS,
                                key=f"compact_popup_out_hour_{row_key}",
                            )
                        with out_time_cols[1]:
                            out_minute = st.selectbox(
                                "Minute",
                                options=TIME_PICKER_MINUTES,
                                key=f"compact_popup_out_min_{row_key}",
                            )
                        with out_time_cols[2]:
                            out_ampm = st.selectbox(
                                "AM/PM",
                                options=TIME_PICKER_AMPM,
                                key=f"compact_popup_out_ampm_{row_key}",
                            )
                    status_current = st.session_state.get(f"compact_popup_status_{row_key}", "")
                    status_options, status_index = _compact_build_select_options(STATUS_OPTIONS, status_current)
                    status_input = st.selectbox(
                        "Status",
                        options=status_options,
                        index=status_index,
                        key=f"compact_popup_status_{row_key}",
                    )
                with form_right:
                    doctor_current = st.session_state.get(f"compact_popup_doctor_{row_key}", "")
                    doctor_options, doctor_index = _compact_build_select_options(DOCTOR_OPTIONS, doctor_current)
                    doctor_input = st.selectbox(
                        "Doctor",
                        options=doctor_options,
                        index=doctor_index,
                        key=f"compact_popup_doctor_{row_key}",
                    )
                    procedure_input = st.text_input(
                        "Procedure",
                        key=f"compact_popup_procedure_{row_key}",
                    )
                    first_current = st.session_state.get(f"compact_popup_first_{row_key}", "")
                    first_options, first_index = _compact_build_select_options(ASSISTANT_OPTIONS, first_current)
                    first_input = st.selectbox(
                        "First",
                        options=first_options,
                        index=first_index,
                        key=f"compact_popup_first_{row_key}",
                    )
                    second_current = st.session_state.get(f"compact_popup_second_{row_key}", "")
                    second_options, second_index = _compact_build_select_options(ASSISTANT_OPTIONS, second_current)
                    second_input = st.selectbox(
                        "Second",
                        options=second_options,
                        index=second_index,
                        key=f"compact_popup_second_{row_key}",
                    )
                    third_current = st.session_state.get(f"compact_popup_third_{row_key}", "")
                    third_options, third_index = _compact_build_select_options(ASSISTANT_OPTIONS, third_current)
                    third_input = st.selectbox(
                        "Third",
                        options=third_options,
                        index=third_index,
                        key=f"compact_popup_third_{row_key}",
                    )
                flag_cols = st.columns(2, gap="small")
                with flag_cols[0]:
                    case_paper_input = st.checkbox(
                        "QTRAQ",
                        key=f"compact_popup_case_{row_key}",
                    )
                with flag_cols[1]:
                    suction_input = st.checkbox(
                        "Suction",
                        key=f"compact_popup_suction_{row_key}",
                    )
                form_actions = st.columns(2, gap="small")
                with form_actions[0]:
                    save_clicked = st.form_submit_button("Save", use_container_width=True)
                with form_actions[1]:
                    cancel_clicked = st.form_submit_button("Cancel", use_container_width=True)
            if cancel_clicked:
                _close_compact_edit_dialog()
                st.rerun()
            if save_clicked:
                in_norm, in_err = _time_from_picker_parts(in_hour, in_minute, in_ampm)
                out_norm, out_err = _time_from_picker_parts(out_hour, out_minute, out_ampm)
                if in_err or out_err:
                    if in_err:
                        st.error(in_err)
                    if out_err:
                        st.error(out_err)
                else:
                    updates = {
                        "Patient Name": str(patient_input or "").strip(),
                        "In Time": in_norm,
                        "Out Time": out_norm,
                        "Procedure": str(procedure_input or "").strip(),
                        "DR.": str(doctor_input or "").strip(),
                        "Doctor": str(doctor_input or "").strip(),
                        "FIRST": str(first_input or "").strip(),
                        "SECOND": str(second_input or "").strip(),
                        "Third": str(third_input or "").strip(),
                        "THIRD": str(third_input or "").strip(),
                        "CASE PAPER": "Yes" if case_paper_input else "",
                        "SUCTION": bool(suction_input),
                        "STATUS": str(status_input or "").strip(),
                        "Status": str(status_input or "").strip(),
                    }
                    if _apply_compact_card_edit(row_id, lookup_patient, lookup_in_time, updates):
                        _close_compact_edit_dialog()
                        st.rerun()
        _dialog_decorator = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)
        if _dialog_decorator:
            @_dialog_decorator("Edit appointment")
            def _render_compact_edit_dialog() -> None:
                _render_compact_edit_dialog_body()
        else:
            def _render_compact_edit_dialog() -> None:
                st.warning("Popup editing requires a newer Streamlit version.")
                _render_compact_edit_dialog_body()
        def _update_row_status(row_id, patient_name, in_time_str, new_status):
            df_source = df_raw if "df_raw" in globals() else df_schedule
            if df_source is None or df_source.empty:
                st.warning("No schedule data to update.")
                return
            df_updated = df_source.copy()
            idx = None
            if row_id and "REMINDER_ROW_ID" in df_updated.columns:
                matches = df_updated["REMINDER_ROW_ID"].astype(str) == str(row_id)
                if matches.any():
                    idx = matches.idxmax()
            if idx is None and "Patient Name" in df_updated.columns and patient_name:
                name_mask = df_updated["Patient Name"].astype(str).str.upper() == patient_name.upper()
                if in_time_str and "In Time" in df_updated.columns:
                    time_mask = df_updated["In Time"].astype(str) == in_time_str
                    match = df_updated[name_mask & time_mask]
                else:
                    match = df_updated[name_mask]
                if not match.empty:
                    idx = match.index[0]
            if idx is None:
                st.warning("Unable to locate row for update.")
                return
            old_status = ""
            if "STATUS" in df_updated.columns:
                old_status = str(df_updated.at[idx, "STATUS"]).strip().upper()
                df_updated.at[idx, "STATUS"] = new_status
            if "Status" in df_updated.columns:
                if not old_status:
                    old_status = str(df_updated.at[idx, "Status"]).strip().upper()
                df_updated.at[idx, "Status"] = new_status
            ts = _now_iso()
            if "STATUS_CHANGED_AT" in df_updated.columns:
                df_updated.at[idx, "STATUS_CHANGED_AT"] = ts
            if ("ONGOING" in new_status or "ON GOING" in new_status) and "ACTUAL_START_AT" in df_updated.columns:
                if not str(df_updated.at[idx, "ACTUAL_START_AT"]).strip():
                    df_updated.at[idx, "ACTUAL_START_AT"] = ts
            if ("DONE" in new_status or "COMPLETED" in new_status) and "ACTUAL_END_AT" in df_updated.columns:
                if not str(df_updated.at[idx, "ACTUAL_END_AT"]).strip():
                    df_updated.at[idx, "ACTUAL_END_AT"] = ts
            if "STATUS_LOG" in df_updated.columns:
                existing_log = str(df_updated.at[idx, "STATUS_LOG"])
                try:
                    df_updated.at[idx, "STATUS_LOG"] = _append_status_log(
                        existing_log,
                        {"at": ts, "from": old_status, "to": new_status},
                    )
                except Exception:
                    df_updated.at[idx, "STATUS_LOG"] = existing_log
            _maybe_save(df_updated, message=f"Status set to {new_status} for {patient_name}")
            st.toast(f"{patient_name} marked {new_status}", icon="✅")
            st.rerun()
        def _update_row_case_paper(row_id, patient_name, in_time_str, case_checked: bool):
            df_source = df_raw if "df_raw" in globals() else df_schedule
            if df_source is None or df_source.empty:
                st.warning("No schedule data to update.")
                return
            df_updated = df_source.copy()
            idx = None
            if row_id and "REMINDER_ROW_ID" in df_updated.columns:
                matches = df_updated["REMINDER_ROW_ID"].astype(str) == str(row_id)
                if matches.any():
                    idx = matches.idxmax()
            if idx is None and "Patient Name" in df_updated.columns and patient_name:
                name_mask = df_updated["Patient Name"].astype(str).str.upper() == str(patient_name).upper()
                if in_time_str and "In Time" in df_updated.columns:
                    time_mask = df_updated["In Time"].astype(str) == str(in_time_str)
                    match = df_updated[name_mask & time_mask]
                else:
                    match = df_updated[name_mask]
                if not match.empty:
                    idx = match.index[0]
            if idx is None:
                st.warning("Unable to locate row for update.")
                return
            if "CASE PAPER" not in df_updated.columns:
                st.warning("No QTRAQ column to update.")
                return
            df_updated.at[idx, "CASE PAPER"] = "Yes" if case_checked else ""
            _maybe_save(df_updated, message=f"Case paper updated for {patient_name}")
            st.toast(f"{patient_name} case paper updated")
            st.rerun()
        search_value = st.session_state.get("compact_search", "").strip()
        df_cards = df_display.copy()
        if search_value:
            query = search_value.lower()
            mask = pd.Series(False, index=df_cards.index)
            for col in ["Patient Name", "Doctor", "DR.", "Procedure", "FIRST", "SECOND", "Third", "THIRD", "Status"]:
                if col in df_cards.columns:
                    mask = mask | df_cards[col].astype(str).str.lower().str.contains(query, na=False)
            df_cards = df_cards[mask]
        if view_mode == "Table":
            df_table = df_display.drop(columns=["REMINDER_ROW_ID"], errors="ignore")
            # Mark busy assistants in the display
            df_table_with_busy = mark_busy_assistants(df_table)
            st.data_editor(df_table_with_busy, use_container_width=True, height=280, key="compact_schedule_editor")
        else:
            show_case = "CASE PAPER" in df_display.columns
            cards_per_row = 3
            # Mark busy assistants in card view
            df_cards_marked = mark_busy_assistants(df_cards)
            card_rows = [
                df_cards_marked.iloc[i:i + cards_per_row]
                for i in range(0, len(df_cards_marked), cards_per_row)
            ]
            if df_cards_marked.empty:
                st.info("No patients found.")
            for chunk in card_rows:
                cols = st.columns(len(chunk), gap="small")
                for col, (row_index, row) in zip(cols, chunk.iterrows()):
                    patient = _clean_text(row.get("Patient Name"))
                    doctor = _clean_text(row.get("Doctor") or row.get("DR."))
                    procedure = _clean_text(row.get("Procedure"))
                    in_time = _clean_text(row.get("In Time") or row.get("In Time Str"))
                    out_time = _clean_text(row.get("Out Time") or row.get("Out Time Str"))
                    status = _clean_text(row.get("Status") or row.get("STATUS") or "WAITING")
                    row_id = _clean_text(row.get("REMINDER_ROW_ID"))
                    staff = [
                        _clean_text(row.get("FIRST")),
                        _clean_text(row.get("SECOND")),
                        _clean_text(row.get("Third") or row.get("THIRD")),
                    ]
                    staff = [name for name in staff if name]
                    time_text = " - ".join([t for t in [in_time, out_time] if t])
                    status_text = (status or "WAITING").strip().upper()
                    if not status_text:
                        status_text = "WAITING"
                    status_class = _status_class(status_text)
                    staff_html = " &bull; ".join(html.escape(name) for name in staff) if staff else "Unassigned"
                    doctor_icon_svg = '<svg class="info-icon-svg" viewBox="0 0 24 24" width="20" height="20"><path d="M19 8h-1.26c-.19-.73-.48-1.42-.85-2.06l.94-.94a.996.996 0 0 0 0-1.41l-1.41-1.41a.996.996 0 0 0-1.41 0l-.94.94c-.64-.37-1.33-.66-2.06-.85V1c0-.55-.45-1-1-1H9c-.55 0-1 .45-1 1v1.26c-.73.19-1.42.48-2.06.85l-.94-.94a.996.996 0 0 0-1.41 0L2.18 3.58a.996.996 0 0 0 0 1.41l.94.94c-.37.64-.66 1.33-.85 2.06H1c-.55 0-1 .45-1 1v2c0 .55.45 1 1 1h1.26c.19.73.48 1.42.85 2.06l-.94.94a.996.996 0 0 0 0 1.41l1.41 1.41c.39.39 1.02.39 1.41 0l.94-.94c.64.37 1.33.66 2.06.85V23c0 .55.45 1 1 1h2c.55 0 1-.45 1-1v-1.26c.73-.19 1.42-.48 2.06-.85l.94.94c.39.39 1.02.39 1.41 0l1.41-1.41a.996.996 0 0 0 0-1.41l-.94-.94c.37-.64.66-1.33.85-2.06H19c.55 0 1-.45 1-1V9c0-.55-.45-1-1-1zm-8 8c-1.66 0-3-1.34-3-3s1.34-3 3-3 3 1.34 3 3-1.34 3-3 3z" fill="currentColor"/></svg>'
                    staff_icon_svg = '<svg class="info-icon-svg" viewBox="0 0 24 24" width="20" height="20"><path d="M16 11c1.66 0 2.99-1.34 2.99-3S17.66 5 16 5c-1.66 0-3 1.34-3 3s1.34 3 3 3zm-8 0c1.66 0 2.99-1.34 2.99-3S9.66 5 8 5C6.34 5 5 6.34 5 8s1.34 3 3 3zm0 2c-2.33 0-7 1.17-7 3.5V19h14v-2.5c0-2.33-4.67-3.5-7-3.5zm8 0c-.29 0-.62.02-.97.05 1.16.84 1.97 1.97 1.97 3.45V19h6v-2.5c0-2.33-4.67-3.5-7-3.5z" fill="currentColor"/></svg>'
                    doctor_line = (
                        f"<div class='info-row'><span class='info-icon doctor-icon'>{doctor_icon_svg}</span><span class='info-text'>{html.escape(doctor)}</span></div>"
                        if doctor
                        else ""
                    )
                    staff_line = f"<div class='info-row'><span class='info-icon staff-icon'>{staff_icon_svg}</span><span class='info-text'>{staff_html}</span></div>"
                    row_key = row_id if row_id else f"compact_{row_index}"
                    with col:
                        with st.container(border=True):
                            st.markdown("<div class='card-shell-marker'></div>", unsafe_allow_html=True)
                            st.markdown(
                                _normalize_html(
                                    f"""
                                    <div class="card-inner">
                                        <div class="card-status-banner {status_class}">
                                            <span class="status-dot"></span>
                                            <span class="status-text">{html.escape(status_text)}</span>
                                        </div>
                                        <div class="card-head">
                                            <div class="card-avatar">{html.escape(_initials(patient))}</div>
                                            <div class="card-title">
                                                <div class="card-name">{html.escape(patient) if patient else "Unknown"}</div>
                                                <div class="card-time">{html.escape(time_text) if time_text else "--"}</div>
                                            </div>
                                        </div>
                                        <div class="card-subdivider"></div>
                                        <div class="card-info">
                                            {doctor_line}
                                            {staff_line}
                                        </div>
                                        <div class="card-divider"></div>
                                    </div>
                                    """
                                ),
                                unsafe_allow_html=True,
                            )
                            st.markdown("<div class='card-actions-row'>", unsafe_allow_html=True)
                            if show_case:
                                row_cols = st.columns([1, 1.15, 1.15, 1.15], gap="small")
                                with row_cols[0]:
                                    case_active = _truthy(row.get("CASE PAPER"))
                                    case_checked = st.checkbox(
                                        "QTRAQ",
                                        value=case_active,
                                        key=f"card_case_{row_key}",
                                    )
                                    if case_checked != case_active:
                                        _update_row_case_paper(row_id, patient, in_time, case_checked)
                                with row_cols[1]:
                                    st.markdown("<div class='card-action-marker card-action-done'></div>", unsafe_allow_html=True)
                                    if st.button("✓ Done", key=f"card_done_{row_key}", use_container_width=True, type="primary"):
                                        _update_row_status(row_id, patient, in_time, "DONE")
                                with row_cols[2]:
                                    st.markdown("<div class='card-action-marker card-action-edit'></div>", unsafe_allow_html=True)
                                    st.button(
                                        "✎ Edit",
                                        key=f"card_edit_{row_key}",
                                        on_click=_open_compact_edit_dialog,
                                        args=(
                                            {
                                                "row_key": row_key,
                                                "row_id": row_id,
                                                "lookup_patient": patient,
                                                "lookup_in_time": in_time,
                                                "patient": patient,
                                                "in_time": in_time,
                                                "out_time": out_time,
                                                "doctor": doctor,
                                                "procedure": procedure,
                                                "status": status,
                                                "staff_first": _clean_text(row.get("FIRST")),
                                                "staff_second": _clean_text(row.get("SECOND")),
                                                "staff_third": _clean_text(row.get("Third") or row.get("THIRD")),
                                                "case_paper": _truthy(row.get("CASE PAPER")),
                                                "suction": _truthy(row.get("SUCTION")),
                                            },
                                        ),
                                        use_container_width=True,
                                        type="secondary",
                                    )
                                with row_cols[3]:
                                    st.markdown("<div class='card-action-marker card-action-cancel'></div>", unsafe_allow_html=True)
                                    if st.button("✕ Cancel", key=f"card_cancel_{row_key}", use_container_width=True, type="secondary"):
                                        _update_row_status(row_id, patient, in_time, "CANCELLED")
                            else:
                                action_cols = st.columns([1.15, 1.15, 1.15], gap="small")
                                with action_cols[0]:
                                    st.markdown("<div class='card-action-marker card-action-done'></div>", unsafe_allow_html=True)
                                    if st.button("✓ Done", key=f"card_done_{row_key}", use_container_width=True, type="primary"):
                                        _update_row_status(row_id, patient, in_time, "DONE")
                                with action_cols[1]:
                                    st.markdown("<div class='card-action-marker card-action-edit'></div>", unsafe_allow_html=True)
                                    st.button(
                                        "✎ Edit",
                                        key=f"card_edit_{row_key}",
                                        on_click=_open_compact_edit_dialog,
                                        args=(
                                            {
                                                "row_key": row_key,
                                                "row_id": row_id,
                                                "lookup_patient": patient,
                                                "lookup_in_time": in_time,
                                                "patient": patient,
                                                "in_time": in_time,
                                                "out_time": out_time,
                                                "doctor": doctor,
                                                "procedure": procedure,
                                                "status": status,
                                                "staff_first": _clean_text(row.get("FIRST")),
                                                "staff_second": _clean_text(row.get("SECOND")),
                                                "staff_third": _clean_text(row.get("Third") or row.get("THIRD")),
                                                "case_paper": _truthy(row.get("CASE PAPER")),
                                                "suction": _truthy(row.get("SUCTION")),
                                            },
                                        ),
                                        use_container_width=True,
                                        type="secondary",
                                    )
                                with action_cols[2]:
                                    st.markdown("<div class='card-action-marker card-action-cancel'></div>", unsafe_allow_html=True)
                                    if st.button("✕ Cancel", key=f"card_cancel_{row_key}", use_container_width=True, type="secondary"):
                                        _update_row_status(row_id, patient, in_time, "CANCELLED")
                            st.markdown("</div>", unsafe_allow_html=True)
                            st.markdown("<div class='card-details-row'>", unsafe_allow_html=True)
                            with st.expander("View Details", expanded=False):
                                st.markdown(f"**Doctor:** {doctor or '--'}")
                                st.markdown(f"**Procedure:** {procedure or '--'}")
                                st.markdown(f"**Staff:** {', '.join(staff) if staff else 'Unassigned'}")
                                st.markdown(f"**Status:** {status}")
                                if show_case:
                                    st.markdown(f"**QTRAQ:** {'Yes' if _truthy(row.get('CASE PAPER')) else 'No'}")
                            st.markdown("</div>", unsafe_allow_html=True)
            if st.session_state.get("compact_edit_open"):
                _render_compact_edit_dialog()
        st.markdown("<div class='summary-bar'>", unsafe_allow_html=True)
        with st.expander("📊 Schedule Summary by Doctor", expanded=False):
            st.write("Summary table / charts here")
        st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)
# Global save-mode flags
if "auto_save_enabled" not in st.session_state:
    st.session_state.auto_save_enabled = True
if "pending_changes" not in st.session_state:
    st.session_state.pending_changes = False
if "pending_changes_reason" not in st.session_state:
    st.session_state.pending_changes_reason = ""
if "unsaved_df" not in st.session_state:
    st.session_state.unsaved_df = None
if "save_debounce_seconds" not in st.session_state:
    st.session_state.save_debounce_seconds = 0
if "last_save_at" not in st.session_state:
    st.session_state.last_save_at = 0.0
if "last_saved_hash" not in st.session_state:
    st.session_state.last_saved_hash = None
if "loaded_save_version" not in st.session_state:
    st.session_state.loaded_save_version = None
if "loaded_save_at" not in st.session_state:
    st.session_state.loaded_save_at = None
if "enable_conflict_checks" not in st.session_state:
    # Disable conflict checking when using Supabase (migrated data has version mismatches)
    st.session_state.enable_conflict_checks = False if USE_SUPABASE else True
if "save_conflict" not in st.session_state:
    st.session_state.save_conflict = None
else:
    # Clear any existing conflict state on fresh load (for Supabase migration)
    if USE_SUPABASE:
        st.session_state.save_conflict = None
if "is_saving" not in st.session_state:
    st.session_state.is_saving = False
if "unsaved_df_version" not in st.session_state:
    st.session_state.unsaved_df_version = 0
if "supabase_ready" not in st.session_state:
    st.session_state.supabase_ready = False
if "supabase_ready_at" not in st.session_state:
    st.session_state.supabase_ready_at = 0.0
if "supabase_profiles_seeded" not in st.session_state:
    st.session_state.supabase_profiles_seeded = False
if "supabase_staff_refreshed" not in st.session_state:
    st.session_state.supabase_staff_refreshed = False
if "profiles_cache_bust" not in st.session_state:
    st.session_state.profiles_cache_bust = 0
if "active_duty_run_id" not in st.session_state:
    st.session_state.active_duty_run_id = None
if "active_duty_due_at" not in st.session_state:
    st.session_state.active_duty_due_at = None
if "active_duty_started_at" not in st.session_state:
    st.session_state.active_duty_started_at = None
if "active_duty_est_minutes" not in st.session_state:
    st.session_state.active_duty_est_minutes = None
if "duty_current_assistant" not in st.session_state:
    st.session_state.duty_current_assistant = ""
# ===== COLOR CUSTOMIZATION SECTION =====
# REMOVED DUPLICATE: COLORS and WEEKLY_OFF now defined at top of file
# Custom CSS with customizable colors
# ================= PROFILE INTEGRATION WITH SCHEDULE =================
# Load assistant and doctor names for dropdowns
st.markdown(
    f"""
    <style>
    :root {{
        --bg-primary: {COLORS['bg_primary']};
        --bg-secondary: {COLORS['bg_secondary']};
        --text-primary: {COLORS['text_primary']};
        --text-secondary: {COLORS['text_secondary']};
        --accent: {COLORS['accent']};
        --success: {COLORS['success']};
        --warning: {COLORS['warning']};
        --danger: {COLORS['danger']};
        --info: {COLORS['info']};
        --glass-bg: {COLORS['glass_bg']};
        --glass-border: {COLORS['glass_border']};
    }}
    
    * {{
        margin: 0;
        padding: 0;
    }}
    
    body, .stApp {{
        background: linear-gradient(135deg, #f8fafc 0%, #e0f2fe 100%) !important;
        color: var(--text-primary) !important;
        font-family: 'Inter', sans-serif;
    }}
    /* Tighten gap below sticky header */
    .block-container {{
        padding-top: 0.25rem !important;
        padding-bottom: 0.5rem !important;
    }}
    /* Compact headings & metrics */
    h1, h2, h3 {{
        margin: 0.15rem 0 0.35rem 0 !important;
    }}
    div[data-testid="stMetric"] {{
        padding: 0.6rem 0.8rem !important;
        border-radius: 14px;
    }}
    /* Glass Cards & Tables */
    .stDataFrame, .stTable, [data-testid="stDataFrameResizable"], [data-testid="stTable"] {{
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.7), rgba(255, 255, 255, 0.3)) !important;
        border: 1px solid rgba(255, 255, 255, 0.18) !important;
        box-shadow: 0 8px 32px 0 rgba(37, 99, 235, 0.12);
        border-radius: 20px;
        backdrop-filter: blur(20px) !important;
        -webkit-backdrop-filter: blur(20px) !important;
    }}
    .stDataFrame table, .stTable table {{
        background: var(--glass-bg) !important;
    }}
    
    header {{
        background-color: {COLORS['bg_primary']} !important;
        border-bottom: none !important;
        padding: 1rem 0 !important;
    }}
    
    [data-testid="stHeader"] {{
        background-color: {COLORS['bg_primary']} !important;
    }}
    /* Hide GitHub/logo link in Streamlit header (Streamlit Cloud toolbar) */
    [data-testid="stToolbar"] a[href*="github.com"],
    [data-testid="stToolbar"] a[aria-label*="View source"],
    [data-testid="stToolbar"] a[title*="View source"],
    [data-testid="stToolbar"] a[aria-label*="GitHub"],
    [data-testid="stToolbar"] a[title*="GitHub"],
    [data-testid="stToolbar"] button[aria-label*="View source"],
    [data-testid="stToolbar"] button[title*="View source"] {{
        display: none !important;
    }}
    [data-testid="stToolbarActions"],
    [data-testid="stMainMenu"] {{
        display: none !important;
        visibility: hidden !important;
    }}
    
    /* Professional main container */
    .main {{
        padding: 2rem 3rem !important;
        max-width: 2200px !important;
        margin: 0 auto !important;
    }}
    html, body, [data-testid="stAppViewContainer"] {{
        overflow-y: auto !important;
        height: auto !important;
    }}
    [data-testid="stAppViewContainer"] > .main {{
        overflow: visible !important;
    }}
    
    /* Professional header styling */
    .header-container {{
        display: flex;
        align-items: center;
        justify-content: center;
        gap: 1.5rem;
        padding: 2rem 0;
        border-bottom: none;
    }}
    
    .header-logo {{
        width: 80px;
        height: auto;
    }}
    
    .header-title {{
        font-size: 2rem;
        font-weight: 700;
        color: {COLORS['text_primary']};
        letter-spacing: 0.5px;
    }}
    
    .st-bw, .st-cq, .st-dx, .stDataFrame, .stDataFrame th, .stDataFrame td {{
        background-color: {COLORS['bg_secondary']} !important;
        color: {COLORS['text_primary']} !important;
    }}
    
    /* Premium Status-based row background colors with dynamic effects */
    /* Upcoming rows - Light blue */
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("WAITING")) {{
        background: rgba(217, 197, 178, 0.35) !important;
        border-left: 5px solid {COLORS['info']} !important;
    }}
    
    /* Ongoing rows - Light green */
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("ON GOING")) {{
        background: rgba(52, 49, 45, 0.12) !important;
        border-left: 5px solid {COLORS['success']} !important;
    }}
    
    /* Arrived rows - Light yellow */
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("ARRIVED")) {{
        background: rgba(217, 197, 178, 0.45) !important;
        border-left: 5px solid {COLORS['warning']} !important;
    }}
    /* Shifted rows - Yellow */
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("SHIFTED")) {{
        background: rgba(217, 197, 178, 0.45) !important;
        border-left: 5px solid {COLORS['warning']} !important;
    }}
    
    /* Cancelled rows - Light red */
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("CANCELLED")) {{
        background: rgba(126, 127, 131, 0.25) !important;
        border-left: 5px solid {COLORS['danger']} !important;
    }}
    
    /* Enhanced Hover effect with shadow lift */
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("WAITING")):hover {{
        background: rgba(217, 197, 178, 0.45) !important;
        box-shadow: 0 4px 12px rgba(52, 49, 45, 0.18) inset !important;
    }}
    
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("ON GOING")):hover {{
        background: rgba(52, 49, 45, 0.18) !important;
        box-shadow: 0 4px 12px rgba(20, 17, 15, 0.18) inset !important;
    }}
    
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("ARRIVED")):hover {{
        background: rgba(217, 197, 178, 0.55) !important;
        box-shadow: 0 4px 12px rgba(52, 49, 45, 0.18) inset !important;
    }}
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("SHIFTED")):hover {{
        background: rgba(217, 197, 178, 0.55) !important;
        box-shadow: 0 4px 12px rgba(52, 49, 45, 0.18) inset !important;
    }}
    
    [data-testid="stDataFrameContainer"] tbody tr:has(td:contains("CANCELLED")):hover {{
        background: rgba(126, 127, 131, 0.35) !important;
        box-shadow: 0 4px 12px rgba(126, 127, 131, 0.25) inset !important;
    }}
    
    /* Table Header Styling - Premium & Elegant */
    [data-testid="stDataFrameContainer"] thead {{
        background: {COLORS['button_bg']} !important;
        border-bottom: 1px solid var(--glass-border) !important;
        box-shadow: 0 6px 18px rgba(20, 17, 15, 0.28) !important;
    }}
    
    [data-testid="stDataFrameContainer"] thead th {{
        color: {COLORS['button_text']} !important;
        font-weight: 800 !important;
        padding: 18px 16px !important;
        text-align: center !important;
        font-size: 0.99rem !important;
        letter-spacing: 1px !important;
        text-transform: uppercase !important;
        background: {COLORS['button_bg']} !important;
        position: relative !important;
        text-shadow: 0 2px 4px rgba(20, 17, 15, 0.3) !important;
        box-shadow: inset 0 1px 0 rgba(243, 243, 244, 0.18) !important;
        transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1) !important;
        border-right: 1px solid rgba(243, 243, 244, 0.22) !important;
    }}
    
    [data-testid="stDataFrameContainer"] thead th:last-child {{
        border-right: none !important;
    }}
    
    [data-testid="stDataFrameContainer"] thead th:hover {{
        filter: brightness(1.08) !important;
        transform: translateY(-2px) !important;
        box-shadow: inset 0 1px 0 rgba(243, 243, 244, 0.10), 0 10px 22px rgba(20, 17, 15, 0.22) !important;
    }}
    
    /* Premium Table Rows */
    [data-testid="stDataFrameContainer"] tbody tr {{
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1) !important;
        border-radius: 0 !important;
        position: relative !important;
    }}
    
    /* Alternating row background for better readability */
    [data-testid="stDataFrameContainer"] tbody tr:nth-child(even) {{
        background-color: rgba(243, 243, 244, 0.04) !important;
    }}
    
    [data-testid="stDataFrameContainer"] tbody tr:hover {{
        background-color: rgba(243, 243, 244, 0.06) !important;
        box-shadow: 0 2px 14px rgba(20, 17, 15, 0.22) inset !important;
    }}
    
    /* Premium Table Cells */
    [data-testid="stDataFrameContainer"] tbody td {{
        padding: 12px 14px !important;
        border-bottom: 1px solid rgba(217, 197, 178, 0.55) !important;
        border-right: 1px solid rgba(217, 197, 178, 0.35) !important;
        font-size: 0.93rem !important;
        line-height: 1.25 !important;
        vertical-align: middle !important;
        transition: all 0.2s ease !important;
        position: relative !important;
    }}
    [data-testid="stDataFrameContainer"] tbody td:last-child {{
        border-right: none !important;
    }}
    
    /* Dropdown and Select Styling (scoped to main content, avoid sidebar) */
    .main [data-baseweb="select"] {{
        background-color: {COLORS['bg_secondary']} !important;
        border-radius: 6px !important;
    }}
    /* Premium dialog styling */
    div[role="dialog"] > div,
    div[data-testid="stDialog"] {{
        background: linear-gradient(180deg, #ffffff 0%, #f3f6fb 100%);
        border: 1px solid #d7dce5;
        border-radius: 18px;
        box-shadow: 0 24px 48px rgba(20, 17, 15, 0.25);
    }}
    div[role="dialog"] h2,
    div[data-testid="stDialog"] h2 {{
        font-size: 20px;
        font-weight: 800;
        letter-spacing: 0.6px;
        color: #2a2d33;
        margin-bottom: 12px;
    }}
    div[role="dialog"] form,
    div[data-testid="stDialog"] form {{
        background: #ffffff;
        border: 1px solid #e0e4eb;
        border-radius: 16px;
        padding: 14px 16px 12px;
        box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.6), 0 12px 26px rgba(20, 17, 15, 0.08);
    }}
    div[role="dialog"] label,
    div[data-testid="stDialog"] label {{
        font-size: 12px;
        font-weight: 700;
        letter-spacing: 0.5px;
        text-transform: uppercase;
        color: #5b6470;
    }}
    div[role="dialog"] [data-testid="stHorizontalBlock"],
    div[data-testid="stDialog"] [data-testid="stHorizontalBlock"] {{
        gap: 0.75rem;
    }}
    div[role="dialog"] [data-baseweb="input"] > div,
    div[data-testid="stDialog"] [data-baseweb="input"] > div,
    div[role="dialog"] [data-baseweb="select"] > div,
    div[data-testid="stDialog"] [data-baseweb="select"] > div {{
        background: #f8fafc;
        border: 1px solid #d7dce5;
        border-radius: 10px;
        box-shadow: inset 0 1px 2px rgba(20, 17, 15, 0.06);
    }}
    div[role="dialog"] [data-baseweb="input"] > div:hover,
    div[data-testid="stDialog"] [data-baseweb="input"] > div:hover,
    div[role="dialog"] [data-baseweb="select"] > div:hover,
    div[data-testid="stDialog"] [data-baseweb="select"] > div:hover {{
        border-color: #c0c8d4;
    }}
    div[role="dialog"] [data-baseweb="input"] > div:focus-within,
    div[data-testid="stDialog"] [data-baseweb="input"] > div:focus-within,
    div[role="dialog"] [data-baseweb="select"] > div:focus-within,
    div[data-testid="stDialog"] [data-baseweb="select"] > div:focus-within {{
        border-color: #2f63e8;
        box-shadow: 0 0 0 3px rgba(47, 99, 232, 0.16);
    }}
    .time-select-marker {{display:none;}}
    div[role="dialog"] div[data-testid="stVerticalBlock"]:has(.time-select-marker) [data-baseweb="select"] svg,
    div[role="dialog"] div[data-testid="stVerticalBlock"]:has(.time-select-marker) [data-baseweb="select"] span[role="img"],
    div[data-testid="stDialog"] div[data-testid="stVerticalBlock"]:has(.time-select-marker) [data-baseweb="select"] svg,
    div[data-testid="stDialog"] div[data-testid="stVerticalBlock"]:has(.time-select-marker) [data-baseweb="select"] span[role="img"] {{
        display: none;
    }}
    div[role="dialog"] button[kind],
    div[data-testid="stDialog"] button[kind] {{
        height: 38px !important;
        border-radius: 12px !important;
        font-weight: 700;
        letter-spacing: 0.4px;
        box-shadow: 0 10px 20px rgba(20, 17, 15, 0.12) !important;
    }}
    div[role="dialog"] button[kind="primary"],
    div[data-testid="stDialog"] button[kind="primary"] {{
        background: #2f63e8 !important;
        border: 1px solid #2f63e8 !important;
        color: #ffffff !important;
    }}
    div[role="dialog"] button[kind="secondary"],
    div[data-testid="stDialog"] button[kind="secondary"] {{
        background: #ffffff !important;
        border: 1px solid #d7dce5 !important;
        color: #3d424a !important;
    }}
    /* Summary chips */
    .summary-row {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
        gap: 12px;
        margin: 8px 0 18px 0;
    }}
    .summary-chip {{
        padding: 12px 14px;
        border-radius: 14px;
        background: var(--glass-bg);
        border: 1px solid var(--glass-border);
        box-shadow: 0 10px 26px rgba(20,17,15,0.12);
        backdrop-filter: none !important;
        display: flex;
        flex-direction: column;
        gap: 6px;
    }}
    .summary-chip .label {{
        font-size: 12px;
        letter-spacing: 0.4px;
        text-transform: uppercase;
        color: var(--text-secondary);
        opacity: 0.8;
    }}
    .summary-chip .value {{
        font-size: 22px;
        font-weight: 800;
        color: var(--text-primary);
        line-height: 1.1;
    }}
    .summary-chip.success {{ border-color: {COLORS['success']}; }}
    .summary-chip.warning {{ border-color: {COLORS['warning']}; }}
    .summary-chip.danger {{ border-color: {COLORS['danger']}; }}
    .summary-chip.info {{ border-color: {COLORS['info']}; }}
    .summary-chip.secondary {{ border-color: {COLORS['accent']}; }}
    
    .main [data-baseweb="select"] button {{
        color: {COLORS['text_primary']} !important;
        background-color: {COLORS['bg_secondary']} !important;
        border: 1px solid #3b82f6 !important;
        border-radius: 6px !important;
        transition: all 0.2s ease !important;
    }}
    
    .main [data-baseweb="select"] button:hover {{
        border-color: {COLORS['button_bg']} !important;
        box-shadow: 0 2px 8px rgba(20, 17, 15, 0.22) !important;
    }}
    
    .main [data-baseweb="select"] button span {{
        color: {COLORS['text_primary']} !important;
    }}
    
    [data-baseweb="popover"] {{
        background-color: {COLORS['bg_secondary']} !important;
        border-radius: 8px !important;
        box-shadow: 0 4px 12px rgba(20, 17, 15, 0.15) !important;
    }}
    
    [data-baseweb="menu"] {{
        background-color: {COLORS['bg_secondary']} !important;
        border-radius: 8px !important;
    }}
    
    [data-baseweb="menu"] li {{
        color: {COLORS['text_primary']} !important;
        background-color: {COLORS['bg_secondary']} !important;
        padding: 8px 12px !important;
    }}
    
    [data-baseweb="menu"] li:hover {{
        background-color: {COLORS['button_bg']} !important;
        color: {COLORS['button_text']} !important;
    }}
    
    [role="option"] {{
        color: {COLORS['text_primary']} !important;
        background-color: {COLORS['bg_secondary']} !important;
        padding: 8px 12px !important;
    }}
    
    [role="option"]:hover {{
        background-color: {COLORS['button_bg']} !important;
        color: {COLORS['button_text']} !important;
    }}
    
    [role="listbox"] {{
        background: var(--glass-bg) !important;
        border-radius: 10px !important;
        border: 1px solid var(--glass-border) !important;
        backdrop-filter: none !important;
        -webkit-backdrop-filter: none !important;
    }}
    
    /* Data editor dropdown text visibility */
    div[data-testid="stDataFrameContainer"] [role="button"] {{
        color: {COLORS['text_primary']} !important;
    }}
    
    div[data-testid="stDataFrameContainer"] [role="option"] {{
        color: {COLORS['text_primary']} !important;
        background-color: {COLORS['bg_secondary']} !important;
    }}
    
    div[data-testid="stDataFrameContainer"] [role="option"]:hover {{
        background-color: {COLORS['button_bg']} !important;
        color: {COLORS['button_text']} !important;
    }}
    
    /* Button Styling - Premium & Attractive */
    .stButton>button {{
        background: {COLORS['button_bg']} !important;
        color: {COLORS['button_text']} !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        font-size: 0.85rem !important;
        padding: 8px 16px !important;
        transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1) !important;
        box-shadow: 0 4px 18px rgba(20, 17, 15, 0.25) !important;
        letter-spacing: 0.3px !important;
        text-transform: uppercase !important;
        cursor: pointer !important;
    }}
    
    .stButton>button:hover {{
        background: {COLORS['button_bg']} !important;
        transform: translateY(-4px) !important;
        box-shadow: 0 10px 28px rgba(20, 17, 15, 0.32) !important;
        letter-spacing: 1px !important;
    }}
    
    .stButton>button:active {{
        transform: translateY(-1px) !important;
        box-shadow: 0 2px 10px rgba(20, 17, 15, 0.28) !important;
    }}
    
    .stButton>button:focus {{
        outline: none !important;
        box-shadow: 0 0 0 3px {COLORS['button_bg']} !important;
    }}
    /* Targeted hover animation: Add Patient + Save only (via unique tooltip/title) */
    button[title="Add a new patient row (uses selected patient if chosen)"] {{
        position: relative !important;
        overflow: hidden !important;
        background: {COLORS['button_bg']} !important;
        color: {COLORS['button_text']} !important;
    }}
    button[title="Save changes to storage"] {{
        position: relative !important;
        overflow: hidden !important;
        background: {COLORS['button_bg']} !important;
        color: {COLORS['button_text']} !important;
    }}
    button[title="Add a new patient row (uses selected patient if chosen)"]:hover,
    button[title="Save changes to storage"]:hover {{
        background: {COLORS['text_primary']} !important;
        color: {COLORS['button_text']} !important;
        animation: pulse-glow 1.4s ease-out infinite !important;
    }}
    button[title="Add a new patient row (uses selected patient if chosen)"]:active,
    button[title="Save changes to storage"]:active {{
        background: {COLORS['text_primary']} !important;
        color: {COLORS['button_text']} !important;
    }}
    
    .st-bv, .st-cv, .st-cw {{
        background-color: {COLORS['bg_secondary']} !important;
    }}
    
    h1, h2, h3, h4, h5, h6 {{
        color: {COLORS['text_primary']} !important;
        font-weight: 600 !important;
    }}
    
    h1 {{
        font-size: 2rem !important;
        margin-bottom: 1.5rem !important;
    }}
    
    h2 {{
        font-size: 1.5rem !important;
        margin-bottom: 1rem !important;
        margin-top: 1.5rem !important;
    }}
    
    .stMarkdown {{
        color: {COLORS['text_primary']} !important;
    }}
    
    /* Data Frame Container - Premium & Beautiful */
    [data-testid="stDataFrameContainer"] {{
        background: var(--glass-bg) !important;
        border-radius: 14px !important;
        border: 1px solid var(--glass-border) !important;
        box-shadow: 0 14px 40px rgba(20, 17, 15, 0.30) !important;
        overflow: hidden !important;
        transition: all 0.3s ease !important;
        backdrop-filter: none !important;
        -webkit-backdrop-filter: none !important;
    }}
    
    [data-testid="stDataFrameContainer"]:hover {{
        box-shadow: 0 18px 52px rgba(20, 17, 15, 0.36) !important;
    }}
    
    /* Tabs Styling */
    .stTabs [data-baseweb="tab-list"] {{
        background-color: transparent !important;
        border-bottom: 1px solid var(--glass-border) !important;
    }}
    
    .stTabs [data-baseweb="tab"] {{
        color: {COLORS['text_secondary']} !important;
        padding: 12px 20px !important;
        border-bottom: 3px solid transparent !important;
        transition: all 0.3s ease !important;
    }}
    
    .stTabs [data-baseweb="tab"]:hover {{
        color: {COLORS['button_bg']} !important;
    }}
    
    .stTabs [aria-selected="true"] {{
        color: {COLORS['button_bg']} !important;
        border-bottom: 3px solid {COLORS['button_bg']} !important;
    }}
    
    /* Alert/Message Styling */
    .st-info {{
        background-color: rgba(217, 197, 178, 0.18) !important;
        border-left: 4px solid {COLORS['info']} !important;
        border-radius: 6px !important;
        padding: 12px 16px !important;
    }}
    
    .st-success {{
        background-color: rgba(52, 49, 45, 0.12) !important;
        border-left: 4px solid {COLORS['success']} !important;
        border-radius: 6px !important;
        padding: 12px 16px !important;
    }}
    
    .st-warning {{
        background-color: rgba(217, 197, 178, 0.22) !important;
        border-left: 4px solid {COLORS['warning']} !important;
        border-radius: 6px !important;
        padding: 12px 16px !important;
    }}
    
    .st-error {{
        background-color: rgba(126, 127, 131, 0.18) !important;
        border-left: 4px solid {COLORS['danger']} !important;
        border-radius: 6px !important;
        padding: 12px 16px !important;
    }}
    
    /* Animations */
    @keyframes bounce-click {{
        0% {{ transform: scale(1); }}
        50% {{ transform: scale(1.2); }}
        100% {{ transform: scale(1); }}
    }}
    
    @keyframes pulse-glow {{
        0% {{ box-shadow: 0 0 0 0 rgba(52, 49, 45, 0.7); }}
        70% {{ box-shadow: 0 0 0 10px rgba(52, 49, 45, 0); }}
        100% {{ box-shadow: 0 0 0 0 rgba(52, 49, 45, 0); }}
    }}
    
    @keyframes spin-check {{
        0% {{ transform: rotate(-10deg) scale(0.8); }}
        50% {{ transform: rotate(5deg) scale(1.1); }}
        100% {{ transform: rotate(0deg) scale(1); }}
    }}
    
    /* Checkbox Styling */
    /* IMPORTANT: Scope checkbox styling to the data editor only.
       Streamlit sidebar widgets use BaseWeb components that also rely on
       checkbox inputs; global overrides can make them appear "frozen".
    */
    [data-testid="stDataFrameContainer"] input[type="checkbox"] {{
        width: 20px !important;
        height: 20px !important;
        cursor: pointer !important;
        transition: transform 140ms ease, filter 0.3s ease !important;
        accent-color: #2563eb !important;
    }}
    /* Keyboard focus for table checkboxes */
    [data-testid="stDataFrameContainer"] input[type="checkbox"]:focus-visible {{
        outline: 2px solid var(--text-secondary) !important;
        outline-offset: 3px !important;
        border-radius: 4px !important;
    }}
    [data-testid="stDataFrameContainer"] input[type="checkbox"]:active {{
        animation: bounce-click 0.4s ease !important;
    }}
    [data-testid="stDataFrameContainer"] input[type="checkbox"]:checked {{
        animation: tdb-native-checkbox-pop 160ms ease-out, spin-check 0.5s ease !important;
    }}
    @keyframes tdb-native-checkbox-pop {{
        0% {{ transform: scale(1); }}
        60% {{ transform: scale(1.12); }}
        100% {{ transform: scale(1.06); }}
    }}
    /* Streamlit (BaseWeb) checkbox animation (for st.checkbox, sidebar toggles, etc.)
       - Adds a slight "pop" on check
       - Animates the SVG checkmark stroke so it draws left-to-right
       - Keeps keyboard accessibility via :focus-visible
       NOTE: This targets BaseWeb checkbox markup and does NOT touch the data editor's native inputs.
    */
    /* BaseWeb checkbox SVG typically contains multiple paths (box + tick).
       Target the tick specifically (usually the 2nd path) + any polyline tick.
    */
    div[data-baseweb="checkbox"] svg path:nth-of-type(2),
    div[data-baseweb="checkbox"] svg polyline {{
        fill: none !important;
        stroke: var(--text-secondary) !important;
        /* Large dash length so the tick fully hides/shows regardless of path length */
        stroke-dasharray: 1000;
        stroke-dashoffset: 1000;
        transition: stroke-dashoffset 220ms ease;
    }}
    /* Draw the tick when checked */
    div[data-baseweb="checkbox"]:has(input[type="checkbox"]:checked) svg path:nth-of-type(2),
    div[data-baseweb="checkbox"]:has(input[type="checkbox"]:checked) svg polyline {{
        stroke-dashoffset: 0;
    }}
    /* Pop the checkbox icon slightly on check (keeps text stable) */
    div[data-baseweb="checkbox"]:has(input[type="checkbox"]:checked) svg {{
        transform-origin: center;
        animation: tdb-checkbox-pop 160ms ease-out;
    }}
    @keyframes tdb-checkbox-pop {{
        0% {{ transform: scale(1); }}
        60% {{ transform: scale(1.12); }}
        100% {{ transform: scale(1.06); }}
    }}
    /* Keyboard focus ring for accessibility */
    div[data-baseweb="checkbox"]:has(input[type="checkbox"]:focus-visible) svg {{
        outline: 2px solid var(--text-secondary);
        outline-offset: 3px;
        border-radius: 4px;
    }}
    /* Reduced motion support */
    @media (prefers-reduced-motion: reduce) {{
        [data-testid="stDataFrameContainer"] input[type="checkbox"] {{
            transition: none !important;
        }}
        [data-testid="stDataFrameContainer"] input[type="checkbox"]:checked {{
            animation: none !important;
        }}
        div[data-baseweb="checkbox"] svg path:nth-of-type(2),
        div[data-baseweb="checkbox"] svg polyline {{
            transition: none;
        }}
        div[data-baseweb="checkbox"]:has(input[type="checkbox"]:checked) svg {{
            animation: none;
        }}
    }}
    
    /* Divider styling */
    hr {{
        border: none !important;
        border-top: 2px solid #3b82f6 !important;
        margin: 2rem 0 !important;
    }}
    
    /* Section cards */
    .section-card {{
        background-color: {COLORS['bg_secondary']} !important;
        border-radius: 8px !important;
        padding: 1.5rem !important;
        border: 1px solid #3b82f6 !important;
        margin-bottom: 1.5rem !important;
        box-shadow: 0 2px 8px rgba(20, 17, 15, 0.08) !important;
    }}
    
    /* Save button styling - aesthetic and smooth */
    button[key="manual_save_btn"] {{
        background: {COLORS['button_bg']} !important;
        color: {COLORS['button_text']} !important;
        border: none !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 2px 8px rgba(52, 49, 45, 0.3) !important;
        padding: 10px 20px !important;
    }}
    
    button[key="manual_save_btn"]:hover {{
        background: {COLORS['button_bg']} !important;
        box-shadow: 0 4px 14px rgba(52, 49, 45, 0.4) !important;
        transform: translateY(-2px) !important;
    }}
    
    button[key="manual_save_btn"]:active {{
        transform: translateY(0) !important;
        box-shadow: 0 2px 6px rgba(52, 49, 45, 0.3) !important;
    }}
    /* Availability dashboard styling */
    .availability-summary {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
        gap: 1.2rem;
        margin: 1.25rem 0 1.75rem 0;
    }}
    .availability-card {{
        background: var(--bg-secondary);
        border: 1px solid var(--glass-border);
        border-radius: 18px;
        padding: 1.25rem 1.35rem;
        box-shadow: 0 12px 32px rgba(20, 17, 15, 0.22);
        backdrop-filter: none !important;
        display: flex;
        flex-direction: column;
        gap: 0.55rem;
        position: relative;
        overflow: hidden;
    }}
    .availability-card::after {{
        content: "";
        position: absolute;
        inset: 1px;
        border-radius: 16px;
        border-top: 4px solid var(--accent);
        opacity: 0.9;
        pointer-events: none;
    }}
    .availability-card.success::after {{ border-top-color: var(--success); }}
    .availability-card.warning::after {{ border-top-color: var(--warning); }}
    .availability-card.danger::after {{ border-top-color: var(--danger); }}
    .availability-card__icon {{
        font-size: 1.8rem;
        line-height: 1;
    }}
    .availability-card h4 {{
        margin: 0;
        font-size: 0.9rem;
        letter-spacing: 0.08em;
        text-transform: uppercase;
        color: var(--text-secondary);
    }}
    .availability-card strong {{
        display: block;
        font-size: 2.4rem;
        margin: 0.1rem 0 0;
        color: var(--text-primary);
        letter-spacing: -0.02em;
    }}
    .availability-card p {{
        margin: 0;
        color: var(--text-primary);
        opacity: 0.7;
        font-size: 0.95rem;
    }}
    .assistant-grid {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
        gap: 16px;
        margin-top: 1rem;
    }}
    .assistant-card {{
        background: var(--glass-bg);
        border: 1px solid var(--glass-border);
        border-top: 4px solid var(--accent);
        border-radius: 14px;
        padding: 1rem 1.1rem;
        box-shadow: 0 12px 28px rgba(20, 17, 15, 0.18);
        min-height: 140px;
        display: flex;
        flex-direction: column;
        gap: 0.5rem;
        backdrop-filter: none !important;
        -webkit-backdrop-filter: none !important;
        transition: transform 0.2s ease, box-shadow 0.2s ease;
    }}
    .assistant-card:hover {{
        transform: translateY(-4px);
        box-shadow: 0 16px 36px rgba(20, 17, 15, 0.24);
    }}
    .assistant-card.status-free {{ border-top-color: var(--success); }}
    .assistant-card.status-busy {{ border-top-color: var(--warning); }}
    .assistant-card.status-blocked {{ border-top-color: var(--danger); }}
    .assistant-card.status-unknown {{ border-top-color: var(--info); }}
    .assistant-card__header {{
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 0.6rem;
        margin-bottom: 0.25rem;
    }}
    .assistant-card__name {{
        font-size: 1.1rem;
        font-weight: 700;
        color: var(--text-primary);
        letter-spacing: 0.02em;
    }}
    .assistant-card__status-pill {{
        font-size: 0.7rem;
        padding: 0.2rem 0.65rem;
        border-radius: 999px;
        text-transform: uppercase;
        letter-spacing: 0.06em;
        font-weight: 600;
        white-space: nowrap;
    }}
    .assistant-card__status-pill.success {{
        background: rgba(52, 49, 45, 0.18);
        color: var(--text-primary);
        border: 1px solid rgba(52, 49, 45, 0.35);
    }}
    
    .assistant-card__status-pill.warning {{
        background: rgba(217, 197, 178, 0.35);
        color: var(--text-primary);
        border: 1px solid rgba(217, 197, 178, 0.6);
    }}
    
    .assistant-card__status-pill.danger {{
        background: rgba(126, 127, 131, 0.25);
        color: var(--text-primary);
        border: 1px solid rgba(126, 127, 131, 0.45);
    }}
    
    .assistant-card__status-pill.info {{
        background: rgba(20, 17, 15, 0.12);
        color: var(--text-primary);
        border: 1px solid rgba(20, 17, 15, 0.35);
    }}
    .assistant-card__details {{
        font-size: 0.88rem;
        color: var(--text-primary);
        opacity: 0.9;
        line-height: 1.4;
        flex-grow: 1;
    }}
    .assistant-card__meta {{
        margin-top: auto;
        padding-top: 0.5rem;
        font-size: 0.75rem;
        text-transform: uppercase;
        letter-spacing: 0.04em;
        color: var(--text-secondary);
        opacity: 0.8;
        display: flex;
        justify-content: space-between;
        border-top: 1px solid var(--glass-border);
    }}
    @media (max-width: 768px) {{
        .assistant-grid {{
            grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
        }}
    }}
    </style>
    """,
    unsafe_allow_html=True
)
_logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "The Dental Bond LOGO_page-0001.jpg")
logo_b64 = ""
if os.path.exists(_logo_path):
    import base64
    with open(_logo_path, "rb") as f:
        logo_b64 = base64.b64encode(f.read()).decode()
header_css = f"""
<style>
header[data-testid="stHeader"] {{
    background: #f8fafc !important;
    min-height: 72px;
    border-bottom: 1px solid #3b82f6;
    box-shadow: 0 12px 32px rgba(20, 17, 15, 0.18);
    position: sticky;
    top: 0;
    z-index: 100;
}}
header[data-testid="stHeader"] [data-testid="stToolbarActions"],
header[data-testid="stHeader"] [data-testid="stMainMenu"] {{
    display: none !important;
    visibility: hidden !important;
}}
header[data-testid="stHeader"]::after {{
    content: "THE DENTAL BOND\\AReal-time Scheduling Management System";
    white-space: pre;
    position: absolute;
    left: 50%;
    top: 50%;
    transform: translate(-50%, -50%);
    text-align: center;
    font-size: 20px;
    font-weight: 800;
    line-height: 1.3;
    letter-spacing: 0.4px;
    color: #1e293b;
    text-shadow: 0 3px 10px rgba(52, 49, 45, 0.2);
    pointer-events: none;
}}
</style>
"""
st.markdown(header_css, unsafe_allow_html=True)
# ================ HELPER FUNCTIONS ================
# Note: Core time functions moved to top of file to fix import order bugs
def _parse_iso_ts(val):
    try:
        if isinstance(val, datetime):
            return val
        if isinstance(val, str):
            return datetime.fromisoformat(val.replace("Z", "+00:00"))
    except Exception:
        return None
    return None
def _safe_int(val, default: int) -> int:
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return default
        return int(float(val))
    except Exception:
        return default
def _date_from_any(val):
    try:
        if isinstance(val, datetime):
            return val.date()
        if hasattr(val, "date"):
            return val.date()
        if isinstance(val, str) and val:
            return datetime.fromisoformat(val.replace("Z", "+00:00")).date()
    except Exception:
        return None
    return None
# Always update 'now' at the top of the main script body for correct time blocking
now = now_ist()
date_line_str = now.strftime('%B %d, %Y - %I:%M:%S %p')
if st.session_state.get("nav_category") != "Dashboard":
    st.markdown(f"""
        <style>
        .divider-line {{
            height: 2px;
            background: {COLORS['accent']};
            margin: 0.8rem 0;
            border-radius: 1px;
        }}
        .sticky-top {{
            position: sticky;
            top: 0;
            z-index: 999;
            background: {COLORS['bg_primary']};
            padding: 0.4rem 0 0.35rem 0;
            box-shadow: none;
        }}
        .date-line {{
            font-size: 1rem;
            font-weight: 600;
            color: var(--text-primary);
            margin-top: 0.5rem;
        }}
        </style>
        <div class="sticky-top">
            <div class="date-line">{date_line_str} IST</div>
        </div>
    """, unsafe_allow_html=True)
    # Assistants Weekly Off display (10mm below date)
    st.markdown("<div style='margin-top:10mm;'></div>", unsafe_allow_html=True)
    weekday_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    today_idx = now.weekday()
    tomorrow_idx = (today_idx + 1) % 7
    weekly_off_map = _get_profiles_cache_snapshot().get("weekly_off_map", WEEKLY_OFF)
    def _render_off_card(title: str, off_list: list[str]):
        has_off = bool(off_list)
        names = ", ".join(off_list) if has_off else "All assistants available"
        icon = "🚫" if has_off else "✅"
        bg = COLORS['danger'] if has_off else COLORS['success']
        border = COLORS['danger'] if has_off else COLORS['success']
        note = "Cannot be allocated" if has_off else "No weekly off"
        st.markdown(
            f"""
            <div style="
                background: {COLORS['bg_secondary']};
                border: 1px solid {border}40;
                border-left: 4px solid {border};
                border-radius: 8px;
                padding: 12px 14px;
                margin: 6px 0 10px 0;
                display: flex;
                align-items: center;
                gap: 10px;
            ">
                <span style="font-size: 1.3em;">{icon}</span>
                <div>
                    <strong style="color: {COLORS['text_primary']};">{title}</strong>
                    <div style="color: {COLORS['text_secondary']}; margin-top: 2px;">
                        <strong>{names}</strong> — {note}
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    st.markdown("### 🗓️ Assistants Weekly Off")
    col_today, col_tomorrow = st.columns(2)
    with col_today:
        _render_off_card(
            f"Today ({weekday_names[today_idx]})",
            weekly_off_map.get(today_idx, []),
        )
    with col_tomorrow:
        _render_off_card(
            f"Tomorrow ({weekday_names[tomorrow_idx]})",
            weekly_off_map.get(tomorrow_idx, []),
        )
def _get_app_version_short() -> str:
    """Best-effort git/version identifier for display.
    Streamlit Cloud does not guarantee a .git directory is present at runtime,
    so we fall back to common CI env vars when available.
    """
    for key in (
        "STREAMLIT_GIT_COMMIT",
        "GIT_COMMIT",
        "GITHUB_SHA",
        "COMMIT_SHA",
        "VERCEL_GIT_COMMIT_SHA",
        "RENDER_GIT_COMMIT",
        "CF_PAGES_COMMIT_SHA",
    ):
        val = (os.environ.get(key) or "").strip()
        if val:
            return val[:7]
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        head_path = os.path.join(base_dir, ".git", "HEAD")
        if os.path.exists(head_path):
            head = (open(head_path, "r", encoding="utf-8").read() or "").strip()
            if head.startswith("ref:"):
                ref_rel = head.split("ref:", 1)[1].strip()
                ref_path = os.path.join(base_dir, ".git", *ref_rel.split("/"))
                if os.path.exists(ref_path):
                    sha = (open(ref_path, "r", encoding="utf-8").read() or "").strip()
                    if sha:
                        return sha[:7]
            elif head:
                return head[:7]
    except Exception:
        pass
    return "unknown"
# Epoch seconds (used for 30-second snooze timing)
now_epoch = int(time_module.time())
# ================ TIME UTILITY FUNCTIONS ================
# Define time conversion functions early so they can be used throughout the code
def _coerce_to_time_obj(time_value: Any) -> Optional[time_type]:
    """Best-effort coercion of many time representations into a datetime.time.
    Supports:
    - datetime.time, datetime
    - strings: HH:MM, HH:MM:SS, HH.MM, and 12-hour formats like '09:30 AM'
    - numbers: 9.30 (meaning 09:30), or Excel serial time 0-1
    """
    if time_value is None or pd.isna(time_value) or time_value == "":
        return None
    if isinstance(time_value, time_type):
        return time_value
    # Strings
    if isinstance(time_value, str):
        s = " ".join(time_value.strip().split())
        if s == "" or s.upper() in {"N/A", "NAT", "NONE"}:
            return None
        # 12-hour formats (e.g., 09:30 AM, 9:30PM, 09:30:00 PM)
        if re.search(r"\b(AM|PM)\b", s, flags=re.IGNORECASE) or re.search(r"(AM|PM)$", s, flags=re.IGNORECASE):
            s_norm = re.sub(r"\s*(AM|PM)\s*$", r" \1", s, flags=re.IGNORECASE).upper()
            for fmt in ("%I:%M %p", "%I:%M:%S %p"):
                try:
                    dt = datetime.strptime(s_norm, fmt)
                    return time_type(dt.hour, dt.minute)
                except ValueError:
                    pass
        # HH:MM or HH:MM:SS
        if ":" in s:
            parts = s.split(":")
            if len(parts) >= 2:
                try:
                    h = int(parts[0])
                    m_part = re.sub(r"\D.*$", "", parts[1])
                    m = int(m_part)
                    if 0 <= h < 24 and 0 <= m < 60:
                        return time_type(h, m)
                except (ValueError, TypeError):
                    pass
        # HH.MM
        if "." in s:
            parts = s.split(".")
            if len(parts) == 2:
                try:
                    h = int(parts[0])
                    m = int(parts[1])
                    if 0 <= h < 24 and 0 <= m < 60:
                        return time_type(h, m)
                except (ValueError, TypeError):
                    pass
        return None
    # Numeric formats
    try:
        num_val = float(time_value)
    except (ValueError, TypeError):
        return None
    # Excel serial time format (0.625 = 15:00)
    if 0 <= num_val <= 1:
        total_minutes = round(num_val * 1440)
        hours = (total_minutes // 60) % 24
        minutes = total_minutes % 60
        return time_type(hours, minutes)
    # 9.30 meaning 09:30 (decimal part is minutes directly)
    if 0 <= num_val < 24:
        hours = int(num_val)
        decimal_part = num_val - hours
        minutes = round(decimal_part * 100)
        if minutes > 59:
            minutes = round(decimal_part * 60)
        if minutes >= 60:
            hours = (hours + 1) % 24
            minutes = 0
        if 0 <= hours < 24 and 0 <= minutes < 60:
            return time_type(hours, minutes)
    return None
TIME_PICKER_HOURS = [""] + [f"{i:02d}" for i in range(1, 13)]
TIME_PICKER_MINUTES = [""] + [f"{i:02d}" for i in range(60)]
TIME_PICKER_AMPM = ["AM", "PM"]
def _time_to_picker_parts(time_value: Any) -> tuple[str, str, str]:
    t = _coerce_to_time_obj(time_value)
    if t is None:
        return "", "", "AM"
    hour = t.hour
    if hour == 0:
        hour_12 = 12
        ampm = "AM"
    elif hour < 12:
        hour_12 = hour
        ampm = "AM"
    elif hour == 12:
        hour_12 = 12
        ampm = "PM"
    else:
        hour_12 = hour - 12
        ampm = "PM"
    return f"{hour_12:02d}", f"{t.minute:02d}", ampm
def _time_from_picker_parts(hour_str: str, minute_str: str, ampm: str) -> tuple[str, Optional[str]]:
    hour_text = str(hour_str or "").strip()
    minute_text = str(minute_str or "").strip()
    if not hour_text and not minute_text:
        return "", None
    if not hour_text or not minute_text:
        return "", "Select hour and minute."
    try:
        hour = int(hour_text)
        minute = int(minute_text)
    except ValueError:
        return "", "Invalid time selection."
    if hour < 1 or hour > 12 or minute < 0 or minute > 59:
        return "", "Invalid time selection."
    ampm_norm = str(ampm or "").strip().upper()
    if ampm_norm not in ("AM", "PM"):
        return "", "Select AM or PM."
    if hour == 12:
        hour_24 = 0 if ampm_norm == "AM" else 12
    else:
        hour_24 = hour if ampm_norm == "AM" else hour + 12
    return f"{hour_24:02d}:{minute:02d}", None
def dec_to_time(time_value: Any) -> str:
    """Convert various time formats to HH:MM string"""
    t = _coerce_to_time_obj(time_value)
    if t is None:
        return "N/A"
    return f"{t.hour:02d}:{t.minute:02d}"
# Removed duplicate: safe_str_to_time_obj (now defined at top of file)
def time_obj_to_str(t: Any) -> str:
    """Convert time object to 24-hour HH:MM string for Excel"""
    if pd.isna(t) or t is None:
        return "N/A"
    try:
        if isinstance(t, time_type):
            return f"{t.hour:02d}:{t.minute:02d}"
        elif isinstance(t, str):
            return t
    except (ValueError, AttributeError):
        pass
    return "N/A"
def time_obj_to_str_12hr(t: Any) -> str:
    """Convert time object to 12-hour format with AM/PM"""
    if pd.isna(t) or t is None:
        return "N/A"
    try:
        if isinstance(t, time_type):
            return t.strftime("%I:%M %p")
        elif isinstance(t, str):
            return t
    except (ValueError, AttributeError):
        pass
    return "N/A"
# Removed duplicate: time_to_minutes (now defined at top of file)
# Note: The top version may be less comprehensive than _coerce_to_time_obj approach
# but handles the common HH:MM string and numeric cases
# ================ DEPARTMENT & STAFF CONFIGURATION ================
# Departments with their doctors and assistants
# NOTE: Keep these lists as the single source of truth for dropdowns + allocation.
def _unique_preserve_order(items: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for item in items:
        key = str(item).strip().upper()
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(key)
    return out
def _norm_staff_key(value: str) -> str:
    """Normalize names like 'DR. NAME' vs 'DR.NAME' to a stable key."""
    try:
        s = str(value or "").strip().upper()
        return re.sub(r"[^A-Z0-9]+", "", s)
    except Exception:
        return ""
def _parse_weekly_off_days(val: str) -> list[int]:
    """Parse weekly off string to list of weekday indices (0=Mon)."""
    if val is None:
        return []
    days_map = {
        "MONDAY": 0, "MON": 0,
        "TUESDAY": 1, "TUE": 1, "TUES": 1,
        "WEDNESDAY": 2, "WED": 2,
        "THURSDAY": 3, "THU": 3, "THURS": 3,
        "FRIDAY": 4, "FRI": 4,
        "SATURDAY": 5, "SAT": 5,
        "SUNDAY": 6, "SUN": 6,
    }
    out: list[int] = []
    parts: list[Any] = []
    if isinstance(val, (list, tuple, set)):
        parts = list(val)
    elif isinstance(val, str):
        raw = val.strip()
        if not raw:
            return []
        if (raw.startswith("[") and raw.endswith("]")) or (raw.startswith("{") and raw.endswith("}")):
            try:
                parsed = json.loads(raw)
                if isinstance(parsed, (list, tuple, set)):
                    parts = list(parsed)
                else:
                    parts = [parsed]
            except Exception:
                parts = raw.replace(";", ",").split(",")
        else:
            parts = raw.replace(";", ",").split(",")
    else:
        parts = [val]
    for part in parts:
        if part is None or (isinstance(part, float) and pd.isna(part)):
            continue
        p = str(part).strip().upper()
        if not p:
            continue
        if p.isdigit():
            idx = int(p)
            if 0 <= idx <= 6:
                out.append(idx)
        elif p in days_map:
            out.append(days_map[p])
    return out
WEEKDAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
def _weekly_off_names(val: str) -> list[str]:
    idxs = _parse_weekly_off_days(val)
    return [WEEKDAY_NAMES[i] for i in idxs if 0 <= i < 7]
def _weekly_off_str_from_list(lst: list[str]) -> str:
    if not lst:
        return ""
    names_upper = [str(x).strip().upper() for x in lst if str(x).strip()]
    out = []
    for nm in names_upper:
        if nm in [d.upper() for d in WEEKDAY_NAMES]:
            out.append(WEEKDAY_NAMES[[d.upper() for d in WEEKDAY_NAMES].index(nm)])
    return ",".join(out)
ALLOCATION_RULES_PATH = Path(__file__).with_name("allocation_rules.json")
def _config_bool(val: Any, default: bool = False) -> bool:
    if isinstance(val, bool):
        return val
    if val is None:
        return default
    s = str(val).strip().lower()
    if s in {"1", "true", "yes", "on"}:
        return True
    if s in {"0", "false", "no", "off"}:
        return False
    return default
@st.cache_data(ttl=30)
def _load_allocation_config_cached(path_str: str, mtime: float) -> dict[str, Any]:
    try:
        with open(path_str, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}
def _get_allocation_config() -> dict[str, Any]:
    try:
        if ALLOCATION_RULES_PATH.exists():
            mtime = ALLOCATION_RULES_PATH.stat().st_mtime
            payload = _load_allocation_config_cached(str(ALLOCATION_RULES_PATH), mtime)
            return payload if isinstance(payload, dict) else {}
    except Exception:
        pass
    return {}
def _get_allocation_department_config(department: str, config: Optional[dict[str, Any]] = None) -> dict[str, Any]:
    dept_upper = str(department).strip().upper()
    if not dept_upper:
        return {}
    cfg = config or _get_allocation_config()
    if not isinstance(cfg, dict):
        return {}
    depts = cfg.get("departments", {})
    if isinstance(depts, dict):
        for key, val in depts.items():
            if str(key).strip().upper() == dept_upper:
                return val if isinstance(val, dict) else {}
    return {}
def _get_global_allocation_config(config: Optional[dict[str, Any]] = None) -> dict[str, bool]:
    cfg = config or _get_allocation_config()
    if not isinstance(cfg, dict):
        return {
            "cross_department_fallback": False,
            "use_profile_role_flags": False,
            "load_balance": False,
        }
    global_cfg = cfg.get("global", {}) if isinstance(cfg.get("global", {}), dict) else {}
    return {
        "cross_department_fallback": _config_bool(global_cfg.get("cross_department_fallback", False)),
        "use_profile_role_flags": _config_bool(global_cfg.get("use_profile_role_flags", False)),
        "load_balance": _config_bool(global_cfg.get("load_balance", False)),
    }
def _get_config_department_maps(config: Optional[dict[str, Any]] = None) -> dict[str, Any]:
    cfg = config or _get_allocation_config()
    doctor_map: dict[str, str] = {}
    assistant_map: dict[str, str] = {}
    dept_list: list[str] = []
    depts = cfg.get("departments") if isinstance(cfg, dict) else None
    if isinstance(depts, dict) and depts:
        for dept, data in depts.items():
            dept_upper = str(dept).strip().upper()
            if dept_upper and dept_upper not in dept_list:
                dept_list.append(dept_upper)
            if not isinstance(data, dict):
                continue
            for name in data.get("doctors", []) or []:
                key = _norm_staff_key(name)
                if key and key not in doctor_map:
                    doctor_map[key] = dept_upper
            for name in data.get("assistants", []) or []:
                key = _norm_staff_key(name)
                if key and key not in assistant_map:
                    assistant_map[key] = dept_upper
    if not dept_list:
        for dept, data in DEPARTMENTS.items():
            dept_upper = str(dept).strip().upper()
            if dept_upper and dept_upper not in dept_list:
                dept_list.append(dept_upper)
            if not isinstance(data, dict):
                continue
            for name in data.get("doctors", []) or []:
                key = _norm_staff_key(name)
                if key and key not in doctor_map:
                    doctor_map[key] = dept_upper
            for name in data.get("assistants", []) or []:
                key = _norm_staff_key(name)
                if key and key not in assistant_map:
                    assistant_map[key] = dept_upper
    return {
        "departments": dept_list,
        "doctors": doctor_map,
        "assistants": assistant_map,
    }
def _seed_supabase_profiles_if_needed(client) -> None:
    """Ensure all configured assistants/doctors exist in Supabase profiles table."""
    if client is None:
        return
    try:
        resp = client.table(PROFILE_SUPABASE_TABLE).select("id,name,kind").execute()
        existing = resp.data or []
        seen = {_norm_staff_key(r.get("name", "")) + "|" + str(r.get("kind", "")).upper() for r in existing}
    except Exception:
        seen = set()
    now_iso = now_ist().isoformat(timespec="seconds")
    to_insert: list[dict[str, Any]] = []
    def _add(name: str, dept: str, kind: str):
        key = _norm_staff_key(name) + "|" + kind.upper()
        if key in seen:
            return
        to_insert.append({
            "id": str(uuid.uuid4()),
            "name": name.upper(),
            "department": dept,
            "contact_email": "",
            "contact_phone": "",
            "status": "ACTIVE",
            "created_at": now_iso,
            "updated_at": now_iso,
            "created_by": "system_seed",
            "updated_by": "system_seed",
            "kind": kind,
        })
    config = _get_allocation_config()
    dept_source = config.get("departments", {}) if isinstance(config, dict) else {}
    if not isinstance(dept_source, dict) or not dept_source:
        dept_source = DEPARTMENTS
    for dept, data in dept_source.items():
        if not isinstance(data, dict):
            continue
        for a in data.get("assistants", []) or []:
            _add(a, dept, PROFILE_ASSISTANT_SHEET)
        for d in data.get("doctors", []) or []:
            _add(d, dept, PROFILE_DOCTOR_SHEET)
    if to_insert:
        try:
            client.table(PROFILE_SUPABASE_TABLE).insert(to_insert).execute()
            st.sidebar.info(f"Seeded {len(to_insert)} profiles to Supabase.")
        except Exception:
            pass
def _refresh_staff_options_from_supabase(client):
    """Override ALL_ASSISTANTS/ALL_DOCTORS and WEEKLY_OFF from Supabase profiles."""
    global ALL_ASSISTANTS, ALL_DOCTORS, WEEKLY_OFF
    try:
        resp = client.table(PROFILE_SUPABASE_TABLE).select("*").execute()
        data = resp.data or []
        df = pd.DataFrame(data)
        if df.empty:
            return
        df["name"] = df["name"].astype(str).str.upper()
        df["department"] = df.get("department", "").astype(str).str.upper()
        assistants = df[df["kind"] == PROFILE_ASSISTANT_SHEET]["name"].dropna().tolist()
        doctors = df[df["kind"] == PROFILE_DOCTOR_SHEET]["name"].dropna().tolist()
        if assistants:
            ALL_ASSISTANTS = _unique_preserve_order(assistants)
        if doctors:
            ALL_DOCTORS = _unique_preserve_order(doctors)
        # Weekly off mapping
        week_map: dict[int, list[str]] = {i: [] for i in range(7)}
        if "weekly_off" in df.columns:
            for _, row in df.iterrows():
                kind = str(row.get("kind", "")).strip()
                if kind != PROFILE_ASSISTANT_SHEET:
                    continue
                wo_days = _parse_weekly_off_days(row.get("weekly_off", ""))
                name = str(row.get("name", "")).strip().upper()
                if not name:
                    continue
                for idx in wo_days:
                    week_map[idx].append(name)
        WEEKLY_OFF = week_map
    except Exception:
        pass
def _is_blank_cell(value: Any) -> bool:
    """True if value is empty/NaN/'nan'/'none'."""
    try:
        if value is None or pd.isna(value):
            return True
    except Exception:
        pass
    s = str(value).strip()
    return (not s) or (s.lower() in {"nan", "none", "nat", "n/a", "na", "null", "-", "--"})
DEPARTMENTS = {
    "PROSTHO": {
        "doctors": _unique_preserve_order([]),  # Empty - add manually via UI
        "assistants": _unique_preserve_order([]),  # Empty - add manually via UI
        "allocation_rules": {
            "FIRST": {
                "default": [],
                "time_override": []
            },
            "SECOND": {
                "when_first_is": {},
                "default": []
            },
        }
    },
    "ENDO": {
        "doctors": _unique_preserve_order([]),  # Empty - add manually via UI
        "assistants": _unique_preserve_order([]),  # Empty - add manually via UI
        "allocation_rules": {
            "FIRST": {
                "default": [],
                "time_override": []
            },
            "SECOND": {
                "default": []
            },
            "Third": {
                "default": []
            },
        }
    },
}
# Combined lists for dropdowns
ALL_DOCTORS = _unique_preserve_order(DEPARTMENTS["PROSTHO"]["doctors"] + DEPARTMENTS["ENDO"]["doctors"])
ALL_ASSISTANTS = _unique_preserve_order(DEPARTMENTS["PROSTHO"]["assistants"] + DEPARTMENTS["ENDO"]["assistants"])
def get_department_for_doctor(doctor_name: str) -> str:
    """Get the department a doctor belongs to"""
    if not doctor_name:
        return ""
    doc_key = _norm_staff_key(doctor_name)
    if not doc_key:
        return ""
    try:
        cache = _get_profiles_cache()
        dept = cache.get("doctor_dept_map", {}).get(doc_key, "")
        if dept:
            return dept
    except Exception:
        pass
    try:
        config_maps = _get_config_department_maps()
        dept = config_maps.get("doctors", {}).get(doc_key, "")
        if dept:
            return dept
    except Exception:
        pass
    for dept, config in DEPARTMENTS.items():
        for d in config["doctors"]:
            d_key = _norm_staff_key(d)
            if not d_key:
                continue
            if doc_key == d_key or doc_key.endswith(d_key) or d_key.endswith(doc_key):
                return dept
    return ""
def get_assistants_for_department(department: str) -> list[str]:
    """Get list of assistants for a specific department"""
    dept_upper = str(department).strip().upper()
    if not dept_upper:
        return _get_all_assistants()
    try:
        cache = _get_profiles_cache()
        dept_list = cache.get("assistants_by_dept", {}).get(dept_upper, [])
        if dept_list:
            return dept_list
    except Exception:
        pass
    try:
        cfg = _get_allocation_config()
        dept_cfg = _get_allocation_department_config(dept_upper, cfg)
        assistants = dept_cfg.get("assistants", []) if isinstance(dept_cfg, dict) else []
        if assistants:
            return _unique_preserve_order(assistants)
    except Exception:
        pass
    if dept_upper in DEPARTMENTS:
        return DEPARTMENTS[dept_upper]["assistants"]
    return _get_all_assistants()
def get_department_for_assistant(assistant_name: str) -> str:
    """Get the department an assistant belongs to"""
    if not assistant_name:
        return ""
    assist_key = _norm_staff_key(assistant_name)
    if not assist_key:
        return ""
    try:
        cache = _get_profiles_cache()
        dept = cache.get("assistant_dept_map", {}).get(assist_key, "")
        if dept:
            return dept
    except Exception:
        pass
    try:
        config_maps = _get_config_department_maps()
        dept = config_maps.get("assistants", {}).get(assist_key, "")
        if dept:
            return dept
    except Exception:
        pass
    for dept, config in DEPARTMENTS.items():
        for a in config["assistants"]:
            a_key = _norm_staff_key(a)
            if not a_key:
                continue
            if assist_key == a_key or assist_key.endswith(a_key) or a_key.endswith(assist_key):
                return dept
    return "SHARED"
# ================ TIME BLOCKING SYSTEM ================
# Initialize time blocks in session state
if "time_blocks" not in st.session_state:
    st.session_state.time_blocks = []  # List of {assistant, start_time, end_time, reason, date}
def add_time_block(assistant: str, start_time: Any, end_time: Any, reason: str = "Backend Work") -> bool:
    """Add a time block for an assistant. Returns True when recorded."""
    today_str = now.strftime("%Y-%m-%d")
    block = {
        "assistant": assistant.upper(),
        "start_time": start_time,
        "end_time": end_time,
        "reason": reason,
        "date": today_str
    }
    st.session_state.time_blocks.append(block)
    return True
def remove_time_block(index: int):
    """Remove a time block by index"""
    if 0 <= index < len(st.session_state.time_blocks):
        st.session_state.time_blocks.pop(index)
        return True
    return False
def is_assistant_blocked(assistant: str, check_time: Any) -> tuple[bool, str]:
    """Check if an assistant is blocked at a specific time. Returns (is_blocked, reason)"""
    if not assistant or not check_time:
        return False, ""
    
    assist_upper = str(assistant).strip().upper()
    today_str = now.strftime("%Y-%m-%d")
    check_minutes = check_time.hour * 60 + check_time.minute
    
    for block in st.session_state.time_blocks:
        if block["date"] != today_str:
            continue
        if block["assistant"].upper() != assist_upper:
            continue
        
        start_min = block["start_time"].hour * 60 + block["start_time"].minute
        end_min = block["end_time"].hour * 60 + block["end_time"].minute
        
        if start_min <= check_minutes <= end_min:
            return True, block.get("reason", "Blocked")
    
    return False, ""
def _time_to_hhmm(t: Optional[time_type]) -> str:
    if t is None:
        return ""
    return f"{t.hour:02d}:{t.minute:02d}"
def _serialize_time_blocks(blocks: list[dict]) -> list[dict]:
    """Convert session_state time blocks into JSON-safe primitives."""
    out: list[dict] = []
    for b in blocks or []:
        try:
            assistant = str(b.get("assistant", "")).strip().upper()
            date = str(b.get("date", "")).strip()
            reason = str(b.get("reason", "Backend Work")).strip() or "Backend Work"
            start_t = b.get("start_time")
            end_t = b.get("end_time")
            start_obj = _coerce_to_time_obj(start_t)
            end_obj = _coerce_to_time_obj(end_t)
            out.append(
                {
                    "assistant": assistant,
                    "date": date,
                    "reason": reason,
                    "start_time": _time_to_hhmm(start_obj),
                    "end_time": _time_to_hhmm(end_obj),
                }
            )
        except Exception:
            continue
    return out
def _deserialize_time_blocks(value) -> list[dict]:
    """Parse stored meta value into session_state-compatible time blocks."""
    if value is None or value == "":
        return []
    raw = value
    if isinstance(value, str):
        try:
            raw = json.loads(value)
        except Exception:
            return []
    if not isinstance(raw, list):
        return []
    out: list[dict] = []
    for b in raw:
        if not isinstance(b, dict):
            continue
        assistant = str(b.get("assistant", "")).strip().upper()
        date = str(b.get("date", "")).strip()
        reason = str(b.get("reason", "Backend Work")).strip() or "Backend Work"
        start_obj = _coerce_to_time_obj(b.get("start_time"))
        end_obj = _coerce_to_time_obj(b.get("end_time"))
        if not assistant or not date or start_obj is None or end_obj is None:
            continue
        out.append(
            {
                "assistant": assistant,
                "date": date,
                "reason": reason,
                "start_time": start_obj,
                "end_time": end_obj,
            }
        )
    return out
def _get_meta_from_df(df_any: Optional[DataFrame]) -> dict:
    try:
        if df_any is not None and hasattr(df_any, "attrs"):
            meta = df_any.attrs.get("meta")
            if isinstance(meta, dict):
                return dict(meta)
    except Exception:
        pass
    return {}
def _set_meta_on_df(df_any: pd.DataFrame, meta: dict) -> None:
    try:
        if hasattr(df_any, "attrs"):
            df_any.attrs["meta"] = dict(meta or {})
    except Exception:
        pass
def _sync_time_blocks_from_meta(df_any: Optional[DataFrame]) -> None:
    """Load persisted time blocks into session_state (once per run)."""
    try:
        meta = _get_meta_from_df(df_any)
        if "time_blocks" in meta:
            blocks = _deserialize_time_blocks(meta.get("time_blocks"))
            st.session_state.time_blocks = blocks
    except Exception:
        pass
def _apply_time_blocks_to_meta(meta: dict) -> dict:
    out = dict(meta or {})
    serialized = _serialize_time_blocks(st.session_state.get("time_blocks", []))
    prev = out.get("time_blocks")
    out["time_blocks"] = serialized
    if prev != serialized or not out.get("time_blocks_updated_at"):
        out["time_blocks_updated_at"] = datetime.now(IST).isoformat()
    return out
# ================ ASSISTANT AVAILABILITY TRACKING ================
def get_assistant_schedule(assistant_name: str, df_schedule: pd.DataFrame) -> list[dict[str, Any]]:
    """Get all appointments where this assistant is assigned"""
    if not assistant_name or df_schedule.empty:
        return []
    
    assist_upper = str(assistant_name).strip().upper()
    appointments = []
    third_col = _get_third_column_name(df_schedule.columns)
    
    for idx, row in df_schedule.iterrows():
        # Check FIRST, SECOND, Third columns
        for col in ["FIRST", "SECOND", third_col]:
            if col in row.index:
                val = str(row.get(col, "")).strip().upper()
                if val == assist_upper:
                    # Skip cancelled/done/completed/shifted appointments
                    status = str(row.get("STATUS", "")).strip().upper()
                    if any(s in status for s in ["CANCELLED", "DONE", "COMPLETED", "SHIFTED"]):
                        continue
                    
                    appointments.append({
                        "row_id": row.get("REMINDER_ROW_ID", ""),
                        "patient": row.get("Patient Name", "Unknown"),
                        "in_time": row.get("In Time"),
                        "out_time": row.get("Out Time"),
                        "doctor": row.get("DR.", ""),
                        "op": row.get("OP", ""),
                        "role": col,
                        "status": status
                    })
    
    return appointments
def is_assistant_available(
    assistant_name: str,
    check_in_time,
    check_out_time,
    df_schedule: pd.DataFrame,
    exclude_row_id: Optional[str] = None,
) -> tuple[bool, str]:
    """
    Check if an assistant is available during a time window.
    Returns (is_available, conflict_reason)
    """
    if not assistant_name:
        return False, "No assistant specified"
    
    assist_upper = str(assistant_name).strip().upper()
    punch_map = _get_today_punch_map()
    punch_state, _, punch_out = _assistant_punch_state(assist_upper, punch_map)
    if punch_state != "IN":
        try:
            today_weekday = now.weekday()  # 0=Monday, 6=Sunday
            weekly_off_map = _get_profiles_cache().get("weekly_off_map", WEEKLY_OFF)
            off_assistants = weekly_off_map.get(today_weekday, [])
            if any(str(a).strip().upper() == assist_upper for a in off_assistants):
                return False, f"Weekly off on {now.strftime('%A')}"
        except Exception:
            pass
        if punch_state == "OUT":
            out_label = _format_punch_time(punch_out)
            return False, f"Punched out at {out_label}" if out_label else "Punched out"
        return False, "Not punched in"
    
    # Convert check times to minutes
    check_in = _coerce_to_time_obj(check_in_time)
    check_out = _coerce_to_time_obj(check_out_time)
    
    if check_in is None or check_out is None:
        return True, ""  # Can't determine, assume available
    
    check_in_min = check_in.hour * 60 + check_in.minute
    check_out_min = check_out.hour * 60 + check_out.minute
    if check_out_min < check_in_min:
        check_out_min += 1440  # Overnight
    
    # Check time blocks first (overlap against the whole appointment window)
    try:
        today_str = now.strftime("%Y-%m-%d")
        for block in st.session_state.get("time_blocks", []):
            if str(block.get("date", "")).strip() != today_str:
                continue
            if str(block.get("assistant", "")).strip().upper() != assist_upper:
                continue
            start_t = _coerce_to_time_obj(block.get("start_time"))
            end_t = _coerce_to_time_obj(block.get("end_time"))
            if start_t is None or end_t is None:
                continue
            start_min = start_t.hour * 60 + start_t.minute
            end_min = end_t.hour * 60 + end_t.minute
            if end_min < start_min:
                end_min += 1440
            if not (check_out_min <= start_min or check_in_min >= end_min):
                return False, f"Blocked: {block.get('reason', 'Blocked')}"
    except Exception:
        pass
    
    # Check existing appointments
    schedule = get_assistant_schedule(assist_upper, df_schedule)
    
    for appt in schedule:
        # Skip if it's the same row we're editing
        if exclude_row_id and str(appt.get("row_id", "")).strip() == str(exclude_row_id).strip():
            continue
        
        appt_in = _coerce_to_time_obj(appt.get("in_time"))
        appt_out = _coerce_to_time_obj(appt.get("out_time"))
        
        if appt_in is None or appt_out is None:
            continue
        
        appt_in_min = appt_in.hour * 60 + appt_in.minute
        appt_out_min = appt_out.hour * 60 + appt_out.minute
        if appt_out_min < appt_in_min:
            appt_out_min += 1440
        
        # Check for overlap
        if not (check_out_min <= appt_in_min or check_in_min >= appt_out_min):
            return False, f"With {appt.get('patient', 'patient')} ({appt_in.strftime('%I:%M %p')}-{appt_out.strftime('%I:%M %p')})"
    
    return True, ""
def _remove_assistant_assignments(df_schedule: Optional[DataFrame], assistant_name: str) -> Optional[DataFrame]:
    """Clear all allotments for an assistant (FIRST/SECOND/Third). Returns updated DF or None if no change."""
    if df_schedule is None or df_schedule.empty:
        return None
    assist_upper = str(assistant_name or "").strip().upper()
    if not assist_upper:
        return None
    df_updated = df_schedule.copy()
    third_col = _get_third_column_name(df_updated.columns)
    cols = ["FIRST", "SECOND", third_col]
    changed = False
    for col in cols:
        if not col or col not in df_updated.columns:
            continue
        mask = df_updated[col].astype(str).str.strip().str.upper() == assist_upper
        if mask.any():
            df_updated.loc[mask, col] = ""
            changed = True
    return df_updated if changed else None
def _pref_allows_role(value: Any) -> bool:
    try:
        s = str(value or "").strip().lower()
    except Exception:
        return True
    if not s:
        return True
    if s in {"no", "n", "false", "0", "off"}:
        return False
    if s in {"yes", "y", "true", "1", "on"}:
        return True
    return True
def _to_float(value: Any) -> Optional[float]:
    try:
        return float(value)
    except Exception:
        return None
def _normalize_name_list(values: Any) -> list[str]:
    if values is None:
        return []
    if isinstance(values, (list, tuple, set)):
        items = list(values)
    else:
        items = [values]
    return _unique_preserve_order([str(x) for x in items if str(x).strip()])
def _get_third_column_name(columns: Any) -> str:
    try:
        if "Third" in columns:
            return "Third"
        if "THIRD" in columns:
            return "THIRD"
    except Exception:
        pass
    return "Third"
def _collect_time_overrides(time_overrides: Any) -> list[tuple[float, list[str]]]:
    overrides: list[tuple[float, list[str]]] = []
    if time_overrides is None:
        return overrides
    if isinstance(time_overrides, dict):
        if "after_hour" in time_overrides:
            after = _to_float(time_overrides.get("after_hour"))
            assistants = _normalize_name_list(
                time_overrides.get("assistant") or time_overrides.get("assistants")
            )
            if after is not None and assistants:
                overrides.append((after, assistants))
        else:
            for key, val in time_overrides.items():
                after = _to_float(key)
                assistants = _normalize_name_list(val)
                if after is not None and assistants:
                    overrides.append((after, assistants))
    elif isinstance(time_overrides, list):
        for item in time_overrides:
            if isinstance(item, dict):
                after = _to_float(item.get("after_hour"))
                assistants = _normalize_name_list(item.get("assistant") or item.get("assistants"))
                if after is not None and assistants:
                    overrides.append((after, assistants))
            elif isinstance(item, (list, tuple)) and len(item) >= 2:
                after = _to_float(item[0])
                assistants = _normalize_name_list(item[1])
                if after is not None and assistants:
                    overrides.append((after, assistants))
    return overrides
def _time_override_candidates(time_overrides: Any, appt_hour: float) -> list[str]:
    overrides = _collect_time_overrides(time_overrides)
    matched = [(after, names) for after, names in overrides if appt_hour >= after]
    matched.sort(key=lambda item: item[0], reverse=True)
    out: list[str] = []
    for _, names in matched:
        out.extend(names)
    return _unique_preserve_order(out)
def _rule_candidates_for_role(
    role: str,
    rule: dict[str, Any],
    doctor: str,
    appt_hour: float,
    first_assistant: str,
) -> list[str]:
    if not isinstance(rule, dict):
        return []
    candidates: list[str] = []
    if role == "SECOND":
        when_map = rule.get("when_first_is", {})
        if isinstance(when_map, dict) and first_assistant:
            first_key = _norm_staff_key(first_assistant)
            for key, val in when_map.items():
                if _norm_staff_key(key) == first_key:
                    candidates.extend(_normalize_name_list(val))
                    break
    doctor_key = _norm_staff_key(doctor)
    doc_list = None
    doctor_overrides = rule.get("doctor_overrides", {})
    if isinstance(doctor_overrides, dict):
        for key, val in doctor_overrides.items():
            if _norm_staff_key(key) == doctor_key:
                doc_list = val
                break
    if doc_list is None:
        for key, val in rule.items():
            if key in {"default", "time_override", "when_first_is", "doctor_overrides"}:
                continue
            if _norm_staff_key(key) == doctor_key:
                doc_list = val
                break
    if doc_list is not None:
        candidates.extend(_normalize_name_list(doc_list))
    if "time_override" in rule:
        candidates.extend(_time_override_candidates(rule.get("time_override"), appt_hour))
    candidates.extend(_normalize_name_list(rule.get("default", [])))
    return _unique_preserve_order(candidates)
def _assistant_loads(df_schedule: pd.DataFrame, exclude_row_id: Optional[str] = None) -> dict[str, int]:
    counts: dict[str, int] = {}
    if df_schedule is None or df_schedule.empty:
        return counts
    third_col = _get_third_column_name(df_schedule.columns)
    for _, row in df_schedule.iterrows():
        if exclude_row_id and str(row.get("REMINDER_ROW_ID", "")).strip() == str(exclude_row_id).strip():
            continue
        for col in ["FIRST", "SECOND", third_col]:
            name = str(row.get(col, "")).strip().upper()
            if not name:
                continue
            counts[name] = counts.get(name, 0) + 1
    return counts
def _order_by_load(names: list[str], load_map: dict[str, int]) -> list[str]:
    if not names:
        return names
    order = {name: idx for idx, name in enumerate(names)}
    return sorted(names, key=lambda n: (load_map.get(n, 0), order.get(n, 0)))
def _select_assistant_from_candidates(
    role: str,
    candidates: list[str],
    available_map: dict[str, str],
    available_order: list[str],
    already: set[str],
    pref_map: dict[str, dict[str, Any]],
    use_role_flags: bool,
    load_map: dict[str, int],
    load_balance: bool,
) -> str:
    filtered: list[str] = []
    for name in candidates:
        key = str(name).strip().upper()
        if not key or key in already:
            continue
        if key not in available_map:
            continue
        if use_role_flags:
            pref_val = pref_map.get(_norm_staff_key(key), {}).get(role, "")
            if not _pref_allows_role(pref_val):
                continue
        filtered.append(key)
    if filtered and load_balance:
        filtered = _order_by_load(filtered, load_map)
    if filtered:
        return available_map[filtered[0]]
    fallback: list[str] = []
    for name in available_order:
        key = str(name).strip().upper()
        if not key or key in already:
            continue
        if use_role_flags:
            pref_val = pref_map.get(_norm_staff_key(key), {}).get(role, "")
            if not _pref_allows_role(pref_val):
                continue
        if key in available_map:
            fallback.append(key)
    if fallback and load_balance:
        fallback = _order_by_load(fallback, load_map)
    if fallback:
        return available_map[fallback[0]]
    return ""
def _allocate_assistants_for_slot(
    doctor: str,
    department: str,
    in_time: Any,
    out_time: Any,
    df_schedule: pd.DataFrame,
    exclude_row_id: Optional[str] = None,
    current_assignments: Optional[dict[str, Any]] = None,
    only_fill_empty: bool = False,
) -> dict[str, str]:
    result = {"FIRST": "", "SECOND": "", "Third": ""}
    if current_assignments:
        for role in result:
            val = current_assignments.get(role, "")
            result[role] = "" if _is_blank_cell(val) else str(val).strip()
    if not doctor:
        return result
    in_obj = _coerce_to_time_obj(in_time)
    out_obj = _coerce_to_time_obj(out_time)
    if in_obj is None or out_obj is None:
        return result
    appt_hour = in_obj.hour + in_obj.minute / 60.0
    config = _get_allocation_config()
    global_cfg = _get_global_allocation_config(config)
    dept_cfg = _get_allocation_department_config(department, config)
    rules = dept_cfg.get("allocation_rules", {}) if isinstance(dept_cfg, dict) else {}
    dept_assistants = get_assistants_for_department(department)
    all_assistants = _get_all_assistants()
    free_now_set, free_status_map = _get_dashboard_free_set(df_schedule, all_assistants)
    avail_dept = get_available_assistants(
        department,
        in_time,
        out_time,
        df_schedule,
        exclude_row_id,
        assistants_override=dept_assistants,
        free_now_set=free_now_set,
        free_status_map=free_status_map,
    )
    available_dept_order = [a["name"] for a in avail_dept if a.get("available")]
    available_dept_map = {name.upper(): name for name in available_dept_order}
    if global_cfg.get("cross_department_fallback", False):
        avail_all = get_available_assistants(
            department,
            in_time,
            out_time,
            df_schedule,
            exclude_row_id,
            assistants_override=all_assistants,
            free_now_set=free_now_set,
            free_status_map=free_status_map,
        )
        available_all_order = [a["name"] for a in avail_all if a.get("available")]
        available_all_map = {name.upper(): name for name in available_all_order}
    else:
        available_all_order = available_dept_order
        available_all_map = available_dept_map
    cache = _get_profiles_cache()
    pref_map = cache.get("assistant_prefs", {})
    load_map = _assistant_loads(df_schedule, exclude_row_id) if global_cfg.get("load_balance", False) else {}
    already = {
        str(x).strip().upper()
        for x in [result["FIRST"], result["SECOND"], result["Third"]]
        if x
    }
    for role in ["FIRST", "SECOND", "Third"]:
        if only_fill_empty and role in result and result[role]:
            continue
        rule = rules.get(role, {}) if isinstance(rules, dict) else {}
        candidates = _rule_candidates_for_role(role, rule, doctor, appt_hour, result.get("FIRST", ""))
        chosen = _select_assistant_from_candidates(
            role,
            candidates,
            available_dept_map,
            available_dept_order,
            already,
            pref_map,
            global_cfg.get("use_profile_role_flags", False),
            load_map,
            global_cfg.get("load_balance", False),
        )
        if not chosen and global_cfg.get("cross_department_fallback", False):
            chosen = _select_assistant_from_candidates(
                role,
                candidates,
                available_all_map,
                available_all_order,
                already,
                pref_map,
                global_cfg.get("use_profile_role_flags", False),
                load_map,
                global_cfg.get("load_balance", False),
            )
        if chosen:
            result[role] = chosen
            already.add(chosen.strip().upper())
    return result
def get_available_assistants(
    department: str,
    check_in_time: Any,
    check_out_time: Any,
    df_schedule: pd.DataFrame,
    exclude_row_id: Optional[str] = None,
    assistants_override: Optional[list[str]] = None,
    free_now_set: Optional[set[str]] = None,
    free_status_map: Optional[dict[str, dict[str, str]]] = None,
) -> list[dict[str, Any]]:
    """
    Get list of available assistants for a department at a specific time.
    Returns list of dicts with assistant name and availability status.
    """
    if assistants_override is not None:
        assistants = _unique_preserve_order(assistants_override)
    else:
        assistants = get_assistants_for_department(department)
    available = []
    
    for assistant in assistants:
        assist_upper = str(assistant).strip().upper()
        if free_now_set is not None and assist_upper not in free_now_set:
            reason = "Not available on dashboard"
            if isinstance(free_status_map, dict):
                info = free_status_map.get(assist_upper, {}) or {}
                status_label = str(info.get("status", "")).strip().upper()
                if info.get("reason"):
                    reason = str(info.get("reason"))
                elif status_label:
                    reason = f"Dashboard: {status_label}"
            available.append({
                "name": assistant,
                "available": False,
                "reason": reason,
            })
            continue
        is_avail, reason = is_assistant_available(assistant, check_in_time, check_out_time, df_schedule, exclude_row_id)
        available.append({
            "name": assistant,
            "available": is_avail,
            "reason": reason if not is_avail else "Available"
        })
    
    return available
def auto_allocate_assistants(
    doctor: str,
    in_time: Any,
    out_time: Any,
    df_schedule: pd.DataFrame,
    exclude_row_id: Optional[str] = None,
) -> dict[str, str]:
    """
    Automatically allocate assistants based on department and availability.
    Returns dict with FIRST, SECOND, Third assignments.
    """
    department = get_department_for_doctor(doctor)
    return _allocate_assistants_for_slot(
        doctor,
        department,
        in_time,
        out_time,
        df_schedule,
        exclude_row_id=exclude_row_id,
        current_assignments=None,
        only_fill_empty=False,
    )
def _auto_fill_assistants_for_row(df_schedule: pd.DataFrame, row_index: int, only_fill_empty: bool = True) -> bool:
    """Auto-fill FIRST/SECOND/Third for a single row based on doctor-specific and time-based allocation rules. Returns True if anything changed."""
    try:
        if row_index < 0 or row_index >= len(df_schedule):
            return False
        row = df_schedule.iloc[row_index]
        doctor = str(row.get("DR.", "")).strip()
        if _is_blank_cell(doctor):
            doctor = str(row.get("Doctor", "")).strip()
        in_time_val = row.get("In Time", None)
        out_time_val = row.get("Out Time", None)
        row_id = str(row.get("REMINDER_ROW_ID", "")).strip()
        if not doctor:
            return False
        if _coerce_to_time_obj(in_time_val) is None or _coerce_to_time_obj(out_time_val) is None:
            return False
        department = get_department_for_doctor(doctor)
        current_first = row.get("FIRST", "")
        current_second = row.get("SECOND", "")
        third_col = _get_third_column_name(df_schedule.columns)
        current_third = row.get(third_col, "")
        if only_fill_empty and (not _is_blank_cell(current_first)) and (not _is_blank_cell(current_second)) and (not _is_blank_cell(current_third)):
            return False
        allocations = _allocate_assistants_for_slot(
            doctor,
            department,
            in_time_val,
            out_time_val,
            df_schedule,
            exclude_row_id=row_id,
            current_assignments={
                "FIRST": current_first,
                "SECOND": current_second,
                "Third": current_third,
            },
            only_fill_empty=only_fill_empty,
        )
        changed = False
        for role, current_val in [("FIRST", current_first), ("SECOND", current_second), ("Third", current_third)]:
            new_val = allocations.get(role, "")
            if _is_blank_cell(new_val):
                continue
            if str(new_val).strip() != str(current_val).strip():
                if role == "Third":
                    if third_col in df_schedule.columns:
                        df_schedule.iloc[row_index, df_schedule.columns.get_loc(third_col)] = new_val
                else:
                    if role in df_schedule.columns:
                        df_schedule.iloc[row_index, df_schedule.columns.get_loc(role)] = new_val
                changed = True
        return changed
    except Exception:
        return False
def get_current_assistant_status(
    df_schedule: pd.DataFrame,
    assistants: Optional[list[str]] = None,
    punch_map: Optional[dict[str, dict[str, str]]] = None,
) -> dict[str, dict[str, str]]:
    """
    Get real-time status of all assistants.
    Returns dict with assistant name -> status info
    """
    status = {}
    if df_schedule is None:
        df_schedule = pd.DataFrame()
    if assistants is None:
        assistants = _get_all_assistants()
    if punch_map is None:
        try:
            punch_map = _get_today_punch_map()
        except Exception:
            punch_map = {}
    current_time = time_type(now.hour, now.minute)
    current_min = now.hour * 60 + now.minute
    today_weekday = now.weekday()
    weekday_name_list = globals().get("weekday_names", [])
    weekday_label = (
        weekday_name_list[today_weekday]
        if isinstance(weekday_name_list, list) and 0 <= today_weekday < len(weekday_name_list)
        else now.strftime("%A")
    )
    weekly_off_map = _get_profiles_cache_snapshot().get("weekly_off_map", WEEKLY_OFF)
    weekly_off_set = {
        str(name).strip().upper()
        for name in weekly_off_map.get(today_weekday, [])
        if str(name).strip()
    }
    
    for assistant in assistants:
        assist_upper = assistant.upper()
        punch_state, punch_in, punch_out = _assistant_punch_state(assist_upper, punch_map)
        if punch_state != "IN":
            if assist_upper in weekly_off_set:
                reason = f"Weekly off ({weekday_label})"
            elif punch_state == "OUT":
                out_label = _format_punch_time(punch_out)
                reason = f"Punched out at {out_label}" if out_label else "Punched out"
            else:
                reason = "Not punched in"
            status[assist_upper] = {
                "status": "BLOCKED",
                "reason": reason,
                "department": get_department_for_assistant(assist_upper),
            }
            continue
        
        # Check if blocked
        is_blocked, block_reason = is_assistant_blocked(assist_upper, current_time)
        if is_blocked:
            status[assist_upper] = {
                "status": "BLOCKED",
                "reason": block_reason,
                "department": get_department_for_assistant(assist_upper)
            }
            continue
        
        # Check current appointments
        schedule = get_assistant_schedule(assist_upper, df_schedule)
        current_appt = None
        
        for appt in schedule:
            status_text = str(appt.get("status", "")).upper()
            appt_in = _coerce_to_time_obj(appt.get("in_time"))
            appt_out = _coerce_to_time_obj(appt.get("out_time"))
            # If status explicitly says ON GOING, treat as busy regardless of time parsing.
            if "ON GOING" in status_text or "ONGOING" in status_text:
                current_appt = appt
                break
            # If timing is missing but status shows ARRIVED, treat as busy to avoid zero-count glitch.
            if (appt_in is None or appt_out is None) and "ARRIVED" in status_text:
                current_appt = appt
                break
            
            if appt_in is None or appt_out is None:
                continue
            
            appt_in_min = appt_in.hour * 60 + appt_in.minute
            appt_out_min = appt_out.hour * 60 + appt_out.minute
            if appt_out_min < appt_in_min:
                appt_out_min += 1440
            
            if appt_in_min <= current_min <= appt_out_min:
                current_appt = appt
                break
        
        # Check for active duty run (duty timer)
        try:
            duty_runs_df = load_duty_runs_sheet()
            active_duty = duty_runs_df[
                (duty_runs_df["assistant"].astype(str).str.strip() == assist_upper) &
                (duty_runs_df["status"].astype(str).str.upper() == "IN_PROGRESS")
            ] if not duty_runs_df.empty else pd.DataFrame()
        except Exception:
            active_duty = pd.DataFrame()
        if not active_duty.empty:
            # Assistant has active duty
            duty_run = active_duty.iloc[0]
            due_dt = _parse_iso_ts(duty_run.get("due_at"))
            remaining_time = ""
            if due_dt:
                delta = due_dt - now_ist()
                if delta.total_seconds() > 0:
                    mins = int(delta.total_seconds() // 60)
                    secs = int(delta.total_seconds() % 60)
                    remaining_time = f" • {mins:02d}:{secs:02d}"
            status[assist_upper] = {
                "status": "BUSY",
                "reason": f"On Duty{remaining_time}",
                "duty_run_id": duty_run.get("id"),
                "duty_id": duty_run.get("duty_id"),
                "due_at": duty_run.get("due_at"),
                "department": get_department_for_assistant(assist_upper)
            }
        elif current_appt:
            status[assist_upper] = {
                "status": "BUSY",
                "reason": f"With {current_appt.get('patient', 'patient')}",
                "patient": current_appt.get("patient", ""),
                "doctor": current_appt.get("doctor", ""),
                "op": current_appt.get("op", ""),
                "department": get_department_for_assistant(assist_upper)
            }
        else:
            status[assist_upper] = {
                "status": "FREE",
                "reason": "Available",
                "department": get_department_for_assistant(assist_upper)
            }
    
    return status
def _get_dashboard_free_set(
    df_schedule: pd.DataFrame,
    assistants: list[str],
) -> tuple[set[str], dict[str, dict[str, str]]]:
    try:
        status_map = get_current_assistant_status(df_schedule, assistants=assistants)
    except Exception:
        return set(), {}
    free_set = {
        name
        for name, info in status_map.items()
        if str(info.get("status", "")).strip().upper() == "FREE"
    }
    return free_set, status_map
STATUS_BADGES = {
    "FREE": {"label": "Free", "emoji": "🟢", "pill": "success", "card_class": "status-free", "default_detail": "Ready for assignment"},
    "BUSY": {"label": "Busy", "emoji": "🔴", "pill": "warning", "card_class": "status-busy", "default_detail": "In procedure"},
    "BLOCKED": {"label": "Blocked", "emoji": "🚫", "pill": "danger", "card_class": "status-blocked", "default_detail": "Unavailable"},
    "UNKNOWN": {"label": "Unknown", "emoji": "❔", "pill": "info", "card_class": "status-unknown", "default_detail": "No schedule"},
}
def _render_availability_summary(total: int, free: int, busy: int, blocked: int) -> None:
    """Render availability summary using native Streamlit components."""
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric(label="👥 Total Assistants", value=total, help="Rostered today")
    with col2:
        st.metric(label="🟢 Free", value=free, help="Ready for allocation")
    with col3:
        st.metric(label="🔴 Busy", value=busy, help="Currently chairside")
    with col4:
        st.metric(label="🚫 Blocked", value=blocked, help="Weekly off / hold")
def _render_assistant_cards(card_entries: list[dict[str, Any]]) -> None:
    """Render assistant cards using native Streamlit components."""
    if not card_entries:
        st.info("No assistants match the selected filters.")
        return
    # Create rows of 4 cards each
    cards_per_row = 4
    for i in range(0, len(card_entries), cards_per_row):
        row_entries = card_entries[i:i + cards_per_row]
        cols = st.columns(cards_per_row)
        
        for j, entry in enumerate(row_entries):
            with cols[j]:
                assistant_name = str(entry.get("name", "Assistant"))
                info = entry.get("info", {}) or {}
                status_raw = str(info.get("status", "UNKNOWN")).upper()
                meta = STATUS_BADGES.get(status_raw, STATUS_BADGES["UNKNOWN"])
                reason = str(info.get("reason", "")).strip()
                patient = str(info.get("patient", "")).strip()
                doctor = str(info.get("doctor", "")).strip()
                op_room = str(info.get("op", "")).strip()
                department = str(info.get("department", "")) or "—"
                # Build detail text
                detail_lines: list[str] = []
                if status_raw == "BUSY" and patient:
                    detail_lines.append(f"With {patient}")
                elif reason:
                    detail_lines.append(reason)
                else:
                    detail_lines.append(meta.get("default_detail", ""))
                if doctor and (status_raw == "BUSY" or not patient):
                    detail_lines.append(f"Doctor: {doctor}")
                if op_room:
                    detail_lines.append(f"OP: {op_room}")
                detail_text = " | ".join(line for line in detail_lines if line)
                # Use expander for card-like appearance
                status_emoji = meta["emoji"]
                status_label = meta["label"]
                
                with st.container(border=True):
                    st.markdown(f"**{assistant_name}**")
                    st.caption(f"{status_emoji} {status_label}")
                    if detail_text:
                        st.write(detail_text)
                    st.caption(f"Dept: {department}")
# --- Reminder settings in sidebar ---
# --- Sidebar: Notifications & Auto-Allotment ---
with st.sidebar:
    st.markdown("## 🔔 Notifications")
    st.checkbox("Enable 15-minute reminders", value=True, key="enable_reminders")
    st.checkbox(
        "Run alerts on all pages",
        value=False,
        key="alerts_background",
        help="When off, reminders and status alerts run only on the Scheduling page for smoother UX.",
    )
    st.selectbox(
        "Default snooze (seconds)",
        options=[30, 60, 90, 120, 150, 180, 300],
        index=0,
        key="default_snooze_seconds",
    )
    st.write("💡 Reminders alert 15 minutes before a patient's In Time.")
    st.markdown("---")
    st.markdown("## 🤖 Auto-Allotment")
    st.session_state.auto_assign_assistants = st.checkbox(
        "Auto-assign assistants",
        value=st.session_state.get("auto_assign_assistants", True),
        help="Automatically fill FIRST/SECOND/Third based on rules and availability."
    )
    st.session_state.auto_assign_only_empty = st.checkbox(
        "Only fill empty slots",
        value=st.session_state.get("auto_assign_only_empty", True),
        help="If enabled, only empty assistant slots will be auto-filled."
    )
# ================ WEEKLY OFF DISPLAY ================
with st.sidebar:
    st.markdown("---")
    st.markdown("## 📋 Assistant Weekly Off Schedule")
    
    today_weekday = now.weekday()  # 0=Monday, 6=Sunday
    weekday_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    today_name = weekday_names[today_weekday]
    weekly_off_map = _get_profiles_cache_snapshot().get("weekly_off_map", WEEKLY_OFF)
    
    # TODAY'S OFF
    today_off = weekly_off_map.get(today_weekday, [])
    st.markdown("**Today:**")
    if today_off:
        off_text = ", ".join(today_off)
        st.warning(f"🔴 {off_text}")
        st.caption("Cannot be allocated today.")
    else:
        st.success(f"✅ All assistants available")
    
    # TOMORROW'S OFF
    tomorrow_weekday = (today_weekday + 1) % 7  # Next day, wrap around if Sunday
    tomorrow_name = weekday_names[tomorrow_weekday]
    tomorrow_off = weekly_off_map.get(tomorrow_weekday, [])
    
    st.markdown("**Tomorrow:**")
    if tomorrow_off:
        off_text = ", ".join(tomorrow_off)
        st.info(f"ℹ️ {tomorrow_name}: {off_text}")
        st.caption("Will be off tomorrow - plan accordingly.")
    else:
        st.success(f"✅ {tomorrow_name}: All assistants available")
with st.sidebar:
    st.markdown("---")
# ================ Data Storage Configuration ================
# REMOVED DUPLICATES: All storage configuration moved to top of file
# SECURITY FIX: Removed hardcoded Supabase credentials
def _safe_secret_get(key: str, default=None):
    """Safely read st.secrets in all environments."""
    try:
        return st.secrets.get(key, default)
    except Exception:
        return default
def _as_bool(val) -> bool:
    try:
        return str(val).strip().lower() in {"1", "true", "yes", "on"}
    except Exception:
        return False
def _ensure_profile_df(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in PROFILE_COLUMNS:
        if col not in out.columns:
            out[col] = ""
    # Normalize text casing
    out["name"] = out["name"].astype(str).str.upper()
    out["department"] = out["department"].astype(str).str.upper()
    return out[PROFILE_COLUMNS]
def _now_iso():
    """Get current time in IST as ISO string."""
    return now_ist().isoformat(timespec="seconds")
def _profiles_table_setup_sql(table_name: str) -> str:
    table = table_name or "profiles"
    return (
        f"create table if not exists {table} (\n"
        "  id text primary key,\n"
        "  kind text not null,\n"
        "  name text not null,\n"
        "  department text,\n"
        "  contact_email text,\n"
        "  contact_phone text,\n"
        "  status text,\n"
        "  weekly_off text,\n"
        "  pref_first text,\n"
        "  pref_second text,\n"
        "  pref_third text,\n"
        "  created_at timestamptz,\n"
        "  updated_at timestamptz,\n"
        "  created_by text,\n"
        "  updated_by text\n"
        ");\n"
    )
@st.cache_data(ttl=30)
def _profiles_table_ready(_supabase, table_name: str) -> tuple[bool, str]:
    if not _supabase or not table_name:
        return False, "Supabase client is not available."
    try:
        _supabase.table(table_name).select("id,kind,name").limit(1).execute()
        return True, ""
    except Exception as e:
        return False, str(e)
def _render_profiles_setup_help(table_name: str, err: Optional[str] = None) -> None:
    st.error("Supabase profiles table is missing or misconfigured.")
    if err:
        st.caption(f"Details: {err}")
    st.markdown("Create the table in Supabase SQL Editor:")
    st.code(_profiles_table_setup_sql(table_name), language="sql")
    st.markdown(
        "If you use an anon key, add RLS policies that allow read and write, "
        "or use a service role key."
    )
def load_profiles(sheet_name: str) -> pd.DataFrame:
    """Load assistant/doctor profiles (Supabase-first).
    Performance: Results are cached by _load_profiles_cached wrapper
    """
    if USE_SUPABASE and supabase_client is not None:
        try:
            # Optimized: Only fetch needed columns for faster query
            resp = (
                supabase_client.table(PROFILE_SUPABASE_TABLE)
                .select("id,name,kind,department,status,weekly_off,pref_first,pref_second,pref_third")
                .eq("kind", sheet_name)
                .execute()
            )
            data = resp.data or []
            df = pd.DataFrame(data)
            if df.empty:
                return _ensure_profile_df(pd.DataFrame())
            # Coerce helper columns
            df["name"] = df["name"].astype(str).str.upper()
            df["department"] = df.get("department", "").astype(str).str.upper()
            # Ensure preference/weekly_off columns exist
            if "weekly_off" not in df.columns:
                df["weekly_off"] = ""
            if "pref_first" not in df.columns:
                df["pref_first"] = ""
            if "pref_second" not in df.columns:
                df["pref_second"] = ""
            if "pref_third" not in df.columns:
                df["pref_third"] = ""
            return _ensure_profile_df(df)
        except Exception:
            return _ensure_profile_df(pd.DataFrame())
    try:
        if not os.path.exists(file_path):
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            wb.create_sheet(sheet_name)
            wb.save(file_path)
        try:
            wb = openpyxl.load_workbook(file_path)
        except (zipfile.BadZipFile, KeyError, Exception):
            # Repair a corrupted workbook by recreating it
            # (catches [Content_Types].xml missing errors and other corruption)
            try:
                os.remove(file_path)
            except Exception:
                pass
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            wb.create_sheet(sheet_name)
            wb.save(file_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(PROFILE_COLUMNS)
            wb.save(file_path)
        ws = wb[sheet_name]
        data = list(ws.values)
        if len(data) > 1:
            df = pd.DataFrame(data[1:], columns=data[0])
        elif len(data) == 1:
            df = pd.DataFrame(columns=data[0])
        else:
            # Sheet is completely empty, use PROFILE_COLUMNS as default
            df = pd.DataFrame(columns=PROFILE_COLUMNS)
        return _ensure_profile_df(df)
    except Exception as e:
        st.error(f"Error loading profiles '{sheet_name}': {e}")
        return _ensure_profile_df(pd.DataFrame())
def save_profiles(df: pd.DataFrame, sheet_name: str) -> bool:
    """Persist assistant/doctor profiles (Supabase-first). Returns True on success."""
    if USE_SUPABASE and supabase_client is not None:
        try:
            clean_df = _ensure_profile_df(df)
            clean_df = clean_df.where(pd.notna(clean_df), None)
            if "id" in clean_df.columns:
                ids = clean_df["id"].astype(str)
                missing = clean_df["id"].isna() | ids.str.strip().isin(["", "nan", "none"])
                if missing.any():
                    clean_df.loc[missing, "id"] = [str(uuid.uuid4()) for _ in range(int(missing.sum()))]
            clean_df["kind"] = sheet_name
            # Flatten weekly_off lists if present
            def _fmt_wo(val):
                if isinstance(val, list):
                    return ",".join([str(v) for v in val if str(v).strip()])
                return str(val or "")
            clean_df["weekly_off"] = clean_df["weekly_off"].apply(_fmt_wo)
            # Upsert per row
            rows = clean_df.to_dict(orient="records")
            for row in rows:
                rid = row.get("id")
                if rid:
                    res = supabase_client.table(PROFILE_SUPABASE_TABLE).upsert(row).execute()
                else:
                    res = supabase_client.table(PROFILE_SUPABASE_TABLE).insert(row).execute()
                err = getattr(res, "error", None)
                if err:
                    raise RuntimeError(str(err))
            try:
                _get_active_assistant_profile_names.clear()
            except Exception:
                pass
            return True
        except Exception as e:
            st.error(f"Error saving profiles to Supabase '{sheet_name}': {e}")
            st.info("Ensure the profiles table exists and has all required columns.")
            st.code(_profiles_table_setup_sql(PROFILE_SUPABASE_TABLE), language="sql")
            return False
    try:
        clean_df = _ensure_profile_df(df)
        clean_df = clean_df.where(pd.notna(clean_df), None)
        if "id" in clean_df.columns:
            ids = clean_df["id"].astype(str)
            missing = clean_df["id"].isna() | ids.str.strip().isin(["", "nan", "none"])
            if missing.any():
                clean_df.loc[missing, "id"] = [str(uuid.uuid4()) for _ in range(int(missing.sum()))]
        try:
            wb = openpyxl.load_workbook(file_path)
        except (zipfile.BadZipFile, KeyError, Exception):
            wb = openpyxl.Workbook()
        # Use ExcelWriter to write the sheet (replaces if exists, creates if not)
        # Use mode='a' (append) with if_sheet_exists='replace' to keep other sheets intact
        with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            clean_df.to_excel(writer, sheet_name=sheet_name, index=False)
        # After saving, reload and ensure at least one sheet is visible
        try:
            wb = openpyxl.load_workbook(file_path)
            if not any(ws.sheet_state == 'visible' for ws in wb.worksheets):
                # If no sheets are visible, make the first one visible
                if wb.sheetnames:
                    wb[wb.sheetnames[0]].sheet_state = 'visible'
                    wb.save(file_path)
        except Exception:
            pass  # If we can't fix visibility, that's ok
        try:
            _get_active_assistant_profile_names.clear()
        except Exception:
            pass
        return True
    except Exception as e:
        st.error(f"Error saving profiles '{sheet_name}': {e}")
        return False
@st.cache_data(ttl=600, show_spinner="Loading profiles...")
def _load_profiles_cached(sheet_name: str, cache_bust: int) -> pd.DataFrame:
    """Load profiles with aggressive caching (10 minutes).
    Profiles don't change frequently, so we cache for longer to improve performance.
    Use cache_bust parameter to force refresh when needed.
    """
    return load_profiles(sheet_name)
def _is_active_status(value: Any) -> bool:
    try:
        s = str(value or "").strip().upper()
    except Exception:
        return True
    return (not s) or s == "ACTIVE"
def _get_profiles_cache() -> dict[str, Any]:
    cache_bust = int(st.session_state.get("profiles_cache_bust", 0))
    cached = st.session_state.get("profiles_cache", {})
    if isinstance(cached, dict) and cached.get("cache_bust") == cache_bust:
        return cached
    assistants_df = _load_profiles_cached(PROFILE_ASSISTANT_SHEET, cache_bust)
    doctors_df = _load_profiles_cached(PROFILE_DOCTOR_SHEET, cache_bust)
    if assistants_df is None:
        assistants_df = _ensure_profile_df(pd.DataFrame())
    if doctors_df is None:
        doctors_df = _ensure_profile_df(pd.DataFrame())
    assistants_df = _ensure_profile_df(assistants_df)
    doctors_df = _ensure_profile_df(doctors_df)
    config = _get_allocation_config()
    config_maps = _get_config_department_maps(config)
    config_doctor_map = config_maps.get("doctors", {})
    config_assistant_map = config_maps.get("assistants", {})
    assistants_list: list[str] = []
    assistant_dept_map: dict[str, str] = {}
    assistant_pref_map: dict[str, dict[str, Any]] = {}
    weekly_off_map: dict[int, list[str]] = {i: [] for i in range(7)}
    for _, row in assistants_df.iterrows():
        name = str(row.get("name", "")).strip().upper()
        if not name:
            continue
        if "status" in assistants_df.columns and not _is_active_status(row.get("status", "")):
            continue
        assistants_list.append(name)
        key = _norm_staff_key(name)
        dept = str(row.get("department", "")).strip().upper()
        if not dept:
            dept = config_assistant_map.get(key, "")
        if not dept:
            dept = "SHARED"
        assistant_dept_map[key] = dept
        assistant_pref_map[key] = {
            "FIRST": row.get("pref_first", ""),
            "SECOND": row.get("pref_second", ""),
            "Third": row.get("pref_third", ""),
        }
        try:
            for idx in _parse_weekly_off_days(row.get("weekly_off", "")):
                weekly_off_map[idx].append(name)
        except Exception:
            pass
    doctors_list: list[str] = []
    doctor_dept_map: dict[str, str] = {}
    for _, row in doctors_df.iterrows():
        name = str(row.get("name", "")).strip().upper()
        if not name:
            continue
        if "status" in doctors_df.columns and not _is_active_status(row.get("status", "")):
            continue
        doctors_list.append(name)
        key = _norm_staff_key(name)
        dept = str(row.get("department", "")).strip().upper()
        if not dept:
            dept = config_doctor_map.get(key, "")
        if dept:
            doctor_dept_map[key] = dept
    assistants_list = _unique_preserve_order(assistants_list)
    doctors_list = _unique_preserve_order(doctors_list)
    dept_set = set(config_maps.get("departments", []) or [])
    dept_set.update([d for d in assistant_dept_map.values() if d])
    dept_set.update([d for d in doctor_dept_map.values() if d])
    if not dept_set:
        dept_set.update([str(d).strip().upper() for d in DEPARTMENTS.keys()])
    def _build_config_lists(key: str) -> dict[str, list[str]]:
        out: dict[str, list[str]] = {}
        depts = config.get("departments", {}) if isinstance(config, dict) else {}
        if isinstance(depts, dict):
            for dept_name, data in depts.items():
                if not isinstance(data, dict):
                    continue
                dept_upper = str(dept_name).strip().upper()
                if not dept_upper:
                    continue
                raw_list = data.get(key, []) or []
                out[dept_upper] = _unique_preserve_order(raw_list)
        return out
    config_assistant_lists = _build_config_lists("assistants")
    config_doctor_lists = _build_config_lists("doctors")
    assistants_by_dept: dict[str, list[str]] = {dept: [] for dept in dept_set}
    if config_assistant_lists:
        for dept, ordered in config_assistant_lists.items():
            for name in ordered:
                if name in assistants_list and name not in assistants_by_dept.setdefault(dept, []):
                    assistants_by_dept[dept].append(name)
    for name in assistants_list:
        dept = assistant_dept_map.get(_norm_staff_key(name), "")
        if not dept:
            continue
        if name not in assistants_by_dept.setdefault(dept, []):
            assistants_by_dept[dept].append(name)
    doctors_by_dept: dict[str, list[str]] = {dept: [] for dept in dept_set}
    if config_doctor_lists:
        for dept, ordered in config_doctor_lists.items():
            for name in ordered:
                if name in doctors_list and name not in doctors_by_dept.setdefault(dept, []):
                    doctors_by_dept[dept].append(name)
    for name in doctors_list:
        dept = doctor_dept_map.get(_norm_staff_key(name), "")
        if not dept:
            continue
        if name not in doctors_by_dept.setdefault(dept, []):
            doctors_by_dept[dept].append(name)
    global ALL_ASSISTANTS, ALL_DOCTORS, WEEKLY_OFF
    if assistants_list:
        ALL_ASSISTANTS = assistants_list
    if doctors_list:
        ALL_DOCTORS = doctors_list
    if assistants_list:
        WEEKLY_OFF = weekly_off_map
    cache = {
        "cache_bust": cache_bust,
        "assistants": assistants_list,
        "doctors": doctors_list,
        "assistant_dept_map": assistant_dept_map,
        "doctor_dept_map": doctor_dept_map,
        "assistant_prefs": assistant_pref_map,
        "weekly_off_map": weekly_off_map,
        "departments": sorted([d for d in dept_set if d]),
        "assistants_by_dept": assistants_by_dept,
        "doctors_by_dept": doctors_by_dept,
    }
    st.session_state.profiles_cache = cache
    return cache
def _get_known_departments() -> list[str]:
    try:
        cache = _get_profiles_cache()
        depts = cache.get("departments", [])
        if depts:
            return depts
    except Exception:
        pass
    config = _get_allocation_config()
    dept_list = []
    depts = config.get("departments", {}) if isinstance(config, dict) else {}
    if isinstance(depts, dict):
        for dept in depts.keys():
            dept_upper = str(dept).strip().upper()
            if dept_upper and dept_upper not in dept_list:
                dept_list.append(dept_upper)
    if dept_list:
        return sorted(dept_list)
    return sorted([str(d).strip().upper() for d in DEPARTMENTS.keys() if str(d).strip()])
def _get_all_doctors() -> list[str]:
    try:
        cache = _get_profiles_cache()
        doctors = cache.get("doctors", [])
        if doctors:
            return doctors
    except Exception:
        pass
    config = _get_allocation_config()
    out: list[str] = []
    depts = config.get("departments", {}) if isinstance(config, dict) else {}
    if isinstance(depts, dict):
        for data in depts.values():
            if not isinstance(data, dict):
                continue
            out.extend(data.get("doctors", []) or [])
    out = _unique_preserve_order(out)
    if out:
        return out
    return ALL_DOCTORS
def _get_all_assistants() -> list[str]:
    try:
        cache = _get_profiles_cache()
        assistants = cache.get("assistants", [])
        if assistants:
            return assistants
    except Exception:
        pass
    config = _get_allocation_config()
    out: list[str] = []
    depts = config.get("departments", {}) if isinstance(config, dict) else {}
    if isinstance(depts, dict):
        for data in depts.values():
            if not isinstance(data, dict):
                continue
            out.extend(data.get("assistants", []) or [])
    out = _unique_preserve_order(out)
    if out:
        return out
    return ALL_ASSISTANTS
def _restore_profile_hidden_columns(
    edited_df: pd.DataFrame,
    base_df: pd.DataFrame,
    hidden_cols: list[str],
    user_name: str,
) -> pd.DataFrame:
    out = edited_df.copy()
    for col in hidden_cols:
        if col not in out.columns:
            out[col] = ""
    if "id" not in out.columns:
        out["id"] = ""
    base_id_map: dict[str, dict[str, Any]] = {}
    if "id" in base_df.columns:
        for _, row in base_df.iterrows():
            rid = str(row.get("id", "")).strip()
            if not rid or rid.lower() in {"nan", "none"}:
                continue
            base_id_map[rid] = row.to_dict()
    if base_id_map and "name" in out.columns and "department" in out.columns:
        base_key = (
            base_df["name"].astype(str).str.strip().str.upper()
            + "|"
            + base_df["department"].astype(str).str.strip().str.upper()
        )
        base_keys = dict(zip(base_key, base_df["id"].astype(str)))
        missing_id = out["id"].apply(_is_blank_cell)
        if missing_id.any():
            out_key = (
                out["name"].astype(str).str.strip().str.upper()
                + "|"
                + out["department"].astype(str).str.strip().str.upper()
            )
            out.loc[missing_id, "id"] = out_key[missing_id].map(base_keys).fillna("")
    if base_id_map:
        for col in hidden_cols:
            mask = out[col].apply(_is_blank_cell)
            if not mask.any():
                continue
            out.loc[mask, col] = out.loc[mask, "id"].map(
                lambda rid: base_id_map.get(str(rid).strip(), {}).get(col, "")
            )
    now_iso = _now_iso()
    if "created_at" in out.columns:
        mask = out["created_at"].apply(_is_blank_cell)
        if mask.any():
            out.loc[mask, "created_at"] = now_iso
    if "created_by" in out.columns:
        mask = out["created_by"].apply(_is_blank_cell)
        if mask.any():
            out.loc[mask, "created_by"] = user_name
    return out
def render_profile_manager(sheet_name: str, entity_label: str, dept_label: str) -> None:
    """UI to add/edit assistant/doctor profiles with simple role guard."""
    user_role = st.session_state.get("user_role", "viewer")
    user_name = st.session_state.get("current_user", "user")
    if USE_SUPABASE and supabase_client is not None:
        ready, err = _profiles_table_ready(supabase_client, PROFILE_SUPABASE_TABLE)
        if not ready:
            _render_profiles_setup_help(PROFILE_SUPABASE_TABLE, err)
            return
    df_profiles = load_profiles(sheet_name)
    status_options = ["ACTIVE", "INACTIVE"]
    dept_options = [""] + _get_known_departments()
    hidden_cols = ["id", "created_at", "updated_at", "created_by", "updated_by"]
    is_editor = user_role in ("admin", "editor")
    st.markdown(f"### {entity_label} Profiles")
    if not is_editor:
        # Filters (applied to the read-only view)
        f1, f2, f3 = st.columns([0.2, 0.2, 0.6])
        with f1:
            status_filter = st.multiselect(
                "Status",
                options=status_options,
                default=["ACTIVE"],
                key=f"{sheet_name}_status_filter",
            )
        with f2:
            dept_filter = st.selectbox(
                dept_label,
                options=["All"] + dept_options[1:],
                key=f"{sheet_name}_dept_filter",
            )
        with f3:
            search_term = st.text_input("Search name", key=f"{sheet_name}_search")
        filtered = df_profiles.copy()
        if status_filter:
            filtered = filtered[filtered["status"].isin(status_filter)]
        if dept_filter and dept_filter != "All":
            filtered = filtered[filtered["department"].str.upper() == dept_filter.upper()]
        if search_term:
            filtered = filtered[filtered["name"].str.contains(search_term, case=False, na=False)]
        display_filtered = filtered.drop(columns=[c for c in hidden_cols if c in filtered.columns], errors="ignore")
        st.dataframe(display_filtered, use_container_width=True, hide_index=True)
        st.info("You are in read-only mode. Switch to admin/editor to add or edit profiles.")
        return
    def _render_add_profile_dialog_body() -> None:
        st.markdown(f"### Add {entity_label}")
        with st.form(f"add_{sheet_name}_form", clear_on_submit=False):
            name = st.text_input(f"{entity_label} Name")
            dept = st.selectbox(dept_label, options=dept_options, key=f"{sheet_name}_dept_new")
            contact_email = st.text_input("Contact Email", key=f"{sheet_name}_email_new")
            contact_phone = st.text_input("Contact Phone", key=f"{sheet_name}_phone_new")
            status_val = st.selectbox("Status", options=status_options, key=f"{sheet_name}_status_new")
            submitted = st.form_submit_button(f"Add {entity_label}")
            if submitted:
                if not name.strip():
                    st.warning("Name is required.")
                else:
                    new_row = {
                        "id": str(uuid.uuid4()),
                        "name": name.strip(),
                        "department": dept.strip(),
                        "contact_email": contact_email.strip(),
                        "contact_phone": contact_phone.strip(),
                        "status": status_val,
                        "created_at": _now_iso(),
                        "updated_at": _now_iso(),
                        "created_by": user_name,
                        "updated_by": user_name,
                    }
                    df_profiles_local = pd.concat([df_profiles, pd.DataFrame([new_row])], ignore_index=True)
                    ok = save_profiles(df_profiles_local, sheet_name)
                    if not ok:
                        st.error(f"Failed to save {entity_label}.")
                        return
                    st.session_state.profiles_cache_bust += 1
                    if USE_SUPABASE and supabase_client is not None:
                        st.session_state.supabase_staff_refreshed = False
                    st.success(f"{entity_label} added.")
                    st.rerun()
    _dialog_decorator = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)
    if _dialog_decorator:
        @_dialog_decorator(f"Add {entity_label}")
        def _render_add_profile_dialog() -> None:
            _render_add_profile_dialog_body()
    else:
        def _render_add_profile_dialog() -> None:
            st.warning("Popup add requires a newer Streamlit version.")
            _render_add_profile_dialog_body()
    if st.button(f"Add {entity_label}", key=f"add_{sheet_name}_open", use_container_width=True):
        _render_add_profile_dialog()
    def _render_delete_profile_dialog_body() -> None:
        st.markdown(f"### Delete {entity_label} Profiles")
        if df_profiles.empty:
            st.caption("No profiles available to delete.")
            return
        option_meta: dict[str, dict[str, Any]] = {}
        delete_options: list[str] = []
        for idx, row in df_profiles.iterrows():
            name = str(row.get("name", "")).strip()
            dept = str(row.get("department", "")).strip()
            rid = str(row.get("id", "")).strip()
            label_parts = [name.title() if name else f"Row {idx + 1}"]
            if dept:
                label_parts.append(dept.title())
            label = " - ".join(label_parts)
            if rid:
                label = f"{label} ({rid[-6:]})"
            else:
                label = f"{label} (row {idx + 1})"
            if label in option_meta:
                label = f"{label} #{idx + 1}"
            option_meta[label] = {
                "id": rid,
                "index": idx,
                "name": name.upper(),
                "department": dept.upper(),
            }
            delete_options.append(label)
        selected = st.multiselect(
            f"Select {entity_label} profiles",
            options=delete_options,
            key=f"{sheet_name}_delete_select",
        )
        confirm = st.checkbox(
            "Confirm delete",
            key=f"{sheet_name}_delete_confirm",
        )
        if st.button(
            f"Delete selected {entity_label} profiles",
            key=f"{sheet_name}_delete_btn",
            use_container_width=True,
        ):
            if not selected:
                st.warning("Select at least one profile to delete.")
            elif not confirm:
                st.warning("Please confirm delete.")
            else:
                to_delete = [option_meta[label] for label in selected if label in option_meta]
                if USE_SUPABASE and supabase_client is not None:
                    try:
                        ids = [item["id"] for item in to_delete if item.get("id")]
                        if ids:
                            supabase_client.table(PROFILE_SUPABASE_TABLE).delete().in_("id", ids).execute()
                        for item in to_delete:
                            if item.get("id"):
                                continue
                            if not item.get("name"):
                                continue
                            q = (
                                supabase_client.table(PROFILE_SUPABASE_TABLE)
                                .delete()
                                .eq("kind", sheet_name)
                                .eq("name", item["name"])
                            )
                            if item.get("department"):
                                q = q.eq("department", item["department"])
                            q.execute()
                        try:
                            _get_active_assistant_profile_names.clear()
                        except Exception:
                            pass
                        try:
                            _refresh_staff_options_from_supabase(supabase_client)
                        except Exception:
                            pass
                    except Exception as e:
                        st.error(f"Failed to delete {entity_label} profiles: {e}")
                        return
                else:
                    drop_idx = [item["index"] for item in to_delete]
                    df_after = df_profiles.drop(index=drop_idx, errors="ignore").reset_index(drop=True)
                    ok = save_profiles(df_after, sheet_name)
                    if not ok:
                        st.error(f"Failed to delete {entity_label} profiles.")
                        return
                if USE_SUPABASE and supabase_client is not None:
                    st.session_state.supabase_staff_refreshed = False
                st.session_state.profiles_cache_bust += 1
                st.success(f"Deleted {len(to_delete)} {entity_label} profile(s).")
                st.rerun()
    if _dialog_decorator:
        @_dialog_decorator(f"Delete {entity_label}")
        def _render_delete_profile_dialog() -> None:
            _render_delete_profile_dialog_body()
    else:
        def _render_delete_profile_dialog() -> None:
            st.warning("Popup delete requires a newer Streamlit version.")
            _render_delete_profile_dialog_body()
    if st.button(f"Delete {entity_label}", key=f"delete_{sheet_name}_open", use_container_width=True):
        _render_delete_profile_dialog()
    st.markdown("#### Edit All Profiles")
    edited_df = st.data_editor(
        df_profiles,
        hide_index=True,
        use_container_width=True,
        key=f"{sheet_name}_editor",
        column_config={
            "id": None,
            "name": st.column_config.TextColumn(f"{entity_label} Name", required=True),
            "department": st.column_config.SelectboxColumn(dept_label, options=dept_options),
            "contact_email": st.column_config.TextColumn("Contact Email"),
            "contact_phone": st.column_config.TextColumn("Contact Phone"),
            "status": st.column_config.SelectboxColumn("Status", options=status_options, required=True),
            "created_at": None,
            "updated_at": None,
            "created_by": None,
            "updated_by": None,
        },
    )
    if st.button("Save profile changes", key=f"{sheet_name}_save_btn"):
        edited_df = _restore_profile_hidden_columns(edited_df, df_profiles, hidden_cols, user_name)
        edited_df["updated_at"] = _now_iso()
        edited_df["updated_by"] = user_name
        ok = save_profiles(edited_df, sheet_name)
        if not ok:
            st.error("Failed to save profile changes.")
            return
        st.session_state.profiles_cache_bust += 1
        if USE_SUPABASE and supabase_client is not None:
            st.session_state.supabase_staff_refreshed = False
        st.success("Profiles updated.")
        if USE_SUPABASE and supabase_client is not None:
            st.rerun()
# Auto-select backend: Supabase if configured, else local Excel
if not USE_SUPABASE:
    sup_url_hint = _safe_secret_get("supabase_url") or os.environ.get("SUPABASE_URL")
    sup_key_hint = (
        _safe_secret_get("supabase_service_role_key")
        or _safe_secret_get("supabase_key")
        or os.environ.get("SUPABASE_SERVICE_ROLE_KEY")
        or os.environ.get("SUPABASE_KEY")
    )
    # Supabase is disabled - using Excel-only mode
    # if SUPABASE_AVAILABLE and sup_url_hint and sup_key_hint and sup_url_hint.strip() and sup_key_hint.strip():
    #     USE_SUPABASE = True
    #     st.sidebar.success("🔗 Connected to Supabase")
    # else:
    st.sidebar.info("📁 Using local Excel file")
def _get_supabase_config_from_secrets_or_env():
    """Return (url, key, table, row_id, profile_table) from Streamlit secrets/env vars."""
    url = ""
    key = ""
    service_key = ""
    table = supabase_table_name
    row_id = supabase_row_id
    profile_table = PROFILE_SUPABASE_TABLE
    try:
        if hasattr(st, 'secrets'):
            supabase_section = st.secrets.get("supabase", None)
            if isinstance(supabase_section, dict):
                # Support [supabase] table in secrets for legacy configs.
                url = str(supabase_section.get("url", "") or "").strip() or url
                key = str(supabase_section.get("key", "") or "").strip() or key
                service_key = str(supabase_section.get("service_role_key", "") or "").strip() or service_key
                table = str(supabase_section.get("table", table) or table).strip() or table
                row_id = str(supabase_section.get("row_id", row_id) or row_id).strip() or row_id
                profile_table = str(supabase_section.get("profile_table", profile_table) or profile_table).strip() or profile_table
            url = str(st.secrets.get("supabase_url", "") or "").strip() or url
            key = str(st.secrets.get("supabase_key", "") or "").strip() or key
            service_key = str(st.secrets.get("supabase_service_role_key", "") or "").strip() or service_key
            table = str(st.secrets.get("supabase_table", table) or table).strip() or table
            row_id = str(st.secrets.get("supabase_row_id", row_id) or row_id).strip() or row_id
            profile_table = str(st.secrets.get("supabase_profile_table", profile_table) or profile_table).strip() or profile_table
    except Exception:
        pass
    if not url:
        url = os.getenv("SUPABASE_URL", "").strip()
    if not key:
        key = os.getenv("SUPABASE_KEY", "").strip()
    if not service_key:
        service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    if os.getenv("SUPABASE_TABLE"):
        table = os.getenv("SUPABASE_TABLE", table).strip() or table
    if os.getenv("SUPABASE_ROW_ID"):
        row_id = os.getenv("SUPABASE_ROW_ID", row_id).strip() or row_id
    if os.getenv("SUPABASE_PROFILE_TABLE"):
        profile_table = os.getenv("SUPABASE_PROFILE_TABLE", profile_table).strip() or profile_table
    # SECURITY FIX: No hardcoded defaults - require configuration via secrets/env
    if not url or not key:
        # Return None to indicate Supabase is not configured (caller should handle gracefully)
        return None, None, table, row_id, profile_table
    # Prefer service role key when present (avoids RLS setup for server-side app).
    effective_key = service_key or key
    return url, effective_key, table, row_id, profile_table
@st.cache_resource
def _get_supabase_client_cached(_url: str, _key: str):
    return create_client(_url, _key)
def _get_supabase_client(_url: str, _key: str):
    if not SUPABASE_AVAILABLE:
        return None
    if not _url or not _key:
        return None
    try:
        return _get_supabase_client_cached(_url, _key)
    except Exception:
        try:
            return create_client(_url, _key)
        except Exception:
            return None
def _supabase_ready_recent() -> bool:
    """Check if Supabase connectivity was verified recently.
    Performance: Cache connectivity check for 5 minutes to avoid repeated queries
    """
    try:
        if not st.session_state.get("supabase_ready"):
            return False
        last = float(st.session_state.get("supabase_ready_at") or 0.0)
        # Increased from 60s to 300s (5 minutes) for better performance
        return (time_module.time() - last) < 300
    except Exception:
        return False
def _get_expected_columns():
    return [
        "Patient ID", "Patient Name", "In Time", "Out Time", "Procedure", "DR.",
        "FIRST", "SECOND", "Third", "CASE PAPER", "OP",
        "SUCTION", "CLEANING", "STATUS", "REMINDER_ROW_ID",
        "REMINDER_SNOOZE_UNTIL", "REMINDER_DISMISSED",
        # Time tracking / status audit (stored in the same allotment table)
        "STATUS_CHANGED_AT", "ACTUAL_START_AT", "ACTUAL_END_AT", "STATUS_LOG",
    ]
# ================ PATIENT STATUS OPTIONS ================
# Keep legacy values for compatibility with existing data.
STATUS_BASE_OPTIONS = [
    "PENDING",
    "WAITING",
    "ARRIVING",
    "ARRIVED",
    "ON GOING",
    "DONE",
    "COMPLETED",
    "CANCELLED",
    "SHIFTED",
    "LATE",  # patient running late
]
def _now_ist_str() -> str:
    return datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
def _append_status_log(existing_value, event: dict) -> str:
    """Append a status change event to a JSON list stored in a cell."""
    items: list[dict] = []
    try:
        if isinstance(existing_value, list):
            items = [x for x in existing_value if isinstance(x, dict)]
        elif isinstance(existing_value, str) and existing_value.strip():
            parsed = json.loads(existing_value)
            if isinstance(parsed, list):
                items = [x for x in parsed if isinstance(x, dict)]
    except Exception:
        items = []
    items.append(dict(event))
    try:
        return json.dumps(items, ensure_ascii=False)
    except Exception:
        return ""
def _get_patients_config_from_secrets_or_env():
    """Return (patients_table, id_col, name_col)."""
    patients_table = "patients"
    id_col = "id"
    name_col = "name"
    try:
        if hasattr(st, 'secrets'):
            patients_table = str(st.secrets.get("supabase_patients_table", patients_table) or patients_table).strip() or patients_table
            id_col = str(st.secrets.get("supabase_patients_id_col", id_col) or id_col).strip() or id_col
            name_col = str(st.secrets.get("supabase_patients_name_col", name_col) or name_col).strip() or name_col
    except Exception:
        pass
    patients_table = os.getenv("SUPABASE_PATIENTS_TABLE", patients_table).strip() or patients_table
    id_col = os.getenv("SUPABASE_PATIENTS_ID_COL", id_col).strip() or id_col
    name_col = os.getenv("SUPABASE_PATIENTS_NAME_COL", name_col).strip() or name_col
    return patients_table, id_col, name_col
@st.cache_data(ttl=60)
def search_patients_from_supabase(
    _url: str,
    _key: str,
    _patients_table: str,
    _id_col: str,
    _name_col: str,
    _query: str,
    _limit: int = 50,
):
    """Search patients (id + name) from a Supabase table."""
    q = (_query or "").strip()
    client = _get_supabase_client(_url, _key)
    if client is None:
        return []
    def _is_simple_ident(name: str) -> bool:
        return bool(re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", str(name or "")))
    def _quote_ident(name: str) -> str:
        n = str(name or "")
        # Quote if it has spaces, punctuation, or uppercase/lowercase sensitivity.
        if _is_simple_ident(n) and n == n.lower():
            return n
        return '"' + n.replace('"', '""') + '"'
    def _run(_id: str, _name: str, *, server_filter: bool) -> Optional[list[dict]]:
        select_str = f"{_quote_ident(_id)},{_quote_ident(_name)}"
        query = client.table(_patients_table).select(select_str)
        # Only apply server-side ilike/order if the column name is a simple identifier.
        if server_filter and q and _is_simple_ident(_name):
            query = query.ilike(_name, f"%{q}%")
        if server_filter and _is_simple_ident(_name):
            query = query.order(_name)
        resp = query.limit(_limit).execute()
        err = getattr(resp, "error", None)
        if err:
            raise RuntimeError(str(err))
        data = getattr(resp, "data", None)
        return data
    # PostgREST supports ilike and order.
    try:
        data = _run(_id_col, _name_col, server_filter=True)
    except Exception as e:
        # Common case: columns are not named exactly `id`/`name`.
        # Postgres error code for unknown column is 42703.
        err_text = str(e)
        if "42703" not in err_text and "does not exist" not in err_text:
            raise
        # First try to infer actual column names by sampling 1 row.
        inferred_id: Optional[str] = None
        inferred_name: Optional[str] = None
        try:
            probe = client.table(_patients_table).select("*").limit(1).execute()
            probe_err = getattr(probe, "error", None)
            if probe_err:
                raise RuntimeError(str(probe_err))
            probe_data = getattr(probe, "data", None)
            if isinstance(probe_data, list) and probe_data and isinstance(probe_data[0], dict):
                keys = [str(k) for k in probe_data[0].keys()]
                keys_l = {k.lower(): k for k in keys}
                # Heuristics: prefer exact matches, else keys containing patient+id/name.
                for cand in ["id", "patient_id", "patientid", "uhid", "pid", "patient id"]:
                    if cand in keys_l:
                        inferred_id = keys_l[cand]
                        break
                for cand in ["name", "patient_name", "patientname", "full_name", "fullname", "patient name"]:
                    if cand in keys_l:
                        inferred_name = keys_l[cand]
                        break
        except Exception:
            inferred_id = None
            inferred_name = None
        if inferred_id and inferred_name:
            data = _run(inferred_id, inferred_name, server_filter=_is_simple_ident(inferred_name))
            _id_col, _name_col = inferred_id, inferred_name
        else:
            # Fall back to trying a broader set of common column names.
            id_candidates = [
                _id_col,
                "id",
                "ID",
                "patient_id",
                "patientId",
                "patientid",
                "uhid",
                "UHID",
                "pid",
                "PID",
                "patient id",
                "Patient ID",
            ]
            name_candidates = [
                _name_col,
                "name",
                "NAME",
                "patient_name",
                "patientName",
                "patientname",
                "full_name",
                "fullName",
                "fullname",
                "patient name",
                "Patient Name",
            ]
            last_err: Optional[Exception] = None
            data = None
            for cid in id_candidates:
                for cname in name_candidates:
                    if not cid or not cname:
                        continue
                    try:
                        data = _run(cid, cname, server_filter=_is_simple_ident(cname))
                        _id_col = cid
                        _name_col = cname
                        last_err = None
                        break
                    except Exception as inner:
                        last_err = inner
                        continue
                if last_err is None and data is not None:
                    break
            if data is None:
                raise last_err if last_err is not None else e
    if not isinstance(data, list):
        return []
    out = []
    for row in data:
        pid = row.get(_id_col)
        name = row.get(_name_col)
        if pid is None or name is None:
            continue
        out.append({"id": str(pid), "name": str(name)})
    # If we couldn't do server-side filtering (e.g., quoted column names), filter locally.
    if q and out:
        ql = q.lower()
        out = [p for p in out if ql in str(p.get("name", "")).lower()]
    return out
@st.cache_data(ttl=300, show_spinner="Loading data from Supabase...")
def load_data_from_supabase(_url: str, _key: str, _table: str, _row_id: str):
    """Load dataframe payload from Supabase.
    Storage model: a single row with `id` and `payload` (jsonb).
    payload = {"columns": [...], "rows": [ {col: val, ...}, ... ]}
    Performance: Cached for 5 minutes (300s) to minimize API calls
    """
    try:
        client = _get_supabase_client(_url, _key)
        if client is None:
            return None
        # Optimized query: only fetch payload field
        resp = client.table(_table).select("payload").eq("id", _row_id).limit(1).execute()
        data = getattr(resp, "data", None)
        if not data:
            return pd.DataFrame(columns=_get_expected_columns())
        payload = data[0].get("payload") if isinstance(data, list) else None
        if not payload:
            return pd.DataFrame(columns=_get_expected_columns())
        columns = payload.get("columns") or _get_expected_columns()
        # Ensure new expected columns are added for older saved payloads.
        try:
            expected = _get_expected_columns()
            for col in expected:
                if col not in columns:
                    columns.append(col)
        except Exception:
            pass
        rows = payload.get("rows") or []
        df = pd.DataFrame(rows)
        # Ensure expected columns are present and ordered
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        df = df[columns]
        # Optional metadata (e.g., assistant time blocks)
        try:
            meta = payload.get("meta")
            if isinstance(meta, dict):
                df.attrs["meta"] = dict(meta)
        except Exception:
            pass
        return df
    except Exception as e:
        st.error(f"Error loading from Supabase: {e}")
        return None
def save_data_to_supabase(_url: str, _key: str, _table: str, _row_id: str, df: pd.DataFrame) -> bool:
    """Save dataframe payload to Supabase (upsert)."""
    try:
        client = _get_supabase_client(_url, _key)
        if client is None:
            return False
        df_clean = df.copy().fillna("")
        # Convert to JSON-serializable primitives; avoid pandas NA
        for col in df_clean.columns:
            df_clean[col] = df_clean[col].astype(object)
        payload = {
            "columns": df_clean.columns.tolist(),
            "rows": df_clean.to_dict(orient="records"),
        }
        # Optional metadata (stored alongside rows/columns)
        try:
            meta = _get_meta_from_df(df)
            meta = _apply_time_blocks_to_meta(meta)
            payload["meta"] = meta
        except Exception:
            pass
        client.table(_table).upsert({"id": _row_id, "payload": payload}).execute()
        # PERFORMANCE: Don't clear cache here - let TTL handle it
        # Cache will auto-refresh after 5 minutes, preventing excessive API calls
        # Only clear session cache to force reload on next access
        if "cached_df_timestamp" in st.session_state:
            st.session_state.cached_df_timestamp = 0  # Force reload from Streamlit cache
        return True
    except Exception as e:
        st.error(f"Error saving to Supabase: {e}")
        return False
# Try to connect to Supabase using credentials from secrets.toml / env vars
# PERFORMANCE: Only run full initialization once per session
try:
    _sb_url, _sb_key, _sb_table, _sb_row_id, _sb_profile_table = (
        _get_supabase_config_from_secrets_or_env()
    )
    if _sb_url and _sb_key and SUPABASE_AVAILABLE:
        _maybe_client = _get_supabase_client(_sb_url, _sb_key)
        if _maybe_client is not None:
            supabase_client = _maybe_client
            supabase_table_name = _sb_table
            supabase_row_id = _sb_row_id
            st.sidebar.info("☁️ Supabase connected: cloud data sync enabled")
except Exception:
    pass
def _data_editor_has_pending_edits(editor_key: str) -> bool:
    """Detect pending edits without touching widget state.
    Streamlit stores data_editor widget edits in st.session_state[editor_key]
    as a dict with keys like edited_rows/added_rows/deleted_rows.
    """
    try:
        state = st.session_state.get(editor_key)
        if not isinstance(state, dict):
            return False
        return bool(state.get("edited_rows") or state.get("added_rows") or state.get("deleted_rows"))
    except Exception:
        return False
def _get_meta_save_version(meta: Optional[dict]) -> Optional[int]:
    if not isinstance(meta, dict):
        return None
    try:
        val = meta.get("save_version")
        if val is None or str(val).strip() == "":
            return None
        return int(float(val))
    except Exception:
        return None
def _meta_for_hash(meta: Optional[dict]) -> dict:
    if not isinstance(meta, dict):
        return {}
    skip = {"time_blocks_updated_at", "saved_at", "save_version"}
    return {k: v for k, v in meta.items() if k not in skip}
def _compute_save_hash(df_any: pd.DataFrame, meta: Optional[dict]) -> str:
    try:
        data_hash = hashlib.md5(pd.util.hash_pandas_object(df_any, index=True).values.tobytes()).hexdigest()
    except Exception:
        data_hash = hashlib.md5(str(df_any).encode("utf-8")).hexdigest()
    try:
        meta_hash = hashlib.md5(
            json.dumps(_meta_for_hash(meta), sort_keys=True, default=str).encode("utf-8")
        ).hexdigest()
    except Exception:
        meta_hash = ""
    return hashlib.md5(f"{data_hash}|{meta_hash}".encode("utf-8")).hexdigest()
def _fetch_remote_save_version() -> Optional[int]:
    try:
        if USE_SUPABASE:
            sup_url, sup_key, sup_table, sup_row, _ = _get_supabase_config_from_secrets_or_env()
            client = _get_supabase_client(sup_url, sup_key)
            if client is None:
                return None
            resp = client.table(sup_table).select("payload").eq("id", sup_row).limit(1).execute()
            data = getattr(resp, "data", None)
            if not data:
                return None
            payload = data[0].get("payload") if isinstance(data, list) else None
            meta = payload.get("meta") if isinstance(payload, dict) else None
            return _get_meta_save_version(meta)
    except Exception:
        return None
    return None
def _get_editor_changed_rows(editor_key: str) -> tuple[list[int], bool]:
    try:
        state = st.session_state.get(editor_key)
        if not isinstance(state, dict):
            return [], False
        if state.get("added_rows"):
            return [], True
        edited = state.get("edited_rows") or {}
        return sorted(int(k) for k in edited.keys()), False
    except Exception:
        return [], False
def _norm_cell(val) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and pd.isna(val):
        return ""
    s = str(val).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s
def _row_has_changes(edited_row, base_row, compare_cols: list[str]) -> bool:
    for col in compare_cols:
        if col not in edited_row.index or col not in base_row.index:
            continue
        if _norm_cell(edited_row.get(col)) != _norm_cell(base_row.get(col)):
            return True
    return False
# ================ Load Data ================
# PERFORMANCE: Use session-based caching to reduce API calls across reruns
def _get_cached_data():
    """Get data with session-level caching for maximum performance."""
    global USE_SUPABASE  # Need to modify this global
    # Check if we have valid cached data in session
    if "cached_df_raw" in st.session_state and "cached_df_timestamp" in st.session_state:
        cached_time = st.session_state.get("cached_df_timestamp", 0)
        current_time = time_module.time()
        # Cache is valid for 2 minutes in session state (in addition to Streamlit cache)
        if current_time - cached_time < 120:
            return st.session_state.cached_df_raw
    # Load fresh data
    df_raw = None
    if USE_SUPABASE:
        sup_url, sup_key, sup_table, sup_row, _ = _get_supabase_config_from_secrets_or_env()
        # Check if Supabase is actually configured
        if sup_url and sup_key:
            df_raw = load_data_from_supabase(sup_url, sup_key, sup_table, sup_row)
            if df_raw is None:
                st.warning("⚠️ Failed to load from Supabase. Falling back to local Excel file.")
                USE_SUPABASE = False  # Disable for this session
        else:
            st.warning("⚠️ Supabase not configured. Falling back to local Excel file.")
            USE_SUPABASE = False  # Disable for this session
    # Fallback to local Excel if cloud storage failed or not configured
    if df_raw is None:
        st.info("📁 Using local Excel file: Putt Allotment.xlsx")
        try:
            if os.path.exists(file_path):
                df_raw = pd.read_excel(file_path, engine="openpyxl")
            else:
                # Create new Excel file with expected columns
                df_raw = pd.DataFrame(columns=_get_expected_columns())
                df_raw.to_excel(file_path, index=False, engine="openpyxl")
                st.success(f"✅ Created new Excel file: {file_path}")
        except Exception as e:
            st.error(f"❌ Failed to load/create Excel file: {e}")
            # Create empty dataframe as last resort
            df_raw = pd.DataFrame(columns=_get_expected_columns())
    # Store in session cache
    if df_raw is not None:
        st.session_state.cached_df_raw = df_raw
        st.session_state.cached_df_timestamp = time_module.time()
    return df_raw
# Use cached data loader
df_raw = _get_cached_data()
# Track base save version/hash from storage unless we have local pending edits.
loaded_meta = _get_meta_from_df(df_raw)
if st.session_state.get("unsaved_df") is None:
    loaded_version = _get_meta_save_version(loaded_meta)
    if loaded_version is not None:
        st.session_state.loaded_save_version = loaded_version
        st.session_state.loaded_save_at = loaded_meta.get("saved_at")
        st.session_state.last_saved_hash = _compute_save_hash(df_raw, loaded_meta)
    elif st.session_state.get("last_saved_hash") is None:
        st.session_state.last_saved_hash = _compute_save_hash(df_raw, loaded_meta)
# Prefer in-session pending changes when auto-save is off
if st.session_state.get("unsaved_df") is not None:
    try:
        df_raw = st.session_state.unsaved_df.copy()
    except Exception:
        df_raw = st.session_state.unsaved_df
# Clean column names
df_raw.columns = [col.strip() for col in df_raw.columns]
# Ensure metadata attribute exists (defensive check)
# Ensure metadata attribute exists (defensive check)
if not hasattr(df_raw, 'attrs'):
    df_raw.attrs = {}
if "meta" not in df_raw.attrs:
    df_raw.attrs["meta"] = {}
# Load persisted time blocks (if present) from storage metadata
_sync_time_blocks_from_meta(df_raw)
# --- AUTO-REPAIR TIME BLOCKS FORMAT ---
def _is_time_block_valid(block):
    # Check for required keys and correct types
    try:
        if not isinstance(block, dict):
            return False
        if not all(k in block for k in ("assistant", "date", "reason", "start_time", "end_time")):
            return False
        # start_time/end_time should be time or string 'HH:MM'
        st_val = block["start_time"]
        et_val = block["end_time"]
        def _is_time(val):
            from datetime import time
            return isinstance(val, time) or (isinstance(val, str) and len(val) == 5 and val[2] == ":")
        return _is_time(st_val) and _is_time(et_val)
    except Exception:
        return False
meta = df_raw.attrs.get("meta", {})
blocks = meta.get("time_blocks", [])
if not isinstance(blocks, list) or not all(_is_time_block_valid(b) for b in blocks):
    # Attempt to repair by re-serializing current session_state.time_blocks
    import streamlit as st
    try:
        meta = _apply_time_blocks_to_meta(meta)
        df_raw.attrs["meta"] = meta
        save_data(df_raw, show_toast=False, message="Auto-repaired time_blocks format")
        _sync_time_blocks_from_meta(df_raw)
    except Exception as e:
        st.warning(f"[Auto-repair] Failed to repair time_blocks format: {e}")
# Ensure expected columns exist (backfills older data/backends)
for _col in _get_expected_columns():
    if _col in df_raw.columns:
        continue
    if _col == "REMINDER_SNOOZE_UNTIL":
        df_raw[_col] = pd.NA
    elif _col == "REMINDER_DISMISSED":
        df_raw[_col] = False
    else:
        df_raw[_col] = ""
def _collect_unique_upper(df_any: pd.DataFrame, col_name: str) -> list[str]:
    try:
        if col_name not in df_any.columns:
            return []
        s = df_any[col_name].astype(str).replace("nan", "").fillna("")
        vals = [str(v).strip().upper() for v in s.tolist() if str(v).strip()]
        return _unique_preserve_order(vals)
    except Exception:
        return []
# Dropdown options: keep configured lists + include any existing values from data
_extra_doctors = _collect_unique_upper(df_raw, "DR.")
DOCTOR_OPTIONS = _unique_preserve_order(_get_all_doctors() + _extra_doctors)
_extra_assistants: list[str] = []
for _c in ["FIRST", "SECOND", "Third", "CASE PAPER"]:
    _extra_assistants.extend(_collect_unique_upper(df_raw, _c))
ASSISTANT_OPTIONS = _unique_preserve_order(_get_all_assistants() + _extra_assistants)
# Status options: configured set + any existing values in data
_extra_statuses = _collect_unique_upper(df_raw, "STATUS")
STATUS_OPTIONS = _unique_preserve_order(STATUS_BASE_OPTIONS + _extra_statuses)
# Convert checkbox columns (SUCTION, CLEANING) - checkmark or content to boolean
def str_to_checkbox(val: Any) -> bool:
    """Convert string values to boolean for checkboxes"""
    # Preserve actual booleans
    if isinstance(val, bool):
        return val
    # Handle numbers (0/1)
    try:
        if isinstance(val, (int, float)) and not pd.isna(val):
            return bool(int(val))
    except Exception:
        pass
    if pd.isna(val):
        return False
    s = str(val).strip()
    if s == "":
        return False
    su = s.upper()
    if su in {"FALSE", "F", "0", "NO", "N", "NONE", "NAN"}:
        return False
    if su in {"TRUE", "T", "1", "YES", "Y"}:
        return True
    if s == "✓":
        return True
    # Any other non-empty content is treated as checked (legacy behavior)
    return True
def _schedule_cache_key() -> tuple:
    if st.session_state.get("unsaved_df") is not None:
        return ("unsaved", st.session_state.get("unsaved_df_version", 0))
    return (
        "saved",
        st.session_state.get("loaded_save_version"),
        st.session_state.get("last_saved_hash"),
    )
def _schedule_change_key() -> tuple:
    return _schedule_cache_key()
def _get_cached_schedule_hash(df_any: pd.DataFrame) -> str:
    cache_key = _schedule_change_key()
    cached_key = st.session_state.get("schedule_hash_key")
    cached_hash = st.session_state.get("schedule_hash")
    if cached_hash and cached_key == cache_key:
        return cached_hash
    meta = None
    try:
        meta = df_any.attrs.get("meta")
    except Exception:
        meta = None
    new_hash = _compute_save_hash(df_any, meta)
    st.session_state.schedule_hash_key = cache_key
    st.session_state.schedule_hash = new_hash
    return new_hash
def _notification_tick_key(schedule_hash: str) -> tuple:
    return (schedule_hash, int(time_module.time() // 60))
def _prepare_schedule_df_static(df_any: pd.DataFrame) -> pd.DataFrame:
    df_local = df_any.copy()
    df_local["In Time Str"] = df_local["In Time"].apply(dec_to_time)
    df_local["Out Time Str"] = df_local["Out Time"].apply(dec_to_time)
    df_local["In Time Obj"] = df_local["In Time Str"].apply(safe_str_to_time_obj)
    df_local["Out Time Obj"] = df_local["Out Time Str"].apply(safe_str_to_time_obj)
    if "SUCTION" in df_local.columns:
        df_local["SUCTION"] = df_local["SUCTION"].apply(str_to_checkbox)
    if "CLEANING" in df_local.columns:
        df_local["CLEANING"] = df_local["CLEANING"].apply(str_to_checkbox)
    df_local["In_min"] = df_local["In Time"].apply(time_to_minutes).astype("Int64")
    df_local["Out_min"] = df_local["Out Time"].apply(time_to_minutes).astype("Int64")
    df_local.loc[df_local["Out_min"] < df_local["In_min"], "Out_min"] += 1440
    return df_local
def _get_processed_schedule_df(df_any: pd.DataFrame) -> pd.DataFrame:
    cache_key = _schedule_cache_key()
    cached_key = st.session_state.get("schedule_df_cache_key")
    cached_df = st.session_state.get("schedule_df_cache")
    if cached_df is not None and cached_key == cache_key:
        try:
            return cached_df.copy(deep=False)
        except Exception:
            return cached_df
    df_local = _prepare_schedule_df_static(df_any)
    st.session_state.schedule_df_cache_key = cache_key
    st.session_state.schedule_df_cache = df_local
    return df_local
# ================ Reminder Persistence Setup ================
# Add stable row IDs and reminder columns if they don't exist
if 'Patient ID' not in df_raw.columns:
    df_raw['Patient ID'] = ""
if 'REMINDER_ROW_ID' not in df_raw.columns:
    df_raw['REMINDER_ROW_ID'] = [str(uuid.uuid4()) for _ in range(len(df_raw))]
    # Save IDs immediately - will use save_data after it's defined
    _needs_id_save = True
else:
    # Backfill missing/blank IDs so every row (including blank rows) can be targeted for delete/reminders.
    _needs_id_save = False
    try:
        rid_series = df_raw['REMINDER_ROW_ID'].astype(str)
        missing_mask = df_raw['REMINDER_ROW_ID'].isna() | rid_series.str.strip().eq("") | rid_series.str.lower().eq("nan")
        if bool(missing_mask.any()):
            df_raw.loc[missing_mask, 'REMINDER_ROW_ID'] = [str(uuid.uuid4()) for _ in range(int(missing_mask.sum()))]
            _needs_id_save = True
    except Exception:
        # If anything goes wrong, keep dashboard usable; IDs will be handled elsewhere.
        pass
if 'REMINDER_SNOOZE_UNTIL' not in df_raw.columns:
    df_raw['REMINDER_SNOOZE_UNTIL'] = pd.NA
if 'REMINDER_DISMISSED' not in df_raw.columns:
    df_raw['REMINDER_DISMISSED'] = False
# Refresh df with new columns
df = _get_processed_schedule_df(df_raw)
# Current time in minutes (same day)
current_min = now.hour * 60 + now.minute
# Mark ongoing
df["Is_Ongoing"] = (df["In_min"] <= current_min) & (current_min <= df["Out_min"])
# ================ Unified Save Function ================
def save_data(dataframe, show_toast=True, message="Data saved!", *, ignore_conflict=False):
    """Save dataframe to Supabase or Excel based on configuration."""
    if st.session_state.get("is_saving"):
        return False
    st.session_state.is_saving = True
    try:
        if not hasattr(dataframe, 'attrs'):
            dataframe.attrs = {}
        meta = _get_meta_from_df(dataframe)
        meta = _apply_time_blocks_to_meta(meta)
        loaded_version = st.session_state.get("loaded_save_version")
        local_version = _get_meta_save_version(meta)
        if local_version is None and loaded_version is not None:
            local_version = _safe_int(loaded_version, 0)
        remote_version = None
        if (
            st.session_state.get("enable_conflict_checks", True)
            and not ignore_conflict
            and USE_SUPABASE
        ):
            remote_version = _fetch_remote_save_version()
            if remote_version is not None and loaded_version is not None:
                if _safe_int(remote_version, -1) != _safe_int(loaded_version, -1):
                    st.session_state.save_conflict = {
                        "local_version": loaded_version,
                        "remote_version": remote_version,
                        "detected_at": now_ist().isoformat(),
                    }
                    st.error("Save blocked: newer data detected in storage.")
                    return False
        save_hash = _compute_save_hash(dataframe, meta)
        if save_hash == st.session_state.get("last_saved_hash"):
            return True
        base_version = max(
            _safe_int(loaded_version, 0),
            _safe_int(remote_version, 0),
            _safe_int(local_version, 0),
        )
        meta["save_version"] = int(base_version) + 1
        meta["saved_at"] = now_ist().isoformat()
        dataframe.attrs["meta"] = meta
        if USE_SUPABASE:
            sup_url, sup_key, sup_table, sup_row, _ = _get_supabase_config_from_secrets_or_env()
            if sup_url and sup_key:
                success = save_data_to_supabase(sup_url, sup_key, sup_table, sup_row, dataframe)
                if success and show_toast:
                    st.toast(message, icon="✅")
            else:
                st.warning("⚠️ Supabase not configured. Saving to local Excel instead.")
                # Use our safe sheet saving function to preserve other sheets
                save_excel_sheet(dataframe, 'Sheet1')
                if show_toast:
                    st.toast(f"{message} (local)", icon="💾")
        else:
            # Use our safe sheet saving function to preserve other sheets
            save_excel_sheet(dataframe, 'Sheet1')
            try:
                meta = _apply_time_blocks_to_meta(_get_meta_from_df(dataframe))
                meta_rows = []
                for k, v in meta.items():
                    if isinstance(v, (dict, list)):
                        meta_rows.append({"key": str(k), "value": json.dumps(v)})
                    else:
                        meta_rows.append({"key": str(k), "value": str(v)})
                save_excel_sheet(pd.DataFrame(meta_rows), 'Meta')
            except Exception:
                pass
            success = True
            if show_toast:
                st.toast(message, icon="✅")
        if success:
            st.session_state.last_saved_hash = save_hash
            st.session_state.loaded_save_version = meta.get("save_version")
            st.session_state.loaded_save_at = meta.get("saved_at")
            st.session_state.save_conflict = None
            st.session_state.last_save_at = time_module.time()
        return success
    except Exception as e:
        st.error(f"Error saving data: {e}")
        return False
    finally:
        st.session_state.is_saving = False
def _queue_unsaved_df(df_pending: pd.DataFrame, reason: str = "") -> None:
    """Keep changes in memory when auto-save is disabled or delayed."""
    try:
        st.session_state.unsaved_df = df_pending.copy(deep=False)
    except Exception:
        st.session_state.unsaved_df = df_pending
    try:
        st.session_state.unsaved_df_version = int(st.session_state.get("unsaved_df_version", 0)) + 1
    except Exception:
        st.session_state.unsaved_df_version = 1
    st.session_state.pending_changes = True
    st.session_state.pending_changes_reason = reason
def _maybe_save(dataframe, show_toast=True, message="Data saved!", force=False, ignore_conflict=False):
    """Respect auto-save toggle; queue changes if disabled or debounced."""
    if st.session_state.get("is_saving"):
        _queue_unsaved_df(dataframe, reason=message)
        return True
    if force:
        result = save_data(dataframe, show_toast=show_toast, message=message, ignore_conflict=ignore_conflict)
        if result:
            st.session_state.unsaved_df = None
            st.session_state.pending_changes = False
            st.session_state.pending_changes_reason = ""
        else:
            _queue_unsaved_df(dataframe, reason=message)
        return result
    if st.session_state.get("auto_save_enabled", False):
        debounce_s = st.session_state.get("save_debounce_seconds", 0)
        try:
            debounce_s = float(debounce_s or 0)
        except Exception:
            debounce_s = 0.0
        if debounce_s > 0:
            now_ts = time_module.time()
            last_at = float(st.session_state.get("last_save_at", 0.0) or 0.0)
            if (now_ts - last_at) < debounce_s:
                _queue_unsaved_df(dataframe, reason=message)
                return True
        result = save_data(dataframe, show_toast=show_toast, message=message, ignore_conflict=ignore_conflict)
        if result:
            st.session_state.unsaved_df = None
            st.session_state.pending_changes = False
            st.session_state.pending_changes_reason = ""
        else:
            _queue_unsaved_df(dataframe, reason=message)
        return result
    _queue_unsaved_df(dataframe, reason=message)
    if show_toast:
        st.toast("Auto-save disabled. Click 'Save Changes' to persist.", icon="⚠")
    return True
def _build_schedule_backups(df_any: pd.DataFrame) -> tuple[bytes, bytes]:
    """Return (csv_bytes, xlsx_bytes) for the current schedule."""
    csv_bytes = df_any.to_csv(index=False).encode("utf-8")
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_any.to_excel(writer, sheet_name="Sheet1", index=False)
        # Include metadata (time blocks) if present
        try:
            meta = _apply_time_blocks_to_meta(_get_meta_from_df(df_any))
            meta_rows = []
            for k, v in meta.items():
                if isinstance(v, (dict, list)):
                    meta_rows.append({"key": str(k), "value": json.dumps(v)})
                else:
                    meta_rows.append({"key": str(k), "value": str(v)})
            pd.DataFrame(meta_rows).to_excel(writer, sheet_name="Meta", index=False)
        except Exception:
            pass
    xlsx_bytes = buf.getvalue()
    return csv_bytes, xlsx_bytes
def _get_cached_schedule_backups(df_any: pd.DataFrame) -> tuple[bytes, bytes]:
    cache_key = _schedule_cache_key()
    cached_key = st.session_state.get("schedule_backup_key")
    cached_bytes = st.session_state.get("schedule_backup_cache")
    if cached_bytes is not None and cached_key == cache_key:
        return cached_bytes
    csv_bytes, xlsx_bytes = _build_schedule_backups(df_any)
    st.session_state.schedule_backup_key = cache_key
    st.session_state.schedule_backup_cache = (csv_bytes, xlsx_bytes)
    return csv_bytes, xlsx_bytes
def _make_cleared_schedule(df_existing: pd.DataFrame) -> pd.DataFrame:
    """Create an empty schedule dataframe while preserving metadata (e.g., time blocks)."""
    cols = list(df_existing.columns)
    df_empty = pd.DataFrame(columns=cols)
    try:
        meta = _apply_time_blocks_to_meta(_get_meta_from_df(df_existing))
        _set_meta_on_df(df_empty, meta)
    except Exception:
        pass
    return df_empty
# ================ TIME BLOCKING UI (persisted) ================
with st.sidebar:
    st.markdown("## 💾 Save Mode")
    st.session_state.auto_save_enabled = st.checkbox(
        "Enable auto-save",
        value=st.session_state.get("auto_save_enabled", False),
        help="When off, changes stay in session until you click 'Save Changes'."
    )
    save_now_disabled = bool(st.session_state.get("is_saving")) or bool(st.session_state.get("save_conflict"))
    if st.button("Save Now", key="save_now_btn", use_container_width=True, disabled=save_now_disabled):
        df_to_save = st.session_state.get("unsaved_df")
        if df_to_save is None:
            df_to_save = df_raw if "df_raw" in locals() else None
        if df_to_save is not None:
            _maybe_save(df_to_save, message="Saved", force=True)
        else:
            st.warning("Nothing to save yet.")
    debounce_options = [0, 1, 2, 3, 5, 10]
    try:
        debounce_index = debounce_options.index(int(st.session_state.get("save_debounce_seconds", 2)))
    except Exception:
        debounce_index = 2
    st.session_state.save_debounce_seconds = st.selectbox(
        "Auto-save debounce (seconds)",
        options=debounce_options,
        index=debounce_index,
        help="Delay auto-save slightly to merge quick edits.",
    )
    st.session_state.enable_conflict_checks = st.checkbox(
        "Block saves on external changes",
        value=st.session_state.get("enable_conflict_checks", True),
        help="Prevents overwriting if storage changed since you loaded.",
    )
    if st.session_state.get("loaded_save_at"):
        st.caption(f"Last saved: {st.session_state.loaded_save_at}")
    if st.session_state.get("is_saving"):
        st.caption("Saving...")
    if st.session_state.get("save_conflict"):
        st.error("Save conflict: storage changed since you loaded.")
        col_conflict_a, col_conflict_b = st.columns(2)
        with col_conflict_a:
            if st.button("Reload from storage", key="reload_storage_btn"):
                st.session_state.unsaved_df = None
                st.session_state.pending_changes = False
                st.session_state.pending_changes_reason = ""
                st.session_state.save_conflict = None
                try:
                    load_data_from_supabase.clear()
                except Exception:
                    pass
                st.rerun()
        with col_conflict_b:
            if st.button("Force Save", key="force_save_btn"):
                df_to_save = st.session_state.get("unsaved_df")
                if df_to_save is None:
                    df_to_save = df_raw
                _maybe_save(
                    df_to_save,
                    message="Force saved (conflict override)",
                    force=True,
                    ignore_conflict=True,
                )
                st.session_state.save_conflict = None
                st.rerun()
    if st.session_state.get("pending_changes"):
        st.caption("Pending changes not yet saved. Click 'Save Changes'.")
        if (
            st.session_state.auto_save_enabled
            and st.session_state.get("unsaved_df") is not None
            and not st.session_state.get("save_conflict")
        ):
            _maybe_save(
                st.session_state.unsaved_df,
                show_toast=False,
                message=st.session_state.get("pending_changes_reason") or "Auto-saved pending changes",
            )
    st.markdown("---")
    st.markdown("## ⏰ Time Blocking")
    st.caption("Block assistants for backend work")
    with st.expander("➕ Add Time Block", expanded=False):
        block_assistant = st.selectbox(
            "Assistant",
            options=[""] + _get_all_assistants(),
            key="block_assistant_select",
        )
        # For debug/demo: auto-fill start and end time to cover current time
        now_dt = now_ist()
        block_start_default = (now_dt - timedelta(minutes=2)).time().replace(second=0, microsecond=0)
        block_end_default = (now_dt + timedelta(minutes=2)).time().replace(second=0, microsecond=0)
        col_start, col_end = st.columns(2)
        with col_start:
            block_start = st.time_input("Start Time", value=block_start_default, key="block_start_time")
        with col_end:
            block_end = st.time_input("End Time", value=block_end_default, key="block_end_time")
        block_reason = st.text_input(
            "Reason",
            value="Backend Work",
            key="block_reason_input",
            placeholder="e.g., Lunch, Training, Backend Work",
        )
        if st.button("🔒 Add Block", key="add_block_btn", use_container_width=True):
            if not block_assistant:
                st.warning("Please select an assistant")
            else:
                add_time_block(block_assistant, block_start, block_end, block_reason)
                _maybe_save(df_raw, show_toast=False, message="Time block saved")
                st.success(
                    f"✅ Blocked {block_assistant} from {block_start.strftime('%I:%M %p')} to {block_end.strftime('%I:%M %p')}"
                )
                st.rerun()
    # Show current time blocks
    if st.session_state.get("time_blocks"):
        st.markdown("**Current Blocks:**")
        today_str = now.strftime("%Y-%m-%d")
        today_blocks = [b for b in st.session_state.time_blocks if b.get("date") == today_str]
        for i, block in enumerate(today_blocks):
            col_info, col_del = st.columns([4, 1])
            with col_info:
                st.caption(
                    f"🚫 {block['assistant']}: {block['start_time'].strftime('%I:%M %p')}-{block['end_time'].strftime('%I:%M %p')} ({block.get('reason','')})"
                )
            with col_del:
                if st.button("❌", key=f"del_block_{i}", help="Remove this block"):
                    try:
                        actual_idx = st.session_state.time_blocks.index(block)
                        remove_time_block(actual_idx)
                        _maybe_save(df_raw, show_toast=False, message="Time block removed")
                        st.success("Time block removed.")
                        st.rerun()
                    except Exception:
                        pass
        # Debug: Show raw time_blocks and meta
        st.markdown("---")
        st.markdown("**[DEBUG] Time Blocks (formatted):**")
        def _format_block(block):
            return {
                'Assistant': block.get('assistant', ''),
                'Start': block.get('start_time').strftime('%I:%M %p') if block.get('start_time') else '',
                'End': block.get('end_time').strftime('%I:%M %p') if block.get('end_time') else '',
                'Date': block.get('date', ''),
                'Reason': block.get('reason', '')
            }
        # Developer debug removed from sidebar per request
    else:
        st.caption("No time blocks set for today")
# ================ RESET / CLEAR ALL ALLOTMENTS ================
with st.sidebar:
    st.markdown("---")
    st.markdown("## 🧹 Reset Schedule")
    st.caption("Clear all current patient appointments/allotments (keeps time blocks).")
    backup_name_base = f"tdb_allotment_backup_{now.strftime('%Y%m%d_%H%M')}"
    try:
        csv_bytes, xlsx_bytes = _get_cached_schedule_backups(df_raw)
        st.download_button(
            "⬇️ Download backup (CSV)",
            data=csv_bytes,
            file_name=f"{backup_name_base}.csv",
            mime="text/csv",
            use_container_width=True,
        )
        st.download_button(
            "⬇️ Download backup (Excel)",
            data=xlsx_bytes,
            file_name=f"{backup_name_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception:
        st.caption("Backup download unavailable.")
    if "confirm_clear_all_check" not in st.session_state:
        st.session_state.confirm_clear_all_check = False
    if "confirm_clear_all_text" not in st.session_state:
        st.session_state.confirm_clear_all_text = ""
    st.checkbox(
        "I understand this will delete ALL rows",
        key="confirm_clear_all_check",
    )
    st.text_input(
        "Type CLEAR to confirm",
        key="confirm_clear_all_text",
        placeholder="CLEAR",
    )
    if st.button(
        "🧹 Clear All Allotments",
        key="clear_all_allotments_btn",
        use_container_width=True,
        help="Permanently clears all current schedule rows",
    ):
        ok_check = bool(st.session_state.get("confirm_clear_all_check"))
        ok_text = str(st.session_state.get("confirm_clear_all_text", "") or "").strip().upper() == "CLEAR"
        if not (ok_check and ok_text):
            st.warning("Please check the box and type CLEAR to confirm.")
        else:
            try:
                df_cleared = _make_cleared_schedule(df_raw)
                success = _maybe_save(df_cleared, message="Schedule cleared")
                if success:
                    # Clear local notification/reminder state so we don't toast old rows.
                    st.session_state.prev_hash = None
                    st.session_state.prev_ongoing = set()
                    st.session_state.prev_upcoming = set()
                    st.session_state.prev_raw = pd.DataFrame()
                    st.session_state.reminder_sent = set()
                    st.session_state.snoozed = {}
                    st.session_state.reminder_state_key = None
                    st.session_state.notification_tick_key = None
                    st.session_state.delete_row_id = ""
                    st.toast("🧹 Schedule cleared", icon="✅")
                    st.rerun()
            except Exception as e:
                st.error(f"Error clearing schedule: {e}")
# Helper to persist reminder state
def _persist_reminder_to_storage(row_id, until, dismissed):
    """Persist snooze/dismiss fields back to storage by row ID."""
    try:
        if df_raw is None or not isinstance(df_raw, pd.DataFrame):
            st.error("Schedule not loaded; cannot persist reminder.")
            return False
        if 'REMINDER_ROW_ID' not in df_raw.columns:
            st.error("Reminder column missing; cannot persist reminder.")
            return False
        match = df_raw[df_raw['REMINDER_ROW_ID'] == row_id]
        if match.empty:
            return False
        ix = match.index[0]
        df_raw.at[ix, 'REMINDER_SNOOZE_UNTIL'] = int(until) if until is not None else pd.NA
        df_raw.at[ix, 'REMINDER_DISMISSED'] = bool(dismissed)
        if st.session_state.get("auto_save_enabled", False):
            return _maybe_save(df_raw, show_toast=False, message="Reminder updates pending")
        _queue_unsaved_df(df_raw, reason="Reminder updates pending")
        return True
    except Exception as e:
        st.error(f"Error persisting reminder: {e}")
    return False
# Save reminder IDs if they were just generated
if _needs_id_save:
    _maybe_save(df_raw, message="Generated stable row IDs for reminders")
# ================ Change Detection & Notifications ================
if 'prev_hash' not in st.session_state:
    st.session_state.prev_hash = None
    st.session_state.prev_ongoing = set()
    st.session_state.prev_upcoming = set()
    st.session_state.prev_raw = pd.DataFrame()
    st.session_state.reminder_sent = set()  # Track reminders by row ID
    st.session_state.snoozed = {}  # Map row_id -> snooze_until_epoch_seconds
active_category = st.session_state.get("nav_category", "Scheduling")
run_alerts = st.session_state.get("alerts_background", False) or active_category == "Scheduling"
if run_alerts:
    enable_reminders = st.session_state.get("enable_reminders", True)
    schedule_key = _schedule_change_key()
    current_hash = _get_cached_schedule_hash(df_raw)
    if st.session_state.prev_hash != current_hash:
        st.toast("📊 ALLOTMENT UPDATED", icon="🔄")
        # Reset tracked sets on file change
        st.session_state.prev_ongoing = set()
        st.session_state.prev_upcoming = set()
        st.session_state.reminder_sent = set()
        st.session_state.snoozed = {}
        st.session_state.reminder_state_key = None
        st.session_state.notification_tick_key = None
    st.session_state.prev_hash = current_hash
    if enable_reminders and st.session_state.get("reminder_state_key") != schedule_key:
        st.session_state.reminder_sent = set()
        st.session_state.snoozed = {}
        # Load persisted reminders from storage
        for idx, row in df_raw.iterrows():
            try:
                row_id = row.get('REMINDER_ROW_ID')
                if pd.notna(row_id):
                    until_raw = row.get('REMINDER_SNOOZE_UNTIL')
                    until_epoch = None
                    if pd.notna(until_raw) and until_raw != "":
                        try:
                            # Normalize numeric strings
                            if isinstance(until_raw, str) and until_raw.strip().isdigit():
                                until_raw = int(until_raw.strip())
                            if isinstance(until_raw, (int, float)):
                                val = int(until_raw)
                                # Legacy values were stored as minutes since midnight (small numbers)
                                if val < 100000:
                                    midnight_ist = datetime(now.year, now.month, now.day, tzinfo=IST)
                                    until_epoch = int(midnight_ist.timestamp()) + (val * 60)
                                else:
                                    until_epoch = val
                            elif isinstance(until_raw, str):
                                s = until_raw.strip().replace("Z", "+00:00")
                                dt = datetime.fromisoformat(s)
                                until_epoch = int(dt.timestamp())
                        except Exception:
                            until_epoch = None
                    if until_epoch is not None and until_epoch > now_epoch:
                        st.session_state.snoozed[row_id] = until_epoch
                    dismissed = row.get('REMINDER_DISMISSED')
                    if str(dismissed).strip().upper() in ['TRUE','1','T','YES']:
                        st.session_state.reminder_sent.add(row_id)
            except Exception:
                continue
        st.session_state.reminder_state_key = schedule_key
    tick_key = _notification_tick_key(current_hash)
    if st.session_state.get("notification_tick_key") != tick_key:
        # Ensure Is_Ongoing column exists before using it
        if "Is_Ongoing" not in df.columns:
            df["Is_Ongoing"] = (df["In_min"] <= current_min) & (current_min <= df["Out_min"])
        # Currently Ongoing (filtered)
        ongoing_df = df[
            df["Is_Ongoing"] &
            ~df["STATUS"].astype(str).str.upper().str.contains("CANCELLED|DONE|COMPLETED|SHIFTED", na=True)
        ]
        current_ongoing = set(ongoing_df["Patient Name"].dropna())
        # New ongoing (either from time passing or manual status update)
        new_ongoing = current_ongoing - st.session_state.prev_ongoing
        for patient in new_ongoing:
            row = ongoing_df[ongoing_df["Patient Name"] == patient].iloc[0]
            st.toast(f"🚨 NOW ONGOING: {patient} – {row['Procedure']} with {row['DR.']} (Chair {row['OP']})", icon="🟢")
        # Upcoming in next 15 minutes
        upcoming_min = current_min + 15
        upcoming_df = df[
            (df["In_min"] > current_min) &
            (df["In_min"] <= upcoming_min) &
            ~df["STATUS"].astype(str).str.upper().str.contains("CANCELLED|DONE|COMPLETED|SHIFTED", na=True)
        ]
        current_upcoming = set(upcoming_df["Patient Name"].dropna())
        # New upcoming (just entered the 15-minute window)
        new_upcoming = current_upcoming - st.session_state.prev_upcoming
        for patient in new_upcoming:
            row = upcoming_df[upcoming_df["Patient Name"] == patient].iloc[0]
            mins_left = row["In_min"] - current_min
            st.toast(f"⏰ Upcoming in ~{mins_left} min: {patient} – {row['Procedure']} with {row['DR.']}", icon="⚠️")
        # New arrivals (manual status change in Excel)
        current_arrived = set(df_raw[df_raw["STATUS"].astype(str).str.upper() == "ARRIVED"]["Patient Name"].dropna())
        if ("STATUS" in st.session_state.prev_raw.columns) and ("Patient Name" in st.session_state.prev_raw.columns):
            prev_arrived = set(
                st.session_state.prev_raw[
                    st.session_state.prev_raw["STATUS"].astype(str).str.upper() == "ARRIVED"
                ]["Patient Name"].dropna()
            )
        else:
            prev_arrived = set()
        new_arrived = current_arrived - prev_arrived
        for patient in new_arrived:
            row = df[df["Patient Name"] == patient].iloc[0]
            st.toast(f"👤 Patient ARRIVED: {patient} – {row['Procedure']}", icon="🟡")
        # Update session state for next run
        st.session_state.prev_ongoing = current_ongoing
        st.session_state.prev_upcoming = current_upcoming
        st.session_state.prev_raw = df_raw.copy()
        st.session_state.notification_tick_key = tick_key
    # ================ 15-Minute Reminder System ================
    if enable_reminders:
        # Clean up expired snoozes
        expired = [rid for rid, until in list(st.session_state.snoozed.items()) if until <= now_epoch]
        for rid in expired:
            del st.session_state.snoozed[rid]
            # Don't persist clears on natural expiry; we'll overwrite when re-snoozing.
        # Find patients needing reminders (0-15 min before In Time)
        reminder_df = df[
            (df["In_min"].notna()) &
            (df["In_min"] - current_min > 0) &
            (df["In_min"] - current_min <= 15) &
            ~df["STATUS"].astype(str).str.upper().str.contains("CANCELLED|DONE|COMPLETED|SHIFTED|ARRIVED|ARRIVING|ON GOING|ONGOING", na=True)
        ].copy()
        # Show toast for new reminders (not snoozed, not dismissed)
        for idx, row in reminder_df.iterrows():
            row_id = row.get('REMINDER_ROW_ID')
            if pd.isna(row_id):
                continue
            patient = row.get("Patient Name", "Unknown")
            mins_left = int(row["In_min"] - current_min)
            # Skip if snoozed (still active) or dismissed
            snooze_until = st.session_state.snoozed.get(row_id)
            if (snooze_until is not None and snooze_until > now_epoch) or (row_id in st.session_state.reminder_sent):
                continue
            assistants = ", ".join(
                [
                    a
                    for a in [
                        str(row.get("FIRST", "")).strip(),
                        str(row.get("SECOND", "")).strip(),
                        str(row.get("Third", "")).strip(),
                    ]
                    if a and a.lower() not in {"nan", "none"}
                ]
            )
            assistants_text = f" | Assist: {assistants}" if assistants else ""
            st.toast(
                f"🔔 Reminder: {patient} in ~{mins_left} min at {row['In Time Str']} with {row.get('DR.','')} (OP {row.get('OP','')}){assistants_text}",
                icon="🔔",
            )
            # Auto-snooze for 30 seconds, and re-alert until status changes.
            next_until = now_epoch + 30
            st.session_state.snoozed[row_id] = next_until
            _persist_reminder_to_storage(row_id, next_until, False)
        # Reminder management UI
        def _safe_key(s):
            return re.sub(r"\W+", "_", str(s))
        with st.expander("🔔 Manage Reminders", expanded=False):
            if reminder_df.empty:
                st.caption("No upcoming appointments in the next 15 minutes.")
            else:
                for idx, row in reminder_df.iterrows():
                    row_id = row.get('REMINDER_ROW_ID')
                    if pd.isna(row_id):
                        continue
                    patient = row.get('Patient Name', 'Unknown')
                    mins_left = int(row["In_min"] - current_min)
                    assistants = ", ".join(
                        [
                            a
                            for a in [
                                str(row.get("FIRST", "")).strip(),
                                str(row.get("SECOND", "")).strip(),
                                str(row.get("Third", "")).strip(),
                            ]
                            if a and a.lower() not in {"nan", "none"}
                        ]
                    )
                    assistants_text = f" — Assist: {assistants}" if assistants else ""
                    col1, col2, col3, col4, col5 = st.columns([4,1,1,1,1])
                    col1.markdown(
                        f"**{patient}** — {row.get('Procedure','')} (in ~{mins_left} min at {row.get('In Time Str','')}){assistants_text}"
                    )  
                    default_snooze_seconds = int(st.session_state.get("default_snooze_seconds", 30))
                    if col2.button(f"💤 {default_snooze_seconds}s", key=f"snooze_{_safe_key(row_id)}_default"):
                        until = now_epoch + default_snooze_seconds
                        st.session_state.snoozed[row_id] = until
                        st.session_state.reminder_sent.discard(row_id)
                        _persist_reminder_to_storage(row_id, until, False)
                        st.toast(f"😴 Snoozed {patient} for {default_snooze_seconds} sec", icon="💤")
                        st.rerun()
                    if col3.button("💤 30s", key=f"snooze_{_safe_key(row_id)}_30s"):
                        until = now_epoch + 30
                        st.session_state.snoozed[row_id] = until
                        st.session_state.reminder_sent.discard(row_id)
                        _persist_reminder_to_storage(row_id, until, False)
                        st.toast(f"😴 Snoozed {patient} for 30 sec", icon="💤")
                        st.rerun()
                    if col4.button("💤 60s", key=f"snooze_{_safe_key(row_id)}_60s"):
                        until = now_epoch + 60
                        st.session_state.snoozed[row_id] = until
                        st.session_state.reminder_sent.discard(row_id)
                        _persist_reminder_to_storage(row_id, until, False)
                        st.toast(f"😴 Snoozed {patient} for 60 sec", icon="💤")
                        st.rerun()
                    if col5.button("🗑️", key=f"dismiss_{_safe_key(row_id)}"):
                        st.session_state.reminder_sent.add(row_id)
                        _persist_reminder_to_storage(row_id, None, True)
                        st.toast(f"✅ Dismissed reminder for {patient}", icon="✅")
                        st.rerun()
                # Show snoozed reminders
                if st.session_state.snoozed:
                    st.markdown("---")
                    st.markdown("**Snoozed Reminders**")
                    for row_id, until in list(st.session_state.snoozed.items()):
                        remaining_sec = int(until - now_epoch)
                        if remaining_sec > 0:
                            match_row = df[df.get('REMINDER_ROW_ID') == row_id]
                            if not match_row.empty:
                                name = match_row.iloc[0].get('Patient Name', row_id)
                                c1, c2 = st.columns([4,1])
                                c1.write(f"🕐 {name} — {remaining_sec} sec remaining")
                                if c2.button("Cancel", key=f"cancel_{_safe_key(row_id)}"):
                                    del st.session_state.snoozed[row_id]
                                    _persist_reminder_to_storage(row_id, None, False)
                                    st.toast(f"✅ Cancelled snooze for {name}", icon="✅")
                                    st.rerun()
# Sidebar header + attendance punch widget
with st.sidebar:
    st.markdown('<div class="sidebar-title">🦷 TDB Dashboard</div>', unsafe_allow_html=True)
    st.markdown('<div class="live-pill"><span class="live-dot"></span> Live • Auto refresh</div>', unsafe_allow_html=True)
    st.divider()
    schedule_for_punch = df if "df" in locals() else df_raw if "df_raw" in locals() else pd.DataFrame()
    try:
        if USE_SUPABASE and supabase_client is not None:
            sidebar_punch_widget_supabase(schedule_for_punch, supabase_client)
        else:
            sidebar_punch_widget(schedule_for_punch, file_path)
    except Exception as e:
        st.caption(f"Punch widget unavailable: {e}")
    st.divider()
    try:
        # Duties now use Excel backend - supabase parameter kept for backward compatibility
        render_duty_reminder_widget(schedule_for_punch, None)
    except Exception as e:
        st.caption(f"Duty reminder unavailable: {e}")
    st.divider()
# ================ MAIN DASHBOARD NAVIGATION ================
category = st.sidebar.radio(
    "Categories",
    ["Scheduling", "Assistants", "Doctors", "Admin/Settings"],
    index=0,
    key="nav_category",
)
s_sidebar_role_options = ["admin", "editor", "viewer"]
# Role/current user controls removed as requested
sched_view = assist_view = doctor_view = admin_view = None
if category == "Scheduling":
    sched_view = st.sidebar.radio(
        "Scheduling",
        ["Full Schedule", "Schedule by OP", "Ongoing", "Upcoming"],
        index=0,
        key="nav_sched",
    )
elif category == "Assistants":
    assist_view = st.sidebar.radio(
        "Assistants",
        ["Manage Profiles", "Availability", "Auto Allocation", "Workload", "Attendance"],
        index=0,
        key="nav_assist",
    )
elif category == "Doctors":
    doctor_view = st.sidebar.radio(
        "Doctors",
        ["Manage Profiles", "Overview", "Summary", "Per-Doctor Schedule"],
        index=0,
        key="nav_doc",
    )
else:
    admin_view = st.sidebar.radio(
        "Admin/Settings",
        ["Storage/Backup", "Notifications", "Duties Manager"],
        index=0,
        key="nav_admin",
    )
if category == "Assistants" and assist_view == "Manage Profiles":
    render_profile_manager(PROFILE_ASSISTANT_SHEET, "Assistant", "Department")
if category == "Doctors" and doctor_view == "Manage Profiles":
    render_profile_manager(PROFILE_DOCTOR_SHEET, "Doctor", "Department")
if category == "Doctors" and doctor_view == "Overview":
    def render_doctor_overview():
        st.markdown("### 🩺 Doctors Overview")
        today_idx = now.weekday()
        tomorrow_idx = (today_idx + 1) % 7
        weekday_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
        doctors_df = load_profiles(PROFILE_DOCTOR_SHEET)
        doctors_df["status"] = doctors_df.get("status", "ACTIVE").astype(str).str.upper()
        total = len(doctors_df)
        active = (doctors_df["status"] == "ACTIVE").sum() if not doctors_df.empty else 0
        def _off_list(idx):
            if "weekly_off" not in doctors_df.columns:
                return []
            offs = []
            for _, row in doctors_df.iterrows():
                days = _parse_weekly_off_days(row.get("weekly_off", ""))
                if idx in days:
                    offs.append(str(row.get("name", "")).strip().upper())
            return offs
        today_off = _off_list(today_idx)
        tomorrow_off = _off_list(tomorrow_idx)
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Total Doctors", total)
        c2.metric("Active", active)
        c3.metric(f"Off Today ({weekday_names[today_idx]})", len(today_off))
        c4.metric(f"Off Tomorrow ({weekday_names[tomorrow_idx]})", len(tomorrow_off))
        st.markdown("#### Today's Off Doctors")
        if today_off:
            st.warning(", ".join(sorted(today_off)))
        else:
            st.success("All doctors available today.")
        st.markdown("#### Tomorrow's Off Doctors")
        if tomorrow_off:
            st.info(", ".join(sorted(tomorrow_off)))
        else:
            st.success("All doctors available tomorrow.")
        # Card view
        st.markdown("#### Cards")
        entries = []
        for _, r in doctors_df.iterrows():
            name = str(r.get("name", "")).title()
            dept = str(r.get("department", "")).title()
            wo_days = _parse_weekly_off_days(r.get("weekly_off", ""))
            status = "BLOCKED" if today_idx in wo_days else "FREE"
            entries.append({
                "name": name,
                "info": {
                    "status": status,
                    "reason": f"Weekly off {weekday_names[today_idx]}" if today_idx in wo_days else "Available",
                    "department": dept or "N/A",
                }
            })
        _render_assistant_cards(entries)
    render_doctor_overview()
if category == "Scheduling":
    # ================ Status Colors ================
    def get_status_background(status):
        # Return subtle styling without bright backgrounds
        s = str(status).strip().upper()
        if "ON GOING" in s or "ONGOING" in s:
            return f"border-left: 4px solid {COLORS['success']}"
        elif "DONE" in s or "COMPLETED" in s:
            return f"border-left: 4px solid {COLORS['info']}"
        elif "CANCELLED" in s:
            return f"border-left: 4px solid {COLORS['danger']}"
        elif "ARRIVED" in s:
            return f"border-left: 4px solid {COLORS['warning']}"
        elif "LATE" in s:
            return f"border-left: 4px solid {COLORS['warning']}"
        elif "SHIFTED" in s:
            return f"border-left: 4px solid {COLORS['button_bg']}"
        return ""
    
    def highlight_row(row):
        color = get_status_background(row["STATUS"])
        return [color for _ in row]
    
    all_sorted = df
    
    # Manual save button and patient controls for schedule editor
    st.markdown("### 📋 Full Schedule")
    render_schedule_summary_chips(all_sorted)
    
    if not st.session_state.get("auto_save_enabled", False):
        st.caption("Auto-save is OFF. Use 'Save Changes' to persist updates.")
    if st.session_state.get("pending_changes"):
        st.warning("You have pending changes that are not saved yet.")
    
    # Add new patient button and save button
    
    # Automatically add a new empty patient row if the last row is not empty
    def is_row_empty(row):
        # Consider a row empty if Patient Name, In Time, Out Time, Procedure, DR. are all blank/None
        return all(
            not str(row.get(col, '')).strip()
            for col in ["Patient Name", "In Time", "Out Time", "Procedure", "DR."]
        )
    
    if not df_raw.empty:
        last_row = df_raw.iloc[-1]
        if not is_row_empty(last_row):
            # Add a new empty row
            new_row = {
                "Patient ID": "",
                "Patient Name": "",
                "In Time": None,
                "Out Time": None,
                "Procedure": "",
                "DR.": "",
                "FIRST": "",
                "SECOND": "",
                "Third": "",
                "CASE PAPER": "",
                "OP": "",
                "SUCTION": False,
                "CLEANING": False,
                "STATUS": "WAITING",
                "REMINDER_ROW_ID": str(uuid.uuid4()),
                "REMINDER_SNOOZE_UNTIL": pd.NA,
                "REMINDER_DISMISSED": False
            }
            df_raw = pd.concat([df_raw, pd.DataFrame([new_row])], ignore_index=True)
    
    col_add, col_save, col_del_pick, col_del_btn = st.columns([0.15, 0.20, 0.30, 0.15])
    with col_add:
        if st.button(
            "➕ Add Patient",
            key="add_patient_btn",
            use_container_width=True,
        ):
            # Create a new empty row
            new_row = {
                "Patient ID": "",
                "Patient Name": "",
                "In Time": None,
                "Out Time": None,
                "Procedure": "",
                "DR.": "",
                "FIRST": "",
                "SECOND": "",
                "Third": "",
                "CASE PAPER": "",
                "OP": "",
                "SUCTION": False,
                "CLEANING": False,
                "STATUS": "WAITING",
                "REMINDER_ROW_ID": str(uuid.uuid4()),
                "REMINDER_SNOOZE_UNTIL": pd.NA,
                "REMINDER_DISMISSED": False
            }
            # Append to the dataframe
            new_row_df = pd.DataFrame([new_row])
            df_raw_with_new = pd.concat([df_raw, new_row_df], ignore_index=True)
            # FORCE save the updated dataframe immediately
            _maybe_save(df_raw_with_new, show_toast=False, message="New patient row added!", force=True)
            st.success("New patient row added!")
            # Clear cache so fresh data loads on rerun
            if "cached_df_raw" in st.session_state:
                del st.session_state.cached_df_raw
            if "cached_df_timestamp" in st.session_state:
                del st.session_state.cached_df_timestamp
            st.rerun()
    with col_save:
        # Save button for the data editor
        if st.button(
            "?? Save Changes",
            key="manual_save_full",
            use_container_width=True,
            type="primary",
            disabled=bool(st.session_state.get("is_saving")) or bool(st.session_state.get("save_conflict")),
        ):
            st.session_state.manual_save_triggered = True
    
    with col_del_pick:
        # Compact delete row control (uses stable REMINDER_ROW_ID)
        try:
            candidates = df_raw.copy()
            if "Patient Name" in candidates.columns:
                candidates["Patient Name"] = candidates["Patient Name"].astype(str).replace("nan", "").fillna("")
            if "REMINDER_ROW_ID" in candidates.columns:
                candidates["REMINDER_ROW_ID"] = candidates["REMINDER_ROW_ID"].astype(str).replace("nan", "").fillna("")
    
            candidates = candidates[
                (candidates.get("REMINDER_ROW_ID", "").astype(str).str.strip() != "")
            ]
    
            option_map: dict[str, str] = {}
            if not candidates.empty:
                for row_ix, r in candidates.iterrows():
                    rid = str(r.get("REMINDER_ROW_ID", "")).strip()
                    if not rid:
                        continue
                    pname_raw = str(r.get("Patient Name", "")).strip()
                    pname = pname_raw if pname_raw else "(blank row)"
                    in_t = str(r.get("In Time", "")).strip()
                    op = str(r.get("OP", "")).strip()
                    row_no = f"#{int(row_ix) + 1}" if str(row_ix).isdigit() else str(row_ix)
                    label = " · ".join([p for p in [row_no, pname, in_t, op] if p])
                    # Make option text unique even if labels repeat.
                    opt = f"{label} — {rid[:8]}" if label else rid[:8]
                    option_map[opt] = rid
    
            if "delete_row_id" not in st.session_state:
                st.session_state.delete_row_id = ""
    
            if option_map:
                # Use a visible sentinel option instead of `placeholder` for wider Streamlit compatibility.
                # Also: guard against Streamlit selectbox failing when the previously selected value
                # is no longer present in the new options list (common after edits/deletes).
                sentinel = "Select row to delete…"
                options = [sentinel] + sorted(option_map.keys())
    
                # IMPORTANT: Do not mutate st.session_state["delete_row_select"] here.
                # Streamlit raises if you modify a widget key after it has been instantiated.
                prev_choice = st.session_state.get("delete_row_select", sentinel)
                default_index = options.index(prev_choice) if prev_choice in options else 0
    
                chosen = st.selectbox(
                    "Delete row",
                    options=options,
                    key="delete_row_select",
                    label_visibility="collapsed",
                    index=default_index,
                )
                if chosen and chosen != sentinel:
                    st.session_state.delete_row_id = option_map.get(chosen, "")
                else:
                    st.session_state.delete_row_id = ""
            else:
                st.session_state.delete_row_id = ""
                st.caption("Delete row")
        except Exception:
            # Keep dashboard usable even if data is incomplete
            st.caption("Delete row")
    
    with col_del_btn:
        if st.button("⌫", key="delete_row_btn", help="Delete selected row"):
            rid = str(st.session_state.get("delete_row_id", "") or "").strip()
            if not rid:
                st.warning("Select a row to delete")
            else:
                try:
                    if "REMINDER_ROW_ID" not in df_raw.columns:
                        raise ValueError("Missing REMINDER_ROW_ID column")
                    df_updated = df_raw[df_raw["REMINDER_ROW_ID"].astype(str) != rid].copy()
    
                    # Clear local reminder state for this row id.
                    try:
                        if "snoozed" in st.session_state and rid in st.session_state.snoozed:
                            del st.session_state.snoozed[rid]
                        if "reminder_sent" in st.session_state:
                            st.session_state.reminder_sent.discard(rid)
                    except Exception:
                        pass
    
                    _maybe_save(df_updated, message="Row deleted")
                    st.session_state.delete_row_id = ""
                    st.rerun()
                except Exception as e:
                    st.error(f"Error deleting row: {e}")
    view_cols = st.columns([0.2, 0.8], gap="small")
    with view_cols[0]:
        view_mode = st.radio(
            "View",
            ["Cards", "Table"],
            horizontal=True,
            key="full_schedule_view_mode",
            label_visibility="collapsed",
        )
    with view_cols[1]:
        card_search = st.text_input(
            "Search schedule",
            value="",
            key="full_schedule_card_search",
            placeholder="Search schedule...",
            label_visibility="collapsed",
        )
    display_all = all_sorted[[
        "Patient Name",
        "In Time Obj",
        "Out Time Obj",
        "Procedure",
        "DR.",
        "FIRST",
        "SECOND",
        "Third",
        "CASE PAPER",
        "OP",
        "SUCTION",
        "CLEANING",
        "REMINDER_ROW_ID",
        "STATUS",
        "STATUS_CHANGED_AT",
        "ACTUAL_START_AT",
        "ACTUAL_END_AT",
    ]].copy()
    display_all = display_all.rename(columns={"In Time Obj": "In Time", "Out Time Obj": "Out Time"})
    # Preserve original index for mapping edits back to df_raw
    display_all["_orig_idx"] = display_all.index
    display_all = display_all.reset_index(drop=True)
    
    # Convert text columns to string to avoid type compatibility issues (BUT NOT TIME/BOOL COLUMNS)
    for col in ["Patient Name", "Procedure", "DR.", "FIRST", "SECOND", "Third", "CASE PAPER", "OP", "STATUS"]:
        if col in display_all.columns:
            display_all[col] = display_all[col].astype(str).replace('nan', '')
    
    # Keep In Time and Out Time as time objects for proper display
    display_all["In Time"] = display_all["In Time"].apply(lambda v: v if isinstance(v, time_type) else None)
    display_all["Out Time"] = display_all["Out Time"].apply(lambda v: v if isinstance(v, time_type) else None)
    
    # Computed overtime indicator (uses scheduled Out Time vs current time)
    def _compute_overtime_min(_row) -> Optional[int]:
        try:
            s = str(_row.get("STATUS", "")).strip().upper()
            if ("ON GOING" not in s) and ("ONGOING" not in s):
                return None
            out_min = _row.get("Out_min")
            if pd.isna(out_min):
                return None
            diff = int(current_min) - int(out_min)
            return diff if diff > 0 else None
        except Exception:
            return None
    
    display_all["Overtime (min)"] = all_sorted.apply(_compute_overtime_min, axis=1)
    
    
    st.markdown(
        """
        <style>
        .full-schedule-cards {margin-top: 8px;}
        .schedule-card {background:#f8fafc; border:1px solid #3b82f6; border-radius:18px; padding:14px; box-shadow:0 10px 20px rgba(20,17,15,0.08); display:flex; flex-direction:column; gap:10px; min-height:220px;}
        .card-shell-marker {display:none;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) {background:linear-gradient(180deg, #ffffff 0%, #f2f4f7 100%); border:1px solid #e3e6ec; border-radius:24px; box-shadow:0 22px 44px rgba(24, 28, 36, 0.18); overflow:hidden;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) > div {padding:0 20px 18px 20px; display:flex; flex-direction:column; gap:12px; min-height:260px;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) [data-testid="stHorizontalBlock"] {gap: 0.6rem; align-items:center; justify-content:flex-start;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stButton>button,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) [data-testid="stButton"] > button,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind] {height: 18px !important; border-radius: 10px !important; font-weight: 700; text-transform: none; letter-spacing: 0; white-space: nowrap; word-break: keep-all; overflow-wrap: normal; min-width: 60px; padding: 0 6px !important; font-size: 10px; line-height: 1; flex-shrink: 0; display: inline-flex; align-items: center; justify-content: center; gap: 4px; width: 100%; box-shadow: 0 6px 14px rgba(22, 24, 31, 0.14); color:#3f434a;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stButton>button *,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind] * {white-space: nowrap;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="stHorizontalBlock"]:has(.card-action-marker) {flex-wrap: wrap; row-gap: 0.5rem; align-items:center; justify-content:flex-end;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="stHorizontalBlock"]:has(.card-action-marker) > div {min-width: 96px; flex: 1 1 96px;}
        @media (min-width: 1100px) {
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="stHorizontalBlock"]:has(.card-action-marker) {flex-wrap: nowrap;}
            div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="stHorizontalBlock"]:has(.card-action-marker) > div {min-width: 100px; flex: 0 0 auto;}
        }
        .card-actions-row {
            margin-top: 8px;
            display: flex;
            flex-wrap: wrap;
            gap: 0.5rem;
            align-items: center;
            justify-content: flex-start;
        }
        .card-actions-row .stCheckbox {
            margin: 0;
        }
        .card-actions-row .stCheckbox label {
            font-size: 13px;
            font-weight: 600;
            letter-spacing: 0.1px;
        }
        .card-actions-row .stButton>button,
        .card-actions-row button[kind] {
            min-width: 70px;
            border-radius: 10px !important;
            height: 30px !important;
            font-size: 12px !important;
        }
        .card-actions-row div[data-testid="stHorizontalBlock"] {align-items:center;}
        .card-actions-row div[data-testid="stHorizontalBlock"]:has(.stCheckbox) {justify-content:flex-start !important;}
        .card-actions-row div[data-testid="stHorizontalBlock"]:has(.stCheckbox) div[data-testid="column"]:has(.card-action-done) {margin-left:auto;}
        .card-actions-row div[data-testid="column"]:has(.card-action-done),
        .card-actions-row div[data-testid="column"]:has(.card-action-edit),
        .card-actions-row div[data-testid="column"]:has(.card-action-cancel) {min-width: 72px;}
        .card-details-row {
            margin-top: 10px;
            border: 1px solid #d9dde3;
            border-radius: 14px;
            padding: 6px 10px;
            background: #fff;
            box-shadow: 0 10px 20px rgba(24, 28, 36, 0.12);
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox {margin-top: 10px;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox label {font-size: 14px; font-weight: 600; color:#2f333a; white-space: normal; line-height: 1.2;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox div[data-baseweb="checkbox"] > div,
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox div[data-baseweb="checkbox"] > label > div {width: 22px; height: 22px; border-radius: 6px; border: 1.5px solid #c3c8d0; background: #ffffff;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) .stCheckbox div[data-baseweb="checkbox"] input:checked + div {background:#2f63e8; border-color:#2f63e8;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind="primary"] {background:#2f63e8 !important; border:1px solid #2f63e8 !important; color:#ffffff !important; box-shadow:0 8px 18px rgba(47,99,232,0.28) !important;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) button[kind="secondary"] {background:#ffffff !important; border:1px solid #d5d8de !important; color:#4b4f56 !important; box-shadow:0 6px 14px rgba(24, 28, 36, 0.08) !important;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="column"]:has(.card-action-cancel) button {border-color:#e1b0b0 !important; color:#b15454 !important; background:#ffffff !important;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) div[data-testid="stHorizontalBlock"]:has(.stCheckbox) div[data-testid="column"]:has(.card-action-done) {margin-left:auto;}
        .card-action-marker {display:none;}
        .card-status-banner {display:flex; align-items:center; gap:10px; padding:12px 20px; border-radius:20px 20px 12px 12px; font-weight:800; font-size:13px; letter-spacing:0.8px; text-transform:uppercase; margin:0 -20px 14px -20px;}
        .card-status-banner.waiting {background:linear-gradient(90deg, #f7e6b7, #fff2d6); color:#8a775b;}
        .card-status-banner.ongoing {background:linear-gradient(90deg, #dfe9ff, #f1f5ff); color:#2f4f86;}
        .card-status-banner.arrived {background:linear-gradient(90deg, #e9e9ea, #f5f5f6); color:#50545a;}
        .card-status-banner.completed {background:linear-gradient(90deg, #def3e6, #eef8f1); color:#3d6b4a;}
        .card-status-banner.cancelled {background:linear-gradient(90deg, #f6d1d1, #fde8e8); color:#9a4b4b;}
        .status-dot {width:14px; height:14px; border-radius:50%;}
        .card-status-banner.waiting .status-dot {background:#f1b400; box-shadow:0 0 0 4px rgba(241,180,0,0.22);}
        .card-status-banner.ongoing .status-dot {background:#3b6fd8; box-shadow:0 0 0 4px rgba(59,111,216,0.22);}
        .card-status-banner.arrived .status-dot {background:#7a7a7a; box-shadow:0 0 0 4px rgba(122,122,122,0.22);}
        .card-status-banner.completed .status-dot {background:#4caf6b; box-shadow:0 0 0 4px rgba(76,175,107,0.22);}
        .card-status-banner.cancelled .status-dot {background:#d45c5c; box-shadow:0 0 0 4px rgba(212,92,92,0.22);}
        .card-head {display:flex; align-items:center; gap:16px;}
        .card-title {display:flex; flex-direction:column; gap:3px;}
        .card-avatar {width:56px; height:56px; border-radius:50%; background:radial-gradient(circle at 30% 30%, #f6e4c9, #e4cca4); border:1px solid #ead8be; color:#5d4a35; font-weight:800; display:flex; align-items:center; justify-content:center; font-size:16px; box-shadow: inset 0 1px 0 rgba(255,255,255,0.7);}
        .card-name {font-size:18px; font-weight:800; color:#2a2d33; letter-spacing:0.4px; text-transform:uppercase;}
        .card-time {font-size:13px; color:#6f757d;}
        .card-info {
            display: flex;
            flex-direction: column;
            gap: 12px;
            margin: 16px 0;
        }
        .info-row {
            display: flex;
            align-items: center;
            gap: 12px;
            padding: 10px 12px;
            border-radius: 12px;
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.6), rgba(255, 255, 255, 0.3));
            backdrop-filter: blur(8px);
            -webkit-backdrop-filter: blur(8px);
            border: 1px solid rgba(255, 255, 255, 0.3);
            transition: all 0.3s ease;
        }
        .info-row:hover {
            background: linear-gradient(135deg, rgba(255, 255, 255, 0.8), rgba(255, 255, 255, 0.5));
            border-color: rgba(59, 130, 246, 0.2);
            transform: translateX(4px);
        }
        .info-icon {
            width: 36px;
            height: 36px;
            border-radius: 12px;
            background: linear-gradient(135deg, rgba(59, 130, 246, 0.1), rgba(37, 99, 235, 0.05));
            border: 1px solid rgba(59, 130, 246, 0.15);
            display: flex;
            align-items: center;
            justify-content: center;
            color: #2563eb;
            backdrop-filter: blur(8px);
            -webkit-backdrop-filter: blur(8px);
            transition: all 0.3s ease;
            flex-shrink: 0;
        }
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker):hover .info-icon {
            background: linear-gradient(135deg, rgba(59, 130, 246, 0.15), rgba(37, 99, 235, 0.08));
            border-color: rgba(59, 130, 246, 0.25);
            transform: scale(1.05);
        }
        .info-icon.doctor-icon {font-size:16px;}
        .info-icon.staff-icon {font-size:16px;}
        .info-icon-svg {
            display: block;
            width: 20px;
            height: 20px;
        }
        .info-text {
            font-size: 14px;
            font-weight: 500;
            color: #1e293b;
            flex: 1;
        }
        .info-label {
            font-size: 11px;
            font-weight: 600;
            color: #64748b;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 2px;
        }
        .card-subdivider {height:1px; background:#e4e6eb; margin: 10px 0 6px;}
        .card-divider {height:1px; background:#e4e6eb; margin: 12px 0;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) [data-testid="stExpander"] {border:1px solid #d9dde3; border-radius:12px; background:#f7f8fa; margin-top:6px;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) summary {padding:10px 12px; font-weight:600; color:#60656c; display:flex; align-items:center; gap:10px; font-size:13px;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) summary::before {content:"›"; color:#7a8087; font-size:18px;}
        div[data-testid="stVerticalBlockBorderWrapper"]:has(.card-shell-marker) summary::after {content:"⋯"; margin-left:auto; color:#9aa0a7; font-size:18px;}
        </style>
        """,
        unsafe_allow_html=True,
    )
    def _clean_text(val) -> str:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return ""
        text = str(val).strip()
        if text.lower() in {"nan", "none"}:
            return ""
        return text
    def _truthy(val) -> bool:
        if isinstance(val, bool):
            return val
        text = _clean_text(val).lower()
        return text in {"yes", "y", "true", "1", "done", "checked"}
    def _initials(name: str) -> str:
        parts = [p for p in name.strip().split() if p]
        if not parts:
            return "--"
        if len(parts) == 1:
            return parts[0][:2].upper()
        return (parts[0][0] + parts[-1][0]).upper()
    def _status_class(status: str) -> str:
        status_up = status.upper()
        if "WAIT" in status_up:
            return "waiting"
        if "ONGOING" in status_up or "ON GOING" in status_up:
            return "ongoing"
        if "ARRIVED" in status_up:
            return "arrived"
        if "DONE" in status_up or "COMPLETED" in status_up:
            return "completed"
        if "CANCEL" in status_up or "SHIFT" in status_up:
            return "cancelled"
        return "waiting"
    def _open_full_edit_dialog(context: dict[str, Any]) -> None:
        st.session_state["full_edit_context"] = context
        st.session_state["full_edit_open"] = True
        row_key = str(context.get("row_key", "")).strip()
        if not row_key:
            return
        in_time_value = str(context.get("in_time", "") or "").strip()
        out_time_value = str(context.get("out_time", "") or "").strip()
        if in_time_value.upper() in {"N/A", "NONE", "NAT"}:
            in_time_value = ""
        if out_time_value.upper() in {"N/A", "NONE", "NAT"}:
            out_time_value = ""
        st.session_state[f"full_popup_patient_{row_key}"] = str(context.get("patient", "") or "")
        in_hour, in_minute, in_ampm = _time_to_picker_parts(in_time_value)
        out_hour, out_minute, out_ampm = _time_to_picker_parts(out_time_value)
        st.session_state[f"full_popup_in_hour_{row_key}"] = in_hour
        st.session_state[f"full_popup_in_min_{row_key}"] = in_minute
        st.session_state[f"full_popup_in_ampm_{row_key}"] = in_ampm
        st.session_state[f"full_popup_out_hour_{row_key}"] = out_hour
        st.session_state[f"full_popup_out_min_{row_key}"] = out_minute
        st.session_state[f"full_popup_out_ampm_{row_key}"] = out_ampm
        st.session_state[f"full_popup_status_{row_key}"] = str(context.get("status", "") or "")
        st.session_state[f"full_popup_doctor_{row_key}"] = str(context.get("doctor", "") or "")
        st.session_state[f"full_popup_procedure_{row_key}"] = str(context.get("procedure", "") or "")
        st.session_state[f"full_popup_op_{row_key}"] = str(context.get("op", "") or "")
        st.session_state[f"full_popup_first_{row_key}"] = str(context.get("staff_first", "") or "")
        st.session_state[f"full_popup_second_{row_key}"] = str(context.get("staff_second", "") or "")
        st.session_state[f"full_popup_third_{row_key}"] = str(context.get("staff_third", "") or "")
        st.session_state[f"full_popup_case_{row_key}"] = bool(context.get("case_paper", False))
        st.session_state[f"full_popup_suction_{row_key}"] = bool(context.get("suction", False))
        st.session_state[f"full_popup_cleaning_{row_key}"] = bool(context.get("cleaning", False))
    def _close_full_edit_dialog() -> None:
        st.session_state["full_edit_open"] = False
        st.session_state["full_edit_context"] = {}
    def _full_normalize_time_input(raw_value: str) -> tuple[str, Optional[str]]:
        text = str(raw_value or "").strip()
        if not text:
            return "", None
        t = _coerce_to_time_obj(text)
        if t is None:
            return "", "Invalid time format. Use HH:MM or 09:30 AM."
        return f"{t.hour:02d}:{t.minute:02d}", None
    def _full_build_select_options(options: list[str], current_value: str) -> tuple[list[str], int]:
        current = str(current_value or "").strip()
        opts = [opt for opt in options if str(opt).strip()]
        if current and current not in opts:
            opts = [current] + opts
        opts = [""] + opts
        index = opts.index(current) if current in opts else 0
        return opts, index
    def _apply_full_card_edit(row_id, patient_name, in_time_val, updates: dict[str, Any]) -> bool:
        df_source = df_raw if "df_raw" in globals() else df
        if df_source is None or df_source.empty:
            st.warning("No schedule data to update.")
            return False
        df_updated = df_source.copy()
        idx = None
        if row_id and "REMINDER_ROW_ID" in df_updated.columns:
            matches = df_updated["REMINDER_ROW_ID"].astype(str) == str(row_id)
            if matches.any():
                idx = matches.idxmax()
        if idx is None and "Patient Name" in df_updated.columns and patient_name:
            name_mask = df_updated["Patient Name"].astype(str).str.upper() == str(patient_name).upper()
            if in_time_val and "In Time" in df_updated.columns:
                time_mask = df_updated["In Time"].astype(str) == str(in_time_val)
                match = df_updated[name_mask & time_mask]
            else:
                match = df_updated[name_mask]
            if not match.empty:
                idx = match.index[0]
        if idx is None:
            st.warning("Unable to locate row for update.")
            return False
        status_col = "STATUS" if "STATUS" in df_updated.columns else "Status" if "Status" in df_updated.columns else ""
        old_status_norm = str(df_updated.at[idx, status_col]).strip().upper() if status_col else ""
        for col, val in updates.items():
            if col in df_updated.columns:
                df_updated.at[idx, col] = val
        if status_col:
            new_status_norm = str(df_updated.at[idx, status_col]).strip().upper()
            if new_status_norm and new_status_norm != old_status_norm:
                ts = _now_iso()
                if "STATUS_CHANGED_AT" in df_updated.columns:
                    df_updated.at[idx, "STATUS_CHANGED_AT"] = ts
                if ("ONGOING" in new_status_norm or "ON GOING" in new_status_norm) and "ACTUAL_START_AT" in df_updated.columns:
                    if not str(df_updated.at[idx, "ACTUAL_START_AT"]).strip():
                        df_updated.at[idx, "ACTUAL_START_AT"] = ts
                if ("DONE" in new_status_norm or "COMPLETED" in new_status_norm) and "ACTUAL_END_AT" in df_updated.columns:
                    if not str(df_updated.at[idx, "ACTUAL_END_AT"]).strip():
                        df_updated.at[idx, "ACTUAL_END_AT"] = ts
                if "STATUS_LOG" in df_updated.columns:
                    existing_log = str(df_updated.at[idx, "STATUS_LOG"])
                    try:
                        df_updated.at[idx, "STATUS_LOG"] = _append_status_log(
                            existing_log,
                            {"at": ts, "from": old_status_norm, "to": new_status_norm},
                        )
                    except Exception:
                        df_updated.at[idx, "STATUS_LOG"] = existing_log
        if bool(st.session_state.get("auto_assign_assistants", True)):
            only_empty = bool(st.session_state.get("auto_assign_only_empty", True))
            _auto_fill_assistants_for_row(df_updated, int(idx), only_fill_empty=only_empty)
        _maybe_save(df_updated, show_toast=False, message=f"Updated {patient_name or 'patient'}")
        if st.session_state.get("auto_save_enabled", False):
            st.toast("Changes saved.", icon="✅")
        else:
            st.toast("Changes queued. Click 'Save Changes'.", icon="📝")
        return True
    def _render_full_edit_dialog_body() -> None:
        context = st.session_state.get("full_edit_context") or {}
        if not context:
            _close_full_edit_dialog()
            return
        row_key = str(context.get("row_key", "")).strip()
        if not row_key:
            _close_full_edit_dialog()
            return
        lookup_patient = str(context.get("lookup_patient", "") or "")
        lookup_in_time = str(context.get("lookup_in_time", "") or "")
        row_id = str(context.get("row_id", "") or "")
        with st.form(key=f"full_popup_form_{row_key}"):
            patient_input = st.text_input(
                "Patient Name",
                key=f"full_popup_patient_{row_key}",
            )
            time_cols = st.columns(2, gap="small")
            with time_cols[0]:
                with st.container():
                    st.markdown("<div class='time-select-marker'></div>", unsafe_allow_html=True)
                    st.markdown("In Time")
                    in_time_cols = st.columns(3, gap="small")
                    with in_time_cols[0]:
                        in_hour = st.selectbox(
                            "Hour",
                            options=TIME_PICKER_HOURS,
                            key=f"full_popup_in_hour_{row_key}",
                        )
                    with in_time_cols[1]:
                        in_minute = st.selectbox(
                            "Minute",
                            options=TIME_PICKER_MINUTES,
                            key=f"full_popup_in_min_{row_key}",
                        )
                    with in_time_cols[2]:
                        in_ampm = st.selectbox(
                            "AM/PM",
                            options=TIME_PICKER_AMPM,
                            key=f"full_popup_in_ampm_{row_key}",
                        )
            with time_cols[1]:
                with st.container():
                    st.markdown("<div class='time-select-marker'></div>", unsafe_allow_html=True)
                    st.markdown("Out Time")
                    out_time_cols = st.columns(3, gap="small")
                    with out_time_cols[0]:
                        out_hour = st.selectbox(
                            "Hour",
                            options=TIME_PICKER_HOURS,
                            key=f"full_popup_out_hour_{row_key}",
                        )
                    with out_time_cols[1]:
                        out_minute = st.selectbox(
                            "Minute",
                            options=TIME_PICKER_MINUTES,
                            key=f"full_popup_out_min_{row_key}",
                        )
                    with out_time_cols[2]:
                        out_ampm = st.selectbox(
                            "AM/PM",
                            options=TIME_PICKER_AMPM,
                            key=f"full_popup_out_ampm_{row_key}",
                        )
            top_cols = st.columns(2, gap="small")
            with top_cols[0]:
                doctor_current = st.session_state.get(f"full_popup_doctor_{row_key}", "")
                doctor_options, doctor_index = _full_build_select_options(DOCTOR_OPTIONS, doctor_current)
                doctor_input = st.selectbox(
                    "Doctor",
                    options=doctor_options,
                    index=doctor_index,
                    key=f"full_popup_doctor_{row_key}",
                )
            with top_cols[1]:
                procedure_input = st.text_input(
                    "Procedure",
                    key=f"full_popup_procedure_{row_key}",
                )
            mid_cols = st.columns(2, gap="small")
            with mid_cols[0]:
                op_input = st.text_input(
                    "OP",
                    key=f"full_popup_op_{row_key}",
                )
            with mid_cols[1]:
                status_current = st.session_state.get(f"full_popup_status_{row_key}", "")
                status_options, status_index = _full_build_select_options(STATUS_OPTIONS, status_current)
                status_input = st.selectbox(
                    "Status",
                    options=status_options,
                    index=status_index,
                    key=f"full_popup_status_{row_key}",
                )
            staff_cols = st.columns(3, gap="small")
            with staff_cols[0]:
                first_current = st.session_state.get(f"full_popup_first_{row_key}", "")
                first_options, first_index = _full_build_select_options(ASSISTANT_OPTIONS, first_current)
                first_input = st.selectbox(
                    "First",
                    options=first_options,
                    index=first_index,
                    key=f"full_popup_first_{row_key}",
                )
            with staff_cols[1]:
                second_current = st.session_state.get(f"full_popup_second_{row_key}", "")
                second_options, second_index = _full_build_select_options(ASSISTANT_OPTIONS, second_current)
                second_input = st.selectbox(
                    "Second",
                    options=second_options,
                    index=second_index,
                    key=f"full_popup_second_{row_key}",
                )
            with staff_cols[2]:
                third_current = st.session_state.get(f"full_popup_third_{row_key}", "")
                third_options, third_index = _full_build_select_options(ASSISTANT_OPTIONS, third_current)
                third_input = st.selectbox(
                    "Third",
                    options=third_options,
                    index=third_index,
                    key=f"full_popup_third_{row_key}",
                )
            flag_cols = st.columns(3, gap="small")
            with flag_cols[0]:
                case_paper_input = st.checkbox(
                    "QTRAQ",
                    key=f"full_popup_case_{row_key}",
                )
            with flag_cols[1]:
                suction_input = st.checkbox(
                    "Suction",
                    key=f"full_popup_suction_{row_key}",
                )
            with flag_cols[2]:
                cleaning_input = st.checkbox(
                    "Cleaning",
                    key=f"full_popup_cleaning_{row_key}",
                )
            form_actions = st.columns(2, gap="small")
            with form_actions[0]:
                save_clicked = st.form_submit_button("Save", use_container_width=True)
            with form_actions[1]:
                cancel_clicked = st.form_submit_button("Cancel", use_container_width=True)
        if cancel_clicked:
            _close_full_edit_dialog()
            st.rerun()
        if save_clicked:
            in_norm, in_err = _time_from_picker_parts(in_hour, in_minute, in_ampm)
            out_norm, out_err = _time_from_picker_parts(out_hour, out_minute, out_ampm)
            if in_err or out_err:
                if in_err:
                    st.error(in_err)
                if out_err:
                    st.error(out_err)
            else:
                updates = {
                    "Patient Name": str(patient_input or "").strip(),
                    "In Time": in_norm,
                    "Out Time": out_norm,
                    "Procedure": str(procedure_input or "").strip(),
                    "DR.": str(doctor_input or "").strip(),
                    "Doctor": str(doctor_input or "").strip(),
                    "OP": str(op_input or "").strip(),
                    "FIRST": str(first_input or "").strip(),
                    "SECOND": str(second_input or "").strip(),
                    "Third": str(third_input or "").strip(),
                    "THIRD": str(third_input or "").strip(),
                    "CASE PAPER": "Yes" if case_paper_input else "",
                    "SUCTION": bool(suction_input),
                    "CLEANING": bool(cleaning_input),
                    "STATUS": str(status_input or "").strip(),
                    "Status": str(status_input or "").strip(),
                }
                if _apply_full_card_edit(row_id, lookup_patient, lookup_in_time, updates):
                    _close_full_edit_dialog()
                    st.rerun()
    _dialog_decorator = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)
    if _dialog_decorator:
        @_dialog_decorator("Edit appointment")
        def _render_full_edit_dialog() -> None:
            _render_full_edit_dialog_body()
    else:
        def _render_full_edit_dialog() -> None:
            st.warning("Popup editing requires a newer Streamlit version.")
            _render_full_edit_dialog_body()
    def _fmt_time(val) -> str:
        if isinstance(val, time_type):
            return val.strftime("%I:%M %p").lstrip("0")
        return _clean_text(val)
    def _update_row_status(row_id, patient_name, in_time_val, new_status):
        df_source = df_raw if "df_raw" in globals() else df
        if df_source is None or df_source.empty:
            st.warning("No schedule data to update.")
            return
        df_updated = df_source.copy()
        idx = None
        if row_id and "REMINDER_ROW_ID" in df_updated.columns:
            matches = df_updated["REMINDER_ROW_ID"].astype(str) == str(row_id)
            if matches.any():
                idx = matches.idxmax()
        if idx is None and "Patient Name" in df_updated.columns and patient_name:
            name_mask = df_updated["Patient Name"].astype(str).str.upper() == str(patient_name).upper()
            if in_time_val and "In Time" in df_updated.columns:
                time_mask = df_updated["In Time"].astype(str) == str(in_time_val)
                match = df_updated[name_mask & time_mask]
            else:
                match = df_updated[name_mask]
            if not match.empty:
                idx = match.index[0]
        if idx is None:
            st.warning("Unable to locate row for update.")
            return
        old_status_norm = ""
        if "STATUS" in df_updated.columns:
            old_status_norm = str(df_updated.at[idx, "STATUS"]).strip().upper()
            df_updated.at[idx, "STATUS"] = new_status
        if "Status" in df_updated.columns:
            if not old_status_norm:
                old_status_norm = str(df_updated.at[idx, "Status"]).strip().upper()
            df_updated.at[idx, "Status"] = new_status
        ts = _now_iso()
        if "STATUS_CHANGED_AT" in df_updated.columns:
            df_updated.at[idx, "STATUS_CHANGED_AT"] = ts
        if ("ONGOING" in new_status or "ON GOING" in new_status) and "ACTUAL_START_AT" in df_updated.columns:
            if not str(df_updated.at[idx, "ACTUAL_START_AT"]).strip():
                df_updated.at[idx, "ACTUAL_START_AT"] = ts
        if ("DONE" in new_status or "COMPLETED" in new_status) and "ACTUAL_END_AT" in df_updated.columns:
            if not str(df_updated.at[idx, "ACTUAL_END_AT"]).strip():
                df_updated.at[idx, "ACTUAL_END_AT"] = ts
        if "STATUS_LOG" in df_updated.columns:
            existing_log = str(df_updated.at[idx, "STATUS_LOG"])
            try:
                df_updated.at[idx, "STATUS_LOG"] = _append_status_log(
                    existing_log,
                    {"at": ts, "from": old_status_norm, "to": new_status},
                )
            except Exception:
                df_updated.at[idx, "STATUS_LOG"] = existing_log
        _maybe_save(df_updated, message=f"Status set to {new_status} for {patient_name}")
        st.toast(f"{patient_name} marked {new_status}", icon="✅")
        st.rerun()
    def _update_row_case_paper(row_id, patient_name, in_time_val, case_checked: bool):
        df_source = df_raw if "df_raw" in globals() else df
        if df_source is None or df_source.empty:
            st.warning("No schedule data to update.")
            return
        df_updated = df_source.copy()
        idx = None
        if row_id and "REMINDER_ROW_ID" in df_updated.columns:
            matches = df_updated["REMINDER_ROW_ID"].astype(str) == str(row_id)
            if matches.any():
                idx = matches.idxmax()
        if idx is None and "Patient Name" in df_updated.columns and patient_name:
            name_mask = df_updated["Patient Name"].astype(str).str.upper() == str(patient_name).upper()
            if in_time_val and "In Time" in df_updated.columns:
                time_mask = df_updated["In Time"].astype(str) == str(in_time_val)
                match = df_updated[name_mask & time_mask]
            else:
                match = df_updated[name_mask]
            if not match.empty:
                idx = match.index[0]
        if idx is None:
            st.warning("Unable to locate row for update.")
            return
        if "CASE PAPER" not in df_updated.columns:
            st.warning("No QTRAQ column to update.")
            return
        df_updated.at[idx, "CASE PAPER"] = "Yes" if case_checked else ""
        _maybe_save(df_updated, message=f"Case paper updated for {patient_name}")
        st.toast(f"{patient_name} case paper updated")
        st.rerun()
    edited_all = None
    if view_mode == "Table":
        edited_all = st.data_editor(
            display_all,
            width="stretch",
            key="full_schedule_editor",
            hide_index=True,
            disabled=["STATUS_CHANGED_AT", "ACTUAL_START_AT", "ACTUAL_END_AT", "Overtime (min)"],
            column_config={
                "_orig_idx": None,  # Hide the original index column
                "REMINDER_ROW_ID": None,
                "Patient Name": st.column_config.TextColumn(label="Patient Name"),
                "In Time": st.column_config.TimeColumn(label="In Time", format="hh:mm A"),
                "Out Time": st.column_config.TimeColumn(label="Out Time", format="hh:mm A"),
                "Procedure": st.column_config.TextColumn(label="Procedure"),
                "DR.": st.column_config.SelectboxColumn(
                    label="DR.",
                    options=DOCTOR_OPTIONS,
                    required=False,
                ),
                "OP": st.column_config.SelectboxColumn(
                    label="OP",
                    options=["OP 1", "OP 2", "OP 3", "OP 4"],
                    required=False,
                ),
                "FIRST": st.column_config.SelectboxColumn(
                    label="FIRST",
                    options=ASSISTANT_OPTIONS,
                    required=False,
                ),
                "SECOND": st.column_config.SelectboxColumn(
                    label="SECOND",
                    options=ASSISTANT_OPTIONS,
                    required=False,
                ),
                "Third": st.column_config.SelectboxColumn(
                    label="Third",
                    options=ASSISTANT_OPTIONS,
                    required=False,
                ),
                "CASE PAPER": st.column_config.SelectboxColumn(
                    label="QTRAQ",
                    options=ASSISTANT_OPTIONS,
                    required=False,
                ),
                "SUCTION": st.column_config.CheckboxColumn(label="SUCTION"),
                "CLEANING": st.column_config.CheckboxColumn(label="CLEANING"),
                "STATUS_CHANGED_AT": None,
                "ACTUAL_START_AT": None,
                "ACTUAL_END_AT": None,
                "Overtime (min)": None,
                "STATUS": st.column_config.SelectboxColumn(
                    label="STATUS",
                    options=STATUS_OPTIONS,
                    required=False,
                ),
            },
        )
    else:
        df_cards = display_all.copy()
        if card_search:
            query = card_search.lower().strip()
            mask = pd.Series(False, index=df_cards.index)
            for col in ["Patient Name", "Procedure", "DR.", "FIRST", "SECOND", "Third", "STATUS"]:
                if col in df_cards.columns:
                    mask = mask | df_cards[col].astype(str).str.lower().str.contains(query, na=False)
            df_cards = df_cards[mask]
        show_case = "CASE PAPER" in df_cards.columns
        if df_cards.empty:
            st.info("No patients found.")
        else:
            cards_per_row = 3
            for start in range(0, len(df_cards), cards_per_row):
                row_chunk = df_cards.iloc[start:start + cards_per_row]
                cols = st.columns(len(row_chunk), gap="small")
                for col, (row_index, row) in zip(cols, row_chunk.iterrows()):
                    patient = _clean_text(row.get("Patient Name"))
                    doctor = _clean_text(row.get("DR."))
                    procedure = _clean_text(row.get("Procedure"))
                    in_time = row.get("In Time")
                    out_time = row.get("Out Time")
                    status = _clean_text(row.get("STATUS") or row.get("Status") or "WAITING")
                    row_id = _clean_text(row.get("REMINDER_ROW_ID"))
                    staff = [
                        _clean_text(row.get("FIRST")),
                        _clean_text(row.get("SECOND")),
                        _clean_text(row.get("Third")),
                    ]
                    staff = [name for name in staff if name]
                    time_parts = [t for t in [_fmt_time(in_time), _fmt_time(out_time)] if t]
                    time_text = " - ".join(time_parts)
                    status_text = (status or "WAITING").strip().upper()
                    if not status_text:
                        status_text = "WAITING"
                    status_class = _status_class(status_text)
                    staff_html = " &bull; ".join(html.escape(name) for name in staff) if staff else "Unassigned"
                    doctor_icon_svg = '<svg class="info-icon-svg" viewBox="0 0 24 24" width="20" height="20"><path d="M19 8h-1.26c-.19-.73-.48-1.42-.85-2.06l.94-.94a.996.996 0 0 0 0-1.41l-1.41-1.41a.996.996 0 0 0-1.41 0l-.94.94c-.64-.37-1.33-.66-2.06-.85V1c0-.55-.45-1-1-1H9c-.55 0-1 .45-1 1v1.26c-.73.19-1.42.48-2.06.85l-.94-.94a.996.996 0 0 0-1.41 0L2.18 3.58a.996.996 0 0 0 0 1.41l.94.94c-.37.64-.66 1.33-.85 2.06H1c-.55 0-1 .45-1 1v2c0 .55.45 1 1 1h1.26c.19.73.48 1.42.85 2.06l-.94.94a.996.996 0 0 0 0 1.41l1.41 1.41c.39.39 1.02.39 1.41 0l.94-.94c.64.37 1.33.66 2.06.85V23c0 .55.45 1 1 1h2c.55 0 1-.45 1-1v-1.26c.73-.19 1.42-.48 2.06-.85l.94.94c.39.39 1.02.39 1.41 0l1.41-1.41a.996.996 0 0 0 0-1.41l-.94-.94c.37-.64.66-1.33.85-2.06H19c.55 0 1-.45 1-1V9c0-.55-.45-1-1-1zm-8 8c-1.66 0-3-1.34-3-3s1.34-3 3-3 3 1.34 3 3-1.34 3-3 3z" fill="currentColor"/></svg>'
                    staff_icon_svg = '<svg class="info-icon-svg" viewBox="0 0 24 24" width="20" height="20"><path d="M16 11c1.66 0 2.99-1.34 2.99-3S17.66 5 16 5c-1.66 0-3 1.34-3 3s1.34 3 3 3zm-8 0c1.66 0 2.99-1.34 2.99-3S9.66 5 8 5C6.34 5 5 6.34 5 8s1.34 3 3 3zm0 2c-2.33 0-7 1.17-7 3.5V19h14v-2.5c0-2.33-4.67-3.5-7-3.5zm8 0c-.29 0-.62.02-.97.05 1.16.84 1.97 1.97 1.97 3.45V19h6v-2.5c0-2.33-4.67-3.5-7-3.5z" fill="currentColor"/></svg>'
                    doctor_line = (
                        f"<div class='info-row'><span class='info-icon doctor-icon'>{doctor_icon_svg}</span><span class='info-text'>{html.escape(doctor)}</span></div>"
                        if doctor
                        else ""
                    )
                    staff_line = f"<div class='info-row'><span class='info-icon staff-icon'>{staff_icon_svg}</span><span class='info-text'>{staff_html}</span></div>"
                    row_key = row_id if row_id else f"full_{start}_{row_index}"
                    with col:
                        with st.container(border=True):
                            st.markdown("<div class='card-shell-marker'></div>", unsafe_allow_html=True)
                            st.markdown(
                                _normalize_html(
                                    f"""
                                    <div class="card-status-banner {status_class}">
                                        <span class="status-dot"></span>
                                        <span class="status-text">{html.escape(status_text)}</span>
                                    </div>
                                    <div class="card-head">
                                        <div class="card-avatar">{html.escape(_initials(patient))}</div>
                                        <div class="card-title">
                                            <div class="card-name">{html.escape(patient) if patient else "Unknown"}</div>
                                            <div class="card-time">{html.escape(time_text) if time_text else "--"}</div>
                                        </div>
                                    </div>
                                    <div class="card-subdivider"></div>
                                    <div class="card-info">
                                        {doctor_line}
                                        {staff_line}
                                    </div>
                                    """
                                ),
                                unsafe_allow_html=True,
                            )
                            st.markdown("<div class='card-divider'></div>", unsafe_allow_html=True)
                            st.markdown("<div class='card-actions-row'>", unsafe_allow_html=True)
                            if show_case:
                                row_cols = st.columns([1, 1.15, 1.15, 1.15], gap="small")
                                with row_cols[0]:
                                    case_active = _truthy(row.get("CASE PAPER"))
                                    case_checked = st.checkbox("QTRAQ", value=case_active, key=f"full_card_case_{row_key}_{start}")
                                    if case_checked != case_active:
                                        _update_row_case_paper(row_id, patient, in_time, case_checked)
                                with row_cols[1]:
                                    st.markdown("<div class='card-action-marker card-action-done'></div>", unsafe_allow_html=True)
                                    if st.button("✓ Done", key=f"full_card_done_{row_key}_{start}", use_container_width=True, type="primary"):
                                        _update_row_status(row_id, patient, in_time, "DONE")
                                with row_cols[2]:
                                    st.markdown("<div class='card-action-marker card-action-edit'></div>", unsafe_allow_html=True)
                                    st.button("✎ Edit", key=f"full_card_edit_{row_key}_{start}", on_click=_open_full_edit_dialog, args=({"row_key": row_key, "row_id": row_id, "lookup_patient": patient, "lookup_in_time": _fmt_time(in_time), "patient": patient, "in_time": _fmt_time(in_time), "out_time": _fmt_time(out_time), "doctor": doctor, "procedure": procedure, "status": status, "op": _clean_text(row.get("OP")), "staff_first": _clean_text(row.get("FIRST")), "staff_second": _clean_text(row.get("SECOND")), "staff_third": _clean_text(row.get("Third")), "case_paper": _truthy(row.get("CASE PAPER")), "suction": _truthy(row.get("SUCTION")), "cleaning": _truthy(row.get("CLEANING"))},), use_container_width=True, type="secondary")
                                with row_cols[3]:
                                    st.markdown("<div class='card-action-marker card-action-cancel'></div>", unsafe_allow_html=True)
                                    if st.button("✕ Cancel", key=f"full_card_cancel_{row_key}_{start}", use_container_width=True, type="secondary"):
                                        _update_row_status(row_id, patient, in_time, "CANCELLED")
                            else:
                                action_cols = st.columns([1.15, 1.15, 1.15], gap="small")
                                with action_cols[0]:
                                    st.markdown("<div class='card-action-marker card-action-done'></div>", unsafe_allow_html=True)
                                    if st.button("✓ Done", key=f"full_card_done_{row_key}_{start}", use_container_width=True, type="primary"):
                                        _update_row_status(row_id, patient, in_time, "DONE")
                                with action_cols[1]:
                                    st.markdown("<div class='card-action-marker card-action-edit'></div>", unsafe_allow_html=True)
                                    st.button("✎ Edit", key=f"full_card_edit_{row_key}_{start}", on_click=_open_full_edit_dialog, args=({"row_key": row_key, "row_id": row_id, "lookup_patient": patient, "lookup_in_time": _fmt_time(in_time), "patient": patient, "in_time": _fmt_time(in_time), "out_time": _fmt_time(out_time), "doctor": doctor, "procedure": procedure, "status": status, "op": _clean_text(row.get("OP")), "staff_first": _clean_text(row.get("FIRST")), "staff_second": _clean_text(row.get("SECOND")), "staff_third": _clean_text(row.get("Third")), "case_paper": _truthy(row.get("CASE PAPER")), "suction": _truthy(row.get("SUCTION")), "cleaning": _truthy(row.get("CLEANING"))},), use_container_width=True, type="secondary")
                                with action_cols[2]:
                                    st.markdown("<div class='card-action-marker card-action-cancel'></div>", unsafe_allow_html=True)
                                    if st.button("✕ Cancel", key=f"full_card_cancel_{row_key}_{start}", use_container_width=True, type="secondary"):
                                        _update_row_status(row_id, patient, in_time, "CANCELLED")
                            st.markdown("</div>", unsafe_allow_html=True)
                            st.markdown("<div class='card-details-row'>", unsafe_allow_html=True)
                            with st.expander("View Details", expanded=False):
                                st.markdown(f"**Doctor:** {doctor or '--'}")
                                st.markdown(f"**Procedure:** {procedure or '--'}")
                                st.markdown(f"**Staff:** {', '.join(staff) if staff else 'Unassigned'}")
                                st.markdown(f"**Status:** {status}")
                                if show_case:
                                    st.markdown(f"**QTRAQ:** {'Yes' if _truthy(row.get('CASE PAPER')) else 'No'}")
                            st.markdown("</div>", unsafe_allow_html=True)
            if st.session_state.get("full_edit_open"):
                _render_full_edit_dialog()
    # ================ Manual save
    
    # ================ Manual save: process edits only when user clicks save button ================
    if st.session_state.get("manual_save_triggered"):
        # If auto-save is off and we already queued a dataframe, persist it immediately
        pending_df = st.session_state.get("unsaved_df")
        if pending_df is not None:
            pending_msg = st.session_state.get("pending_changes_reason") or "Pending changes saved!"
            if _maybe_save(pending_df, message=pending_msg, force=True):
                st.session_state.unsaved_df = None
                st.session_state.pending_changes = False
                st.session_state.pending_changes_reason = ""
            st.session_state.manual_save_triggered = False
    
        if edited_all is not None:
            editor_key = "full_schedule_editor"
            changed_rows, has_additions = _get_editor_changed_rows(editor_key)
            compare_cols = [
                "Patient Name",
                "In Time",
                "Out Time",
                "Procedure",
                "DR.",
                "FIRST",
                "SECOND",
                "Third",
                "CASE PAPER",
                "OP",
                "SUCTION",
                "CLEANING",
                "STATUS",
            ]
            if has_additions:
                changed_rows = list(edited_all.index)
            else:
                filtered_rows = []
                for row_idx in changed_rows:
                    if row_idx not in edited_all.index or row_idx not in display_all.index:
                        continue
                    if _row_has_changes(edited_all.loc[row_idx], display_all.loc[row_idx], compare_cols):
                        filtered_rows.append(row_idx)
                changed_rows = filtered_rows
            if changed_rows:
                try:
                    # Create a copy of the raw data to update
                    df_updated = df_raw.copy()
    
                    # Track which rows are worth attempting auto-allocation for
                    allocation_candidates: set[int] = set()
                    
                    # Process edited data and convert back to original format
                    for idx in changed_rows:
                        row = edited_all.loc[idx]
                        # Use the preserved original index to map back to df_raw; append when new
                        orig_idx_raw = row.get("_orig_idx", idx)
                        if pd.isna(orig_idx_raw):
                            orig_idx_raw = idx
                        orig_idx = int(orig_idx_raw)
    
                        is_new_row = orig_idx >= len(df_updated)
                        if is_new_row:
                            # Append a blank base row with stable reminder fields
                            base_row = {col: "" for col in df_updated.columns}
                            if "REMINDER_ROW_ID" in base_row:
                                base_row["REMINDER_ROW_ID"] = str(uuid.uuid4())
                            if "REMINDER_SNOOZE_UNTIL" in base_row:
                                base_row["REMINDER_SNOOZE_UNTIL"] = pd.NA
                            if "REMINDER_DISMISSED" in base_row:
                                base_row["REMINDER_DISMISSED"] = False
                            if "STATUS" in base_row and not base_row.get("STATUS"):
                                base_row["STATUS"] = "WAITING"
                            df_updated = pd.concat([df_updated, pd.DataFrame([base_row])], ignore_index=True)
                            orig_idx = len(df_updated) - 1
    
                        try:
                            old_status_norm = ""
                            try:
                                if (not is_new_row) and ("STATUS" in df_raw.columns) and (orig_idx < len(df_raw)):
                                    old_status_norm = str(df_raw.iloc[orig_idx, df_raw.columns.get_loc("STATUS")]).strip().upper()
                            except Exception:
                                old_status_norm = ""
    
                            # Handle Patient ID (optional)
                            if "Patient ID" in row.index and "Patient ID" in df_updated.columns:
                                pid = str(row.get("Patient ID", "")).strip()
                                if pid.lower() in {"nan", "none"}:
                                    pid = ""
                                df_updated.iloc[orig_idx, df_updated.columns.get_loc("Patient ID")] = pid
    
                            # Handle Patient Name
                            patient_name_raw = row.get("Patient Name", "")
                            patient_name = "" if pd.isna(patient_name_raw) else str(patient_name_raw).strip()
                            if patient_name == "":
                                # Clear row if patient name is empty, but preserve stable row id
                                # so users can still delete the blank row from the dropdown.
                                for col in df_updated.columns:
                                    if col == "REMINDER_ROW_ID":
                                        continue
                                    if col == "REMINDER_SNOOZE_UNTIL":
                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc(col)] = pd.NA
                                        continue
                                    if col == "REMINDER_DISMISSED":
                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc(col)] = False
                                        continue
                                    df_updated.iloc[orig_idx, df_updated.columns.get_loc(col)] = ""
                                continue
                            df_updated.iloc[orig_idx, df_updated.columns.get_loc("Patient Name")] = patient_name
                            
                            # Handle In Time - properly convert time object to HH:MM string for Excel
                            if "In Time" in row.index:
                                in_time_val = row["In Time"]
                                t = _coerce_to_time_obj(in_time_val)
                                time_str = f"{t.hour:02d}:{t.minute:02d}" if t is not None else ""
                                df_updated.iloc[orig_idx, df_updated.columns.get_loc("In Time")] = time_str
                            
                            # Handle Out Time - properly convert time object to HH:MM string for Excel
                            if "Out Time" in row.index:
                                out_time_val = row["Out Time"]
                                t = _coerce_to_time_obj(out_time_val)
                                time_str = f"{t.hour:02d}:{t.minute:02d}" if t is not None else ""
                                df_updated.iloc[orig_idx, df_updated.columns.get_loc("Out Time")] = time_str
                            
                            # Handle other columns
                            for col in ["Procedure", "DR.", "FIRST", "SECOND", "Third", "CASE PAPER", "OP", "STATUS"]:
                                if col in row.index and col in df_updated.columns:
                                    val = row[col]
                                    clean_val = str(val).strip() if val and str(val) != "nan" else ""
                                    df_updated.iloc[orig_idx, df_updated.columns.get_loc(col)] = clean_val
    
                            # Time tracking: update timestamps + log on STATUS changes
                            try:
                                if "STATUS" in df_updated.columns:
                                    new_status_norm = str(df_updated.iloc[orig_idx, df_updated.columns.get_loc("STATUS")]).strip().upper()
                                    if new_status_norm and new_status_norm != old_status_norm:
                                        ts = _now_ist_str()
                                        if "STATUS_CHANGED_AT" in df_updated.columns:
                                            df_updated.iloc[orig_idx, df_updated.columns.get_loc("STATUS_CHANGED_AT")] = ts
    
                                        # Actual start/end stamps (only fill first time)
                                        if ("ON GOING" in new_status_norm or "ONGOING" in new_status_norm) and "ACTUAL_START_AT" in df_updated.columns:
                                            cur = str(df_updated.iloc[orig_idx, df_updated.columns.get_loc("ACTUAL_START_AT")]).strip()
                                            if not cur or cur.lower() in {"nan", "none"}:
                                                df_updated.iloc[orig_idx, df_updated.columns.get_loc("ACTUAL_START_AT")] = ts
                                        if ("DONE" in new_status_norm or "COMPLETED" in new_status_norm) and "ACTUAL_END_AT" in df_updated.columns:
                                            cur = str(df_updated.iloc[orig_idx, df_updated.columns.get_loc("ACTUAL_END_AT")]).strip()
                                            if not cur or cur.lower() in {"nan", "none"}:
                                                df_updated.iloc[orig_idx, df_updated.columns.get_loc("ACTUAL_END_AT")] = ts
    
                                        if "STATUS_LOG" in df_updated.columns:
                                            existing_log = ""
                                            try:
                                                if (not is_new_row) and (orig_idx < len(df_raw)) and ("STATUS_LOG" in df_raw.columns):
                                                    existing_log = str(df_raw.iloc[orig_idx, df_raw.columns.get_loc("STATUS_LOG")])
                                            except Exception:
                                                existing_log = ""
                                            df_updated.iloc[orig_idx, df_updated.columns.get_loc("STATUS_LOG")] = _append_status_log(
                                                existing_log,
                                                {
                                                    "at": ts,
                                                    "from": old_status_norm,
                                                    "to": new_status_norm,
                                                },
                                            )
                            except Exception:
                                pass
    
                            # Candidate for allocation if doctor+times exist (helper will decide)
                            allocation_candidates.add(orig_idx)
                            
                            # Handle checkbox columns (SUCTION, CLEANING) - convert boolean to check mark or empty
                            for col in ["SUCTION", "CLEANING"]:
                                if col in row.index and col in df_updated.columns:
                                    val = row[col]
                                    # Store True as "✓" checkmark, False/None as empty string
                                    if pd.isna(val) or val is None or val == False:
                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc(col)] = ""
                                    elif val == True:
                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc(col)] = "✓"
                                    else:
                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc(col)] = ""
                        except Exception as col_error:
                            st.warning(f"Warning updating row {orig_idx}: {str(col_error)}")
                            continue
    
                    # Auto-allocate assistants after applying all row edits
                    if bool(st.session_state.get("auto_assign_assistants", True)):
                        only_empty = bool(st.session_state.get("auto_assign_only_empty", True))
                        for ix in sorted(allocation_candidates):
                            _auto_fill_assistants_for_row(df_updated, ix, only_fill_empty=only_empty)
                    
                    # Write back to storage (manual save always persists)
                    save_ok = _maybe_save(df_updated, message="Schedule updated!", force=True)
                    st.session_state.manual_save_triggered = False
                    if save_ok:
                        st.session_state.unsaved_df = None
                        st.session_state.pending_changes = False
                        st.session_state.pending_changes_reason = ""
                except Exception as e:
                    st.error(f"Error saving: {e}")
                    st.session_state.manual_save_triggered = False
            else:
                # Nothing changed; clear the trigger so it doesn't keep firing on rerun
                st.session_state.manual_save_triggered = False
        else:
            st.session_state.manual_save_triggered = False
    
    if sched_view == "Schedule by OP":
        # ================ Per Chair Tabs ================
        st.markdown("###  Schedule by OP")
        
        unique_ops = sorted(df["OP"].dropna().unique())
        
        if unique_ops:
            tabs = st.tabs([str(op) for op in unique_ops])
            for tab, op in zip(tabs, unique_ops):
                with tab:
                    op_df = df[
                        (df["OP"] == op)
                        & ~df["STATUS"].astype(str).str.upper().str.contains("CANCELLED|DONE|COMPLETED", na=True)
                    ]
                    display_op = op_df[[
                        "Patient ID",
                        "Patient Name",
                        "In Time Obj",
                        "Out Time Obj",
                        "Procedure",
                        "DR.",
                        "OP",
                        "FIRST",
                        "SECOND",
                        "Third",
                        "CASE PAPER",
                        "SUCTION",
                        "CLEANING",
                        "STATUS",
                        "STATUS_CHANGED_AT",
                        "ACTUAL_START_AT",
                        "ACTUAL_END_AT",
                    ]].copy()
                    display_op = display_op.rename(columns={"In Time Obj": "In Time", "Out Time Obj": "Out Time"})
                    # Preserve original index for mapping edits back to df_raw
                    display_op["_orig_idx"] = display_op.index
                    display_op = display_op.reset_index(drop=True)
                    # Ensure time objects are preserved; Streamlit TimeColumn edits best with None for missing
                    display_op["In Time"] = display_op["In Time"].apply(lambda v: v if isinstance(v, time_type) else None)
                    display_op["Out Time"] = display_op["Out Time"].apply(lambda v: v if isinstance(v, time_type) else None)
        
                    # Force correct dtypes for Streamlit compatibility
                    # Text columns
                    for col in ["Patient ID", "Patient Name", "Procedure", "DR.", "FIRST", "SECOND", "Third", "CASE PAPER", "OP", "STATUS"]:
                        if col in display_op.columns:
                            display_op[col] = display_op[col].astype("string").replace('nan', '')
                    # Number column
                    if "Overtime (min)" in display_op.columns:
                        display_op["Overtime (min)"] = pd.to_numeric(display_op["Overtime (min)"], errors="coerce")
                    # Checkbox columns
                    for col in ["SUCTION", "CLEANING"]:
                        if col in display_op.columns:
                            display_op[col] = display_op[col].astype("boolean")
        
                    display_op["Overtime (min)"] = op_df.apply(_compute_overtime_min, axis=1)
        
                    edited_op = st.data_editor(
                        display_op, 
                        width="stretch", 
                        key=f"op_{str(op).replace(' ', '_')}_editor", 
                        hide_index=True,
                        disabled=["STATUS_CHANGED_AT", "ACTUAL_START_AT", "ACTUAL_END_AT", "Overtime (min)"],
                        column_config={
                            "_orig_idx": None,
                            "Patient ID": st.column_config.TextColumn(label="Patient ID", required=False),
                            "In Time": st.column_config.TimeColumn(label="In Time", format="hh:mm A"),
                            "Out Time": st.column_config.TimeColumn(label="Out Time", format="hh:mm A"),
                            "DR.": st.column_config.SelectboxColumn(
                                label="DR.",
                                options=DOCTOR_OPTIONS,
                                required=False
                            ),
                            "OP": st.column_config.SelectboxColumn(
                                label="OP",
                                options=["OP 1", "OP 2", "OP 3", "OP 4"],
                                required=False
                            ),
                            "FIRST": st.column_config.SelectboxColumn(
                                label="FIRST",
                                options=ASSISTANT_OPTIONS,
                                required=False
                            ),
                            "SECOND": st.column_config.SelectboxColumn(
                                label="SECOND",
                                options=ASSISTANT_OPTIONS,
                                required=False
                            ),
                            "Third": st.column_config.SelectboxColumn(
                                label="Third",
                                options=ASSISTANT_OPTIONS,
                                required=False
                            ),
                            "CASE PAPER": st.column_config.SelectboxColumn(
                                label="QTRAQ",
                                options=ASSISTANT_OPTIONS,
                                required=False
                            ),
                            "STATUS_CHANGED_AT": st.column_config.TextColumn(label="Status Changed At"),
                            "ACTUAL_START_AT": st.column_config.TextColumn(label="Actual Start"),
                            "ACTUAL_END_AT": st.column_config.TextColumn(label="Actual End"),
                            "Overtime (min)": st.column_config.NumberColumn(label="Overtime (min)"),
                            "STATUS": st.column_config.SelectboxColumn(
                                label="STATUS",
                                options=STATUS_OPTIONS,
                                required=False
                            )
                        }
                    )
        
                    # Persist edits from OP tabs
                    if edited_op is not None:
                        editor_key = f"op_{str(op).replace(' ', '_')}_editor"
                        changed_rows, has_additions = _get_editor_changed_rows(editor_key)
                        compare_cols = [
                            "Patient ID",
                            "Patient Name",
                            "In Time",
                            "Out Time",
                            "Procedure",
                            "DR.",
                            "OP",
                            "FIRST",
                            "SECOND",
                            "Third",
                            "CASE PAPER",
                            "SUCTION",
                            "CLEANING",
                            "STATUS",
                        ]
                        if has_additions:
                            changed_rows = list(edited_op.index)
                        else:
                            filtered_rows = []
                            for row_idx in changed_rows:
                                if row_idx not in edited_op.index or row_idx not in display_op.index:
                                    continue
                                if _row_has_changes(edited_op.loc[row_idx], display_op.loc[row_idx], compare_cols):
                                    filtered_rows.append(row_idx)
                            changed_rows = filtered_rows
                        if changed_rows:
                            try:
                                df_updated = df_raw.copy()
                                allocation_candidates: set[int] = set()
                                for idx in changed_rows:
                                    row = edited_op.loc[idx]
                                    orig_idx_raw = row.get("_orig_idx")
                                    if pd.isna(orig_idx_raw):
                                        orig_idx_raw = len(df_updated)
                                    orig_idx = int(orig_idx_raw)
        
                                    is_new_row = (orig_idx < 0) or (orig_idx >= len(df_updated))
                                    if is_new_row:
                                        base_row = {col: "" for col in df_updated.columns}
                                        if "REMINDER_ROW_ID" in base_row:
                                            base_row["REMINDER_ROW_ID"] = str(uuid.uuid4())
                                        if "REMINDER_SNOOZE_UNTIL" in base_row:
                                            base_row["REMINDER_SNOOZE_UNTIL"] = pd.NA
                                        if "REMINDER_DISMISSED" in base_row:
                                            base_row["REMINDER_DISMISSED"] = False
                                        if "STATUS" in base_row and not base_row.get("STATUS"):
                                            base_row["STATUS"] = "WAITING"
                                        df_updated = pd.concat([df_updated, pd.DataFrame([base_row])], ignore_index=True)
                                        orig_idx = len(df_updated) - 1
        
                                    old_status_norm = ""
                                    try:
                                        if (not is_new_row) and ("STATUS" in df_raw.columns) and (orig_idx < len(df_raw)):
                                            old_status_norm = str(df_raw.iloc[orig_idx, df_raw.columns.get_loc("STATUS")]).strip().upper()
                                    except Exception:
                                        old_status_norm = ""
        
                                    # Patient ID
                                    patient_id = str(row.get("Patient ID", "")).strip()
                                    if "Patient ID" in df_updated.columns:
                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc("Patient ID")] = patient_id
        
                                    # Patient Name
                                    patient_name_raw = row.get("Patient Name", "")
                                    patient_name = "" if pd.isna(patient_name_raw) else str(patient_name_raw).strip()
                                    if patient_name == "":
                                        for c in df_updated.columns:
                                            if c == "REMINDER_ROW_ID":
                                                continue
                                            if c == "REMINDER_SNOOZE_UNTIL":
                                                df_updated.iloc[orig_idx, df_updated.columns.get_loc(c)] = pd.NA
                                                continue
                                            if c == "REMINDER_DISMISSED":
                                                df_updated.iloc[orig_idx, df_updated.columns.get_loc(c)] = False
                                                continue
                                            df_updated.iloc[orig_idx, df_updated.columns.get_loc(c)] = ""
                                        continue
                                    if "Patient Name" in df_updated.columns:
                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc("Patient Name")] = patient_name
        
                                    # Times -> canonical HH:MM strings
                                    if "In Time" in df_updated.columns:
                                        t = _coerce_to_time_obj(row.get("In Time"))
                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc("In Time")] = (
                                            f"{t.hour:02d}:{t.minute:02d}" if t is not None else ""
                                        )
                                    if "Out Time" in df_updated.columns:
                                        t = _coerce_to_time_obj(row.get("Out Time"))
                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc("Out Time")] = (
                                            f"{t.hour:02d}:{t.minute:02d}" if t is not None else ""
                                        )
        
                                    for c in ["Procedure", "DR.", "OP", "FIRST", "SECOND", "Third", "CASE PAPER", "STATUS"]:
                                        if c in row.index and c in df_updated.columns:
                                            val = row.get(c)
                                            clean_val = str(val).strip() if val and str(val) != "nan" else ""
                                            df_updated.iloc[orig_idx, df_updated.columns.get_loc(c)] = clean_val
        
                                    # Time tracking: update timestamps + log on STATUS changes
                                    try:
                                        if "STATUS" in df_updated.columns:
                                            new_status_norm = str(df_updated.iloc[orig_idx, df_updated.columns.get_loc("STATUS")]).strip().upper()
                                            if new_status_norm and new_status_norm != old_status_norm:
                                                ts = _now_ist_str()
                                                if "STATUS_CHANGED_AT" in df_updated.columns:
                                                    df_updated.iloc[orig_idx, df_updated.columns.get_loc("STATUS_CHANGED_AT")] = ts
        
                                                if ("ON GOING" in new_status_norm or "ONGOING" in new_status_norm) and "ACTUAL_START_AT" in df_updated.columns:
                                                    cur = str(df_updated.iloc[orig_idx, df_updated.columns.get_loc("ACTUAL_START_AT")]).strip()
                                                    if not cur or cur.lower() in {"nan", "none"}:
                                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc("ACTUAL_START_AT")] = ts
                                                if ("DONE" in new_status_norm or "COMPLETED" in new_status_norm) and "ACTUAL_END_AT" in df_updated.columns:
                                                    cur = str(df_updated.iloc[orig_idx, df_updated.columns.get_loc("ACTUAL_END_AT")]).strip()
                                                    if not cur or cur.lower() in {"nan", "none"}:
                                                        df_updated.iloc[orig_idx, df_updated.columns.get_loc("ACTUAL_END_AT")] = ts
        
                                                if "STATUS_LOG" in df_updated.columns:
                                                    existing_log = ""
                                                    try:
                                                        if (not is_new_row) and (orig_idx < len(df_raw)) and ("STATUS_LOG" in df_raw.columns):
                                                            existing_log = str(df_raw.iloc[orig_idx, df_raw.columns.get_loc("STATUS_LOG")])
                                                    except Exception:
                                                        existing_log = ""
                                                    df_updated.iloc[orig_idx, df_updated.columns.get_loc("STATUS_LOG")] = _append_status_log(
                                                        existing_log,
                                                        {"at": ts, "from": old_status_norm, "to": new_status_norm},
                                                    )
                                    except Exception:
                                        pass
        
                                    allocation_candidates.add(orig_idx)
        
                                    for c in ["SUCTION", "CLEANING"]:
                                        if c in row.index and c in df_updated.columns:
                                            val = row.get(c)
                                            if pd.isna(val) or val is None or val is False:
                                                df_updated.iloc[orig_idx, df_updated.columns.get_loc(c)] = ""
                                            elif val is True:
                                                df_updated.iloc[orig_idx, df_updated.columns.get_loc(c)] = "✓"
                                            else:
                                                df_updated.iloc[orig_idx, df_updated.columns.get_loc(c)] = ""
        
                                if bool(st.session_state.get("auto_assign_assistants", True)):
                                    only_empty = bool(st.session_state.get("auto_assign_only_empty", True))
                                    for ix in sorted(allocation_candidates):
                                        _auto_fill_assistants_for_row(df_updated, ix, only_fill_empty=only_empty)
        
                                _maybe_save(df_updated, message=f"Schedule updated for {op}!")
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error saving {op} edits: {e}")
        else:
            st.info("No chair data available.")
        
        
    if category in ("Scheduling", "Doctors") and (doctor_view in (None, "Summary", "Per-Doctor Schedule") or category == "Scheduling"):
        # ================ Doctor Statistics ================
        st.markdown("### 👨‍⚕️ Schedule Summary by Doctor")
        groupby_column = "DR."
        if groupby_column in df.columns and not df[groupby_column].isnull().all():
            try:
                doctor_procedures = df[df["DR."].notna()].groupby("DR.").size().reset_index(name="Total Procedures")
                doctor_procedures = doctor_procedures.reset_index(drop=True)
                if not doctor_procedures.empty:
                    edited_doctor = st.data_editor(doctor_procedures, width="stretch", key="doctor_editor", hide_index=True)
                else:
                    st.info(f"No data available for '{groupby_column}'.")
            except Exception as e:
                st.error(f"Error processing doctor data: {e}")
        else:
            st.info(f"Column '{groupby_column}' not found or contains only empty values.")
        
# ================ ASSISTANT AVAILABILITY DASHBOARD ================
if category == "Assistants" and assist_view == "Availability":
    st.markdown("### 👥 Assistant Availability Dashboard")
    st.markdown("---")
    availability_df = df if 'df' in locals() else df_raw if 'df_raw' in locals() else pd.DataFrame()
    assistants_for_view = get_assistants_list(availability_df)
    if not assistants_for_view:
        assistants_for_view = _get_all_assistants()
    punch_map = _get_today_punch_map()
    # Get current status of all assistants
    assistant_status = get_current_assistant_status(
        availability_df,
        assistants=assistants_for_view,
        punch_map=punch_map,
    )
    
    def _norm_status_value(value: Any) -> str:
        try:
            s = str(value or "").strip().upper()
        except Exception:
            s = ""
        return s if s else "UNKNOWN"
    
    assistant_entries: list[dict] = []
    for assistant in assistants_for_view:
        raw_name = assistant.strip().upper()
        info = dict(assistant_status.get(raw_name, {}))
        if not info:
            info = {"status": "UNKNOWN", "reason": "No schedule"}
        if not info.get("department"):
            info["department"] = get_department_for_assistant(raw_name)
        if not info.get("status"):
            info["status"] = "UNKNOWN"
        assistant_entries.append({
            "name": assistant.title(),
            "raw_name": raw_name,
            "info": info,
        })
    
    assistant_lookup = {entry["raw_name"]: entry for entry in assistant_entries}
    
    # Create tabs for each department
    dept_tabs = st.tabs(["📊 All Assistants", "🦷 PROSTHO Department", "🔬 ENDO Department"])
    
    with dept_tabs[0]:
    
        # Calculate numbers before rendering HTML
        total_count = len(assistant_entries)
        # Normalize status and include alternate status values for busy and blocked
        def is_free(status):
            return status in ["FREE"]
        def is_busy(status):
            return status in ["BUSY", "ON GOING", "ARRIVED"]
        def is_blocked(status):
            return status in ["BLOCKED", "CANCELLED", "SHIFTED"]
    
        free_count = sum(1 for entry in assistant_entries if is_free(_norm_status_value(entry["info"].get("status"))))
        busy_count = sum(1 for entry in assistant_entries if is_busy(_norm_status_value(entry["info"].get("status"))))
        blocked_count = sum(1 for entry in assistant_entries if is_blocked(_norm_status_value(entry["info"].get("status"))))
    
        st.markdown(f"""
        <div style='display: flex; align-items: center; gap: 1.5rem; margin-bottom: 1.2rem;'>
            <div style='background: var(--glass-bg, #f8fafc); border: 1.5px solid var(--glass-border, #3b82f6); border-radius: 1.2rem; padding: 1.2rem 2.2rem; box-shadow: 0 2px 8px rgba(20, 17, 15, 0.04); min-width: 220px;'>
                <div style='font-size: 2.2rem; font-weight: 700; color: var(--text-primary, #1e293b); margin-bottom: 0.2rem;'>Overview</div>
                <div style='font-size: 1.1rem; color: var(--text-secondary, #64748b);'>Current Assistant Status</div>
            </div>
            <div style='display: flex; gap: 1.2rem;'>
                <div style='background: rgba(52, 49, 45, 0.15); border-radius: 0.8rem; padding: 0.8rem 1.4rem; text-align: center;'>
                    <div style='font-size: 1.6rem; font-weight: 600; color: #2563eb;'>{free_count}</div>
                    <div style='font-size: 1rem; color: #2563eb;'>🟢 Free</div>
                </div>
                <div style='background: rgba(126, 127, 131, 0.18); border-radius: 0.8rem; padding: 0.8rem 1.4rem; text-align: center;'>
                    <div style='font-size: 1.6rem; font-weight: 600; color: #64748b;'>{busy_count}</div>
                    <div style='font-size: 1rem; color: #64748b;'>🔴 Busy</div>
                </div>
                <div style='background: rgba(20, 17, 15, 0.12); border-radius: 0.8rem; padding: 0.8rem 1.4rem; text-align: center;'>
                    <div style='font-size: 1.6rem; font-weight: 600; color: #1e293b;'>{blocked_count}</div>
                    <div style='font-size: 1rem; color: #1e293b;'>🚫 Blocked</div>
                </div>
                <div style='background: rgba(217, 197, 178, 0.35); border-radius: 0.8rem; padding: 0.8rem 1.4rem; text-align: center;'>
                    <div style='font-size: 1.6rem; font-weight: 600; color: #2563eb;'>{total_count}</div>
                    <div style='font-size: 1rem; color: #2563eb;'>Total</div>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    
        st.markdown("#### Filter Assistants")
        status_label_map = {
            "FREE": "🟢 Free",
            "BUSY": "🔴 Busy",
            "BLOCKED": "🚫 Blocked",
            "UNKNOWN": "❔ Unknown",
        }
        filter_options = list(status_label_map.keys())
        default_filter = [opt for opt in filter_options if opt != "UNKNOWN"]
        
        # Initialize session state for filter if not set
        if "assistant_status_filter" not in st.session_state:
            st.session_state.assistant_status_filter = default_filter
        
        selected_statuses = st.multiselect(
            "Show statuses",
            options=filter_options,
            default=None,  # Use session state instead
            format_func=lambda x: status_label_map.get(x, x.title()),
            key="assistant_status_filter",
        )
        st.caption("💡 Use the filter to focus on assistants who are free, busy, or currently blocked.")
    
        if selected_statuses:
            filtered_entries = [entry for entry in assistant_entries if _norm_status_value(entry["info"].get("status")) in selected_statuses]
        else:
            filtered_entries = assistant_entries
    
        if filtered_entries:
            st.markdown(f"#### Showing {len(filtered_entries)} Assistant{'s' if len(filtered_entries) != 1 else ''}")
            _render_assistant_cards(filtered_entries)
        else:
            st.info("No assistants match the selected filters.")
    
    with dept_tabs[1]:
        st.markdown("#### PROSTHO Department Assistants")
        prostho_entries: list[dict] = []
        for assistant in get_assistants_for_department("PROSTHO"):
            entry = assistant_lookup.get(assistant.upper())
            if entry is None:
                fallback_info = {
                    "status": "UNKNOWN",
                    "reason": "No schedule",
                    "department": "PROSTHO",
                }
                entry = {"name": assistant.title(), "raw_name": assistant.upper(), "info": fallback_info}
            prostho_entries.append(entry)
    
        prostho_counts: dict[str, int] = {}
        for entry in prostho_entries:
            status_key = _norm_status_value(entry["info"].get("status"))
            prostho_counts[status_key] = prostho_counts.get(status_key, 0) + 1
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("🟢 Free", prostho_counts.get('FREE', 0))
        with col2:
            st.metric("🔴 Busy", prostho_counts.get('BUSY', 0))
        with col3:
            st.metric("🚫 Blocked", prostho_counts.get('BLOCKED', 0))
        
        _render_assistant_cards(prostho_entries)
    
    with dept_tabs[2]:
        st.markdown("#### ENDO Department Assistants")
        endo_entries: list[dict] = []
        for assistant in get_assistants_for_department("ENDO"):
            entry = assistant_lookup.get(assistant.upper())
            if entry is None:
                fallback_info = {
                    "status": "UNKNOWN",
                    "reason": "No schedule",
                    "department": "ENDO",
                }
                entry = {"name": assistant.title(), "raw_name": assistant.upper(), "info": fallback_info}
            endo_entries.append(entry)
    
        endo_counts: dict[str, int] = {}
        for entry in endo_entries:
            status_key = _norm_status_value(entry["info"].get("status"))
            endo_counts[status_key] = endo_counts.get(status_key, 0) + 1
        
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("🟢 Free", endo_counts.get('FREE', 0))
        with col2:
            st.metric("🔴 Busy", endo_counts.get('BUSY', 0))
        with col3:
            st.metric("🚫 Blocked", endo_counts.get('BLOCKED', 0))
        
        _render_assistant_cards(endo_entries)
if category == "Assistants" and assist_view == "Auto Allocation":
    # ================ AUTOMATIC ASSISTANT ALLOCATION ================
    with st.expander("🔄 Automatic Assistant Allocation", expanded=False):
        st.caption("Automatically assign assistants based on department, doctor, and availability")
        
        col_doc, col_in, col_out = st.columns(3)
        
        with col_doc:
            alloc_doctor = st.selectbox(
                "Select Doctor",
                options=[""] + _get_all_doctors(),
                key="alloc_doctor_select"
            )
        
        with col_in:
            alloc_in_time = st.time_input("Appointment Start", value=time_type(9, 0), key="alloc_in_time")
        
        with col_out:
            alloc_out_time = st.time_input("Appointment End", value=time_type(10, 0), key="alloc_out_time")
        
        if alloc_doctor:
            dept = get_department_for_doctor(alloc_doctor)
            st.info(f"Department: **{dept}**")
            
            # Get available assistants
            free_now_set, free_status_map = _get_dashboard_free_set(df, _get_all_assistants())
            available = get_available_assistants(
                dept,
                alloc_in_time,
                alloc_out_time,
                df,
                free_now_set=free_now_set,
                free_status_map=free_status_map,
            )
            
            st.markdown("**Assistant Availability:**")
            for a in available:
                if a["available"]:
                    st.success(f"✅ {a['name']} - Available")
                else:
                    st.error(f"❌ {a['name']} - {a['reason']}")
            
            # Auto-allocate button
            if st.button("🎯 Get Recommended Allocation", key="auto_alloc_btn"):
                allocation = auto_allocate_assistants(alloc_doctor, alloc_in_time, alloc_out_time, df)
                
                if any(allocation.values()):
                    st.success("**Recommended Allocation:**")
                    if allocation["FIRST"]:
                        st.write(f"• **FIRST**: {allocation['FIRST']}")
                    if allocation["SECOND"]:
                        st.write(f"• **SECOND**: {allocation['SECOND']}")
                    if allocation["Third"]:
                        st.write(f"• **Third**: {allocation['Third']}")
                else:
                    st.warning("No available assistants found for this time slot in the department.")
        else:
            st.caption("Select a doctor to see department-specific assistant availability")
    
if category == "Assistants" and assist_view == "Workload":
    # ================ ASSISTANT WORKLOAD SUMMARY ================
    st.markdown("### 📊 Assistant Workload Summary")
    
    # Count appointments per assistant
    assistant_workload = {}
    for assistant in _get_all_assistants():
        schedule = get_assistant_schedule(assistant.upper(), df)
        assistant_workload[assistant] = len(schedule)
    
    # Create workload dataframe
    workload_data = []
    for assistant, count in sorted(assistant_workload.items(), key=lambda x: x[1], reverse=True):
        dept = get_department_for_assistant(assistant.upper())
        workload_data.append({
            "Assistant": assistant,
            "Department": dept,
            "Appointments Today": count
        })
    
    if workload_data:
        st.dataframe(pd.DataFrame(workload_data), use_container_width=True, hide_index=True)
    
if category == "Assistants" and assist_view == "Attendance":
    # ================ ASSISTANTS ATTENDANCE (EXPERIMENTAL) ================
    if USE_SUPABASE:
        st.info("Attendance editor (sheet-based) is disabled in Supabase mode. Use the sidebar Punch widget instead.")
        if supabase_client is None:
            st.warning("Supabase is not configured. Configure Supabase to view attendance reports.")
        else:
            with st.expander("Monthly Attendance Report", expanded=True):
                month_base = datetime.now(IST).date().replace(day=1)
                month_options = []
                for i in range(0, 12):
                    idx = (month_base.year * 12 + (month_base.month - 1)) - i
                    year = idx // 12
                    month = idx % 12 + 1
                    month_options.append(datetime(year, month, 1).date())
                selected_month = st.selectbox(
                    "Report month",
                    options=month_options,
                    index=0,
                    format_func=lambda d: d.strftime("%Y-%m"),
                    key="attendance_report_month",
                )
                next_idx = (selected_month.year * 12 + (selected_month.month - 1)) + 1
                next_year = next_idx // 12
                next_month = next_idx % 12 + 1
                start_date = selected_month.isoformat()
                end_date = (datetime(next_year, next_month, 1).date() - timedelta(days=1)).isoformat()
                st.caption(f"Range: {start_date} to {end_date}")
                records = _load_attendance_range_supabase(supabase_client, start_date, end_date)
                if not records:
                    st.info("No attendance records for selected month.")
                else:
                    df_att = pd.DataFrame(records)
                    for col in ["date", "assistant", "punch_in", "punch_out"]:
                        if col not in df_att.columns:
                            df_att[col] = ""
                    df_att["date"] = df_att["date"].astype(str)
                    df_att["assistant"] = df_att["assistant"].astype(str).str.strip().str.upper()
                    df_att["punch_in"] = df_att["punch_in"].astype(str).str.strip()
                    df_att["punch_out"] = df_att["punch_out"].astype(str).str.strip()
                    df_att["STATUS"] = df_att.apply(
                        lambda row: _attendance_status(row["punch_in"], row["punch_out"]),
                        axis=1,
                    )
                    df_att["WORKED MINS"] = df_att.apply(
                        lambda row: _calc_worked_minutes(row["punch_in"], row["punch_out"]),
                        axis=1,
                    )
                    df_att["WORKED HH:MM"] = df_att["WORKED MINS"].apply(mins_to_hhmm)
                    assistant_options = ["All"] + sorted(
                        [a for a in df_att["assistant"].unique().tolist() if a]
                    )
                    selected_assistant = st.selectbox(
                        "Assistant filter",
                        options=assistant_options,
                        index=0,
                        key="attendance_report_assistant",
                    )
                    if selected_assistant != "All":
                        df_att = df_att[df_att["assistant"] == selected_assistant]
                    if df_att.empty:
                        st.info("No attendance records for this assistant in the selected month.")
                    else:
                        df_summary = df_att.copy()
                        df_summary["WORKED MINS FILLED"] = df_summary["WORKED MINS"].fillna(0).astype(int)
                        summary = (
                            df_summary.groupby("assistant", dropna=False)
                            .agg(
                                Days=("date", "nunique"),
                                Completed=("STATUS", lambda s: (s == "COMPLETE").sum()),
                                In_Progress=("STATUS", lambda s: (s == "IN PROGRESS").sum()),
                                Worked_Minutes=("WORKED MINS FILLED", "sum"),
                            )
                            .reset_index()
                        )
                        summary["Worked HH:MM"] = summary["Worked_Minutes"].apply(mins_to_hhmm)
                        summary_display = summary.rename(columns={"assistant": "ASSISTANT"})
                        summary_display = summary_display.sort_values("ASSISTANT")
                        st.markdown("**Summary**")
                        st.dataframe(summary_display, use_container_width=True, hide_index=True)
                        details = df_att[
                            ["date", "assistant", "punch_in", "punch_out", "STATUS", "WORKED MINS", "WORKED HH:MM"]
                        ].copy()
                        details = details.rename(
                            columns={
                                "date": "DATE",
                                "assistant": "ASSISTANT",
                                "punch_in": "PUNCH IN",
                                "punch_out": "PUNCH OUT",
                            }
                        )
                        details = details.sort_values(["DATE", "ASSISTANT"])
                        st.markdown("**Details**")
                        st.dataframe(details, use_container_width=True, hide_index=True)
                        csv = details.to_csv(index=False)
                        st.download_button(
                            "Download CSV",
                            data=csv,
                            file_name=f"attendance_{selected_month.strftime('%Y_%m')}.csv",
                            mime="text/csv",
                        )
    else:
        with st.expander("🕒 Assistants Attendance", expanded=False):
            try:
                render_assistant_attendance_tab(df if 'df' in locals() else pd.DataFrame(), file_path)
            except Exception as e:
                st.error(f"Unable to load attendance editor: {e}")
# ================ ADMIN / SETTINGS ================
if category == "Admin/Settings":
    st.markdown("### 🔧 Admin / Settings")
    if admin_view == "Duties Manager":
        # Duties are now stored in Excel (not Supabase)
        assistants_for_admin = extract_assistants(df if 'df' in locals() else df_raw if 'df_raw' in locals() else pd.DataFrame())
        render_duties_master_admin(None)  # Excel-based, supabase param not used
        st.divider()
        render_duty_assignment_admin(None, assistants_for_admin)  # Excel-based, supabase param not used
    else:
        st.write(f"Using Supabase: {USE_SUPABASE}")
        st.write(f"Excel path: {file_path}")
