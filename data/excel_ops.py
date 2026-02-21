# data/excel_ops.py
"""Generic Excel read/write helpers preserving all other sheets."""

from __future__ import annotations
from typing import Optional
import os
import zipfile
import pandas as pd
import openpyxl

from config.settings import EXCEL_PATH


def load_sheet(
    sheet_name: str,
    expected_columns: Optional[list] = None,
    path: Optional[str] = None,
) -> pd.DataFrame:
    """Load a sheet from the workbook. Returns empty DataFrame if missing."""
    fpath = str(path or EXCEL_PATH)
    try:
        df = pd.read_excel(fpath, sheet_name=sheet_name, engine="openpyxl")
        if expected_columns:
            for col in expected_columns:
                if col not in df.columns:
                    df[col] = ""
            if df.empty:
                return pd.DataFrame(columns=expected_columns)
        return df
    except Exception:
        return pd.DataFrame(columns=expected_columns) if expected_columns else pd.DataFrame()


def save_sheet(df: pd.DataFrame, sheet_name: str, path: Optional[str] = None) -> bool:
    """Save DataFrame to a sheet, preserving all other sheets."""
    fpath = str(path or EXCEL_PATH)
    try:
        if not os.path.exists(fpath):
            wb = openpyxl.Workbook()
            if wb.active:
                wb.remove(wb.active)
        else:
            try:
                wb = openpyxl.load_workbook(fpath)
            except (zipfile.BadZipFile, KeyError, Exception) as e:
                print(f"CRITICAL: Cannot load existing Excel file {fpath}: {e}")
                return False

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(sheet_name)

        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        # Write data
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Ensure at least one visible sheet
        if not any(ws_check.sheet_state == "visible" for ws_check in wb.worksheets):
            if wb.sheetnames:
                wb[wb.sheetnames[0]].sheet_state = "visible"

        wb.save(fpath)
        wb.close()
        return True
    except Exception as e:
        print(f"Error saving sheet '{sheet_name}': {e}")
        return False
